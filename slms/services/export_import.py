"""Export/Import service for CSV and XLSX formats."""
from __future__ import annotations

import csv
import io
from datetime import datetime, date
from typing import Any, Type, BinaryIO

from flask import current_app
from sqlalchemy import inspect
from sqlalchemy.orm import Query

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from slms.extensions import db
from slms.models.models import (
    Organization, User, League, Season, Team, Player, Venue, Game,
    Coach, CoachAssignment, Referee, GameOfficials, Sponsor, Transaction,
    Registration, MediaAsset, Article, ContentAsset
)


class ExportImportService:
    """Service for exporting and importing domain objects."""

    # Define exportable fields for each model
    EXPORTABLE_FIELDS = {
        'Team': ['id', 'name', 'season_id', 'coach_name', 'wins', 'losses', 'ties', 'created_at'],
        'Player': ['id', 'first_name', 'last_name', 'email', 'jersey_number', 'birthdate', 'team_id', 'created_at'],
        'Coach': ['id', 'first_name', 'last_name', 'email', 'phone', 'certification_level', 'years_experience', 'is_active'],
        'Referee': ['id', 'first_name', 'last_name', 'email', 'phone', 'certification_level', 'license_number', 'is_active'],
        'Venue': ['id', 'name', 'address', 'city', 'timezone', 'court_label', 'open_time', 'close_time'],
        'Game': ['id', 'season_id', 'home_team_id', 'away_team_id', 'venue_id', 'start_time', 'status', 'home_score', 'away_score'],
        'League': ['id', 'name', 'sport', 'description', 'status', 'created_at'],
        'Season': ['id', 'league_id', 'name', 'start_date', 'end_date', 'status', 'registration_open', 'registration_fee'],
        'Sponsor': ['id', 'name', 'contact_name', 'contact_email', 'contact_phone', 'tier', 'sponsorship_amount', 'is_active'],
        'Transaction': ['id', 'transaction_date', 'category', 'description', 'amount', 'payment_method', 'reference_number'],
        'Registration': ['id', 'first_name', 'last_name', 'email', 'phone', 'birthdate', 'team_id', 'payment_status', 'registration_date'],
        'MediaAsset': ['id', 'title', 'description', 'media_type', 'category', 'file_url', 'thumbnail_url', 'created_at'],
        'Article': ['id', 'title', 'slug', 'excerpt', 'status', 'published_at', 'view_count', 'is_featured'],
    }

    # Define importable fields (fields that can be set on create/update)
    IMPORTABLE_FIELDS = {
        'Team': ['name', 'season_id', 'coach_name'],
        'Player': ['first_name', 'last_name', 'email', 'jersey_number', 'birthdate', 'team_id'],
        'Coach': ['first_name', 'last_name', 'email', 'phone', 'certification_level', 'years_experience', 'is_active'],
        'Referee': ['first_name', 'last_name', 'email', 'phone', 'certification_level', 'license_number', 'is_active'],
        'Venue': ['name', 'address', 'city', 'timezone', 'court_label', 'open_time', 'close_time'],
        'Game': ['season_id', 'home_team_id', 'away_team_id', 'venue_id', 'start_time', 'status'],
        'League': ['name', 'sport', 'description', 'status'],
        'Season': ['league_id', 'name', 'start_date', 'end_date', 'status', 'registration_open', 'registration_fee'],
        'Sponsor': ['name', 'contact_name', 'contact_email', 'contact_phone', 'tier', 'sponsorship_amount', 'is_active'],
        'Transaction': ['transaction_date', 'category', 'description', 'amount', 'payment_method', 'reference_number'],
        'Registration': ['first_name', 'last_name', 'email', 'phone', 'birthdate', 'team_id', 'payment_status'],
    }

    MODEL_MAP = {
        'Team': Team,
        'Player': Player,
        'Coach': Coach,
        'Referee': Referee,
        'Venue': Venue,
        'Game': Game,
        'League': League,
        'Season': Season,
        'Sponsor': Sponsor,
        'Transaction': Transaction,
        'Registration': Registration,
        'MediaAsset': MediaAsset,
        'Article': Article,
    }

    @staticmethod
    def export_to_csv(query: Query, model_name: str) -> str:
        """
        Export query results to CSV format.

        Args:
            query: SQLAlchemy query
            model_name: Name of the model being exported

        Returns:
            CSV string
        """
        fields = ExportImportService.EXPORTABLE_FIELDS.get(model_name, [])
        if not fields:
            raise ValueError(f"No exportable fields defined for {model_name}")

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()

        for obj in query.all():
            row = {}
            for field in fields:
                value = getattr(obj, field, None)
                # Convert special types to strings
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                elif hasattr(value, 'value'):  # Enum
                    value = value.value
                elif value is None:
                    value = ''
                row[field] = value
            writer.writerow(row)

        return output.getvalue()

    @staticmethod
    def export_to_xlsx(query: Query, model_name: str) -> bytes:
        """
        Export query results to XLSX format.

        Args:
            query: SQLAlchemy query
            model_name: Name of the model being exported

        Returns:
            XLSX bytes
        """
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX export")

        fields = ExportImportService.EXPORTABLE_FIELDS.get(model_name, [])
        if not fields:
            raise ValueError(f"No exportable fields defined for {model_name}")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = model_name

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Write headers
        for col_idx, field in enumerate(fields, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, obj in enumerate(query.all(), start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = getattr(obj, field, None)
                # Convert special types
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                elif hasattr(value, 'value'):  # Enum
                    value = value.value
                elif value is None:
                    value = ''
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def import_from_csv(
        file_content: str,
        model_name: str,
        org_id: str,
        update_existing: bool = False
    ) -> tuple[int, int, list[str]]:
        """
        Import data from CSV.

        Args:
            file_content: CSV file content as string
            model_name: Name of the model to import
            org_id: Organization ID
            update_existing: Whether to update existing records

        Returns:
            Tuple of (created_count, updated_count, errors)
        """
        model_class = ExportImportService.MODEL_MAP.get(model_name)
        if not model_class:
            raise ValueError(f"Unknown model: {model_name}")

        importable_fields = ExportImportService.IMPORTABLE_FIELDS.get(model_name, [])
        if not importable_fields:
            raise ValueError(f"No importable fields defined for {model_name}")

        reader = csv.DictReader(io.StringIO(file_content))
        created_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                # Prepare data
                data = {'org_id': org_id}
                for field in importable_fields:
                    if field in row and row[field]:
                        value = row[field].strip()
                        if value:
                            # Type conversion
                            column = getattr(model_class, field, None)
                            if column is not None:
                                column_type = column.type.python_type
                                if column_type == datetime:
                                    value = datetime.fromisoformat(value)
                                elif column_type == date:
                                    value = date.fromisoformat(value)
                                elif column_type == int:
                                    value = int(value)
                                elif column_type == float:
                                    value = float(value)
                                elif column_type == bool:
                                    value = value.lower() in ('true', 'yes', '1')
                            data[field] = value

                # Check if record exists (by ID or unique field)
                existing = None
                if 'id' in row and row['id']:
                    existing = db.session.get(model_class, row['id'])
                elif update_existing and 'email' in data:
                    # Try to find by email
                    existing = db.session.query(model_class).filter_by(
                        org_id=org_id,
                        email=data['email']
                    ).first()

                if existing and update_existing:
                    # Update existing record
                    for key, value in data.items():
                        if key != 'id' and key != 'org_id':
                            setattr(existing, key, value)
                    updated_count += 1
                elif not existing:
                    # Create new record
                    new_obj = model_class(**data)
                    db.session.add(new_obj)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                current_app.logger.error(f"Import error on row {row_num}: {e}")

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors.append(f"Database commit failed: {str(e)}")

        return created_count, updated_count, errors

    @staticmethod
    def import_from_xlsx(
        file: BinaryIO,
        model_name: str,
        org_id: str,
        update_existing: bool = False
    ) -> tuple[int, int, list[str]]:
        """
        Import data from XLSX.

        Args:
            file: File object
            model_name: Name of the model to import
            org_id: Organization ID
            update_existing: Whether to update existing records

        Returns:
            Tuple of (created_count, updated_count, errors)
        """
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX import")

        model_class = ExportImportService.MODEL_MAP.get(model_name)
        if not model_class:
            raise ValueError(f"Unknown model: {model_name}")

        importable_fields = ExportImportService.IMPORTABLE_FIELDS.get(model_name, [])
        if not importable_fields:
            raise ValueError(f"No importable fields defined for {model_name}")

        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        created_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dict from row
                row_dict = dict(zip(headers, row))

                # Prepare data
                data = {'org_id': org_id}
                for field in importable_fields:
                    if field in row_dict and row_dict[field] is not None:
                        value = row_dict[field]
                        if isinstance(value, str):
                            value = value.strip()
                        if value:
                            # Type conversion
                            column = getattr(model_class, field, None)
                            if column is not None:
                                column_type = column.type.python_type
                                if column_type == datetime and isinstance(value, str):
                                    value = datetime.fromisoformat(value)
                                elif column_type == date and isinstance(value, str):
                                    value = date.fromisoformat(value)
                                elif column_type == int:
                                    value = int(value)
                                elif column_type == float:
                                    value = float(value)
                                elif column_type == bool:
                                    value = str(value).lower() in ('true', 'yes', '1')
                            data[field] = value

                # Check if record exists
                existing = None
                if 'id' in row_dict and row_dict['id']:
                    existing = db.session.get(model_class, row_dict['id'])
                elif update_existing and 'email' in data:
                    existing = db.session.query(model_class).filter_by(
                        org_id=org_id,
                        email=data['email']
                    ).first()

                if existing and update_existing:
                    # Update existing record
                    for key, value in data.items():
                        if key != 'id' and key != 'org_id':
                            setattr(existing, key, value)
                    updated_count += 1
                elif not existing:
                    # Create new record
                    new_obj = model_class(**data)
                    db.session.add(new_obj)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                current_app.logger.error(f"Import error on row {row_num}: {e}")

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors.append(f"Database commit failed: {str(e)}")

        return created_count, updated_count, errors

    @staticmethod
    def get_export_template(model_name: str, format: str = 'csv') -> bytes | str:
        """
        Generate an empty template for import.

        Args:
            model_name: Name of the model
            format: 'csv' or 'xlsx'

        Returns:
            Template file content
        """
        fields = ExportImportService.IMPORTABLE_FIELDS.get(model_name, [])
        if not fields:
            raise ValueError(f"No importable fields defined for {model_name}")

        if format == 'csv':
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fields)
            writer.writeheader()
            return output.getvalue()

        elif format == 'xlsx':
            if not XLSX_AVAILABLE:
                raise ImportError("openpyxl is required for XLSX templates")

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = model_name

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Write headers
            for col_idx, field in enumerate(fields, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=field)
                cell.fill = header_fill
                cell.font = header_font

            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.read()

        else:
            raise ValueError(f"Unsupported format: {format}")
