from flask import Blueprint, render_template, request, redirect, url_for, flash, session, g, jsonify
from datetime import datetime, date, timedelta, timezone
from decimal import Decimal, InvalidOperation
import copy
import json
import random
import re
from urllib.parse import urlparse

from flask_login import current_user, login_required

from slms.services.db import get_db
from slms.extensions import db
from slms.blueprints.common.tenant import tenant_required
from slms.services.sport_config import get_sport_config, get_all_sports
from slms.services.site import (
    _load_site_settings,
    invalidate_site_settings_cache,
    DEFAULT_FEATURE_FLAGS,
    DEFAULT_SOCIAL_LINKS,
    DEFAULT_THEME_CONFIG,
    get_site_theme_preview,
    save_site_theme_preview,
    publish_site_theme,
    discard_site_theme_preview,
    list_site_theme_versions,
    get_site_theme_version,
    apply_site_payload,
    _apply_payload_to_settings,
    default_cta_slots,
    merge_theme_cta_slots,
    normalize_cta_slots,
    ensure_theme_footer,
    normalize_footer_config,
    FOOTER_ALLOWED_STYLES,
    FOOTER_MAX_COLUMNS,
    FOOTER_MAX_LINKS_PER_COLUMN,
)
from slms.services.media_library import (
    create_media_asset,
    update_media_asset,
    delete_media_asset,
    serialize_media_collection,
)

from slms.models import MediaAsset, UserRole, User


def _default_homepage_config():
    return {
        'title': '',
        'subtitle': '',
        'background_url': '',
        'cta_text': '',
        'cta_url': '',
        'highlights': [],
    }


AVERAGE_MATCH_REVENUE_USD = 7500  # fallback estimate when transaction data is unavailable

ICON_PICKER_CHOICES = [
    {"value": "ph ph-football", "label": "Football"},
    {"value": "ph ph-soccer-ball", "label": "Soccer"},
    {"value": "ph ph-basketball", "label": "Basketball"},
    {"value": "ph ph-baseball", "label": "Baseball"},
    {"value": "ph ph-tennis-ball", "label": "Tennis"},
    {"value": "ph ph-volleyball", "label": "Volleyball"},
    {"value": "ph ph-golf", "label": "Golf"},
    {"value": "ph ph-cricket", "label": "Cricket"},
    {"value": "ph ph-hockey", "label": "Hockey"},
    {"value": "ph ph-boxing-glove", "label": "Boxing"},
    {"value": "ph ph-bowling-ball", "label": "Bowling"},
    {"value": "ph ph-court-basketball", "label": "Court"},
    {"value": "ph ph-tennis-racket", "label": "Racket"},
    {"value": "ph ph-trophy", "label": "Trophy"},
    {"value": "ph ph-medal", "label": "Medal"},
    {"value": "ph ph-flag", "label": "Flag"},
    {"value": "ph ph-crown", "label": "Crown"},
    {"value": "ph ph-shield", "label": "Shield"},
    {"value": "ph ph-speedometer", "label": "Dashboard"},
    {"value": "ph ph-clipboard", "label": "Clipboard"},
    {"value": "ph ph-calendar-dots", "label": "Schedule"},
    {"value": "ph ph-clock", "label": "Time"},
    {"value": "ph ph-chart-line", "label": "Stats"},
    {"value": "ph ph-chart-bar", "label": "Charts"},
    {"value": "ph ph-ranking", "label": "Rankings"},
    {"value": "ph ph-house", "label": "Home"},
    {"value": "ph ph-squares-four", "label": "Dashboard"},
    {"value": "ph ph-dots-nine", "label": "Grid"},
    {"value": "ph ph-compass", "label": "Navigate"},
    {"value": "ph ph-map-pin", "label": "Location"},
    {"value": "ph ph-info", "label": "Info"},
    {"value": "ph ph-question", "label": "Help"},
    {"value": "ph ph-gear-six", "label": "Settings"},
    {"value": "ph ph-bell", "label": "Alerts"},
    {"value": "ph ph-envelope", "label": "Messages"},
    {"value": "ph ph-user", "label": "User"},
    {"value": "ph ph-users-three", "label": "Team"},
    {"value": "ph ph-identification-badge", "label": "Badge"},
    {"value": "ph ph-handshake", "label": "Partners"},
    {"value": "ph ph-user-circle", "label": "Profile"},
    {"value": "ph ph-address-book", "label": "Contacts"},
    {"value": "ph ph-building-office", "label": "Office"},
    {"value": "ph ph-bank", "label": "Finance"},
    {"value": "ph ph-coins", "label": "Fees"},
    {"value": "ph ph-credit-card", "label": "Payment"},
    {"value": "ph ph-receipt", "label": "Receipt"},
    {"value": "ph ph-invoice", "label": "Invoice"},
    {"value": "ph ph-facebook-logo", "label": "Facebook"},
    {"value": "ph ph-instagram-logo", "label": "Instagram"},
    {"value": "ph ph-x-logo", "label": "X/Twitter"},
    {"value": "ph ph-youtube-logo", "label": "YouTube"},
    {"value": "ph ph-tiktok-logo", "label": "TikTok"},
    {"value": "ph ph-phone", "label": "Phone"},
    {"value": "ph ph-star", "label": "Star"},
    {"value": "ph ph-lightning", "label": "Lightning"},
    {"value": "ph ph-fire", "label": "Fire"},
    {"value": "ph ph-heart", "label": "Heart"},
    {"value": "ph ph-thumbs-up", "label": "Like"},
    {"value": "ph ph-download", "label": "Download"},
    {"value": "ph ph-upload", "label": "Upload"},
    {"value": "ph ph-link", "label": "Link"},
    {"value": "ph ph-globe", "label": "Web"},
    {"value": "ph ph-archive", "label": "Archive"},
]

admin_bp = Blueprint('admin', __name__)

from slms.auth import admin_required

def get_existing_data(table_name):
    db = get_db()
    cur = db.cursor()
    cur.execute(f'SELECT * FROM {table_name}')
    data = cur.fetchall()
    cur.close()
    return data


def _calc_metric(current_value, previous_value):
    current = current_value or 0
    previous = previous_value or 0
    delta = current - previous
    delta_pct = (delta / previous * 100) if previous else None
    return {
        'current': current,
        'previous': previous,
        'delta': delta,
        'delta_pct': delta_pct,
    }


def _execute_scalar(cur, query, params=None):
    cur.execute(query, params or ())
    row = cur.fetchone()
    return row[0] if row and row[0] is not None else 0



def _parse_amount_to_cents(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int((Decimal(str(value)) * 100).quantize(Decimal('1')))
    value = value.strip()
    if not value:
        return None
    try:
        quant = Decimal(value.replace(',', ''))
    except InvalidOperation:
        raise ValueError('Invalid currency amount')
    cents = int((quant * 100).quantize(Decimal('1')))
    return cents


def _format_cents_to_decimal(cents):
    if cents is None:
        return None
    return (Decimal(cents) / Decimal('100')).quantize(Decimal('0.01'))


def _ensure_finance_hub_tables(db_wrapper):
    """Create tables for comprehensive finance hub"""
    cur = db_wrapper.cursor()
    try:
        # Revenue tracking table for consolidated financial data
        cur.execute("""
            CREATE TABLE IF NOT EXISTS league_revenue (
                revenue_id SERIAL PRIMARY KEY,
                league_id INTEGER NOT NULL REFERENCES leagues(league_id) ON DELETE CASCADE,
                revenue_type VARCHAR(50) NOT NULL, -- 'fees', 'sponsorship', 'in_person', 'square', 'other'
                amount_cents INTEGER NOT NULL DEFAULT 0,
                currency VARCHAR(3) NOT NULL DEFAULT 'USD',
                transaction_date DATE NOT NULL DEFAULT CURRENT_DATE,
                description TEXT,
                reference_id INTEGER, -- References to related records (fee_id, sponsor_id, etc)
                reference_type VARCHAR(50), -- 'installment', 'sponsorship_deal', 'manual_entry'
                payment_method VARCHAR(50), -- 'cash', 'check', 'card', 'bank_transfer', 'square', 'online'
                status VARCHAR(20) NOT NULL DEFAULT 'confirmed', -- 'pending', 'confirmed', 'cancelled'
                created_by INTEGER REFERENCES users(user_id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                metadata JSONB -- For storing additional payment gateway data
            )
        """)

        # In-person payment transactions
        cur.execute("""
            CREATE TABLE IF NOT EXISTS in_person_payments (
                payment_id SERIAL PRIMARY KEY,
                league_id INTEGER NOT NULL REFERENCES leagues(league_id) ON DELETE CASCADE,
                amount_cents INTEGER NOT NULL,
                currency VARCHAR(3) NOT NULL DEFAULT 'USD',
                payment_method VARCHAR(50) NOT NULL, -- 'cash', 'check', 'card'
                payment_date DATE NOT NULL DEFAULT CURRENT_DATE,
                payer_name VARCHAR(255),
                payer_contact VARCHAR(255), -- Phone or email
                description TEXT NOT NULL,
                check_number VARCHAR(50), -- For check payments
                card_last_four VARCHAR(4), -- For card payments (security)
                receipt_number VARCHAR(100),
                notes TEXT,
                created_by INTEGER REFERENCES users(user_id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Enhanced financial reports/budget tracking
        cur.execute("""
            CREATE TABLE IF NOT EXISTS league_budgets (
                budget_id SERIAL PRIMARY KEY,
                league_id INTEGER NOT NULL REFERENCES leagues(league_id) ON DELETE CASCADE,
                budget_year INTEGER NOT NULL,
                category VARCHAR(100) NOT NULL, -- 'registration_fees', 'sponsorships', 'concessions', etc
                budgeted_amount_cents INTEGER NOT NULL DEFAULT 0,
                actual_amount_cents INTEGER NOT NULL DEFAULT 0,
                currency VARCHAR(3) NOT NULL DEFAULT 'USD',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(league_id, budget_year, category)
            )
        """)

        # Square payment integration tracking (ready for future)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS square_payments (
                square_payment_id SERIAL PRIMARY KEY,
                league_id INTEGER NOT NULL REFERENCES leagues(league_id) ON DELETE CASCADE,
                square_payment_id_external VARCHAR(255) UNIQUE NOT NULL,
                amount_cents INTEGER NOT NULL,
                currency VARCHAR(3) NOT NULL DEFAULT 'USD',
                status VARCHAR(50) NOT NULL, -- 'completed', 'pending', 'failed', 'cancelled'
                payment_date TIMESTAMP,
                description TEXT,
                customer_email VARCHAR(255),
                customer_name VARCHAR(255),
                receipt_url TEXT,
                reference_id INTEGER, -- Link to league fees, etc
                reference_type VARCHAR(50),
                square_order_id VARCHAR(255),
                metadata JSONB,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Create indexes for better performance
        cur.execute("CREATE INDEX IF NOT EXISTS idx_league_revenue_league_id ON league_revenue(league_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_league_revenue_date ON league_revenue(transaction_date)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_league_revenue_type ON league_revenue(revenue_type)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_in_person_payments_league_id ON in_person_payments(league_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_in_person_payments_date ON in_person_payments(payment_date)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_league_budgets_league_year ON league_budgets(league_id, budget_year)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_square_payments_league_id ON square_payments(league_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_square_payments_status ON square_payments(status)")

        db_wrapper.commit()
    except Exception:
        db_wrapper.rollback()
        raise
    finally:
        cur.close()


def _ensure_scorer_metrics_table(db_wrapper):
    """Ensure supporting table for extended scorer analytics exists."""
    try:
        from sqlalchemy import text
        from slms.extensions import db

        engine = db.engine
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS scorer_metrics (
                    scorer_id INTEGER PRIMARY KEY REFERENCES scorers(scorer_id) ON DELETE CASCADE,
                    games_played INTEGER DEFAULT 0,
                    minutes_played INTEGER DEFAULT 0,
                    shots_on_target INTEGER DEFAULT 0,
                    shot_attempts INTEGER DEFAULT 0,
                    passes_completed INTEGER DEFAULT 0,
                    passes_attempted INTEGER DEFAULT 0,
                    saves INTEGER DEFAULT 0,
                    rating REAL,
                    custom_metrics TEXT
                )
            """))
    except Exception:
        # Don't block the page if table creation fails - the error will be caught later
        pass


def _get_comprehensive_finance_data(db, league_id_int):
    """Get all financial data for the finance hub"""
    if not league_id_int:
        return {
            'total_revenue': {'amount': 0, 'display': '$0.00'},
            'revenue_breakdown': {},
            'recent_transactions': [],
            'budget_vs_actual': [],
            'sponsor_revenue': {'amount': 0, 'display': '$0.00'},
            'fee_revenue': {'amount': 0, 'display': '$0.00'},
            'in_person_revenue': {'amount': 0, 'display': '$0.00'},
            'fee_plan': None,
            'installments': [],
            'teams': []
        }

    cur = db.cursor()
    try:
        # Get total revenue by type
        cur.execute("""
            SELECT revenue_type,
                   COALESCE(SUM(amount_cents), 0) as total_cents,
                   COUNT(*) as transaction_count
            FROM league_revenue
            WHERE league_id = %s AND status = 'confirmed'
            GROUP BY revenue_type
            ORDER BY total_cents DESC
        """, (league_id_int,))

        revenue_breakdown = {}
        total_revenue_cents = 0

        for row in cur.fetchall():
            revenue_type, amount_cents, count = row
            revenue_breakdown[revenue_type] = {
                'amount_cents': amount_cents,
                'amount_display': _format_cents_to_decimal(amount_cents),
                'transaction_count': count
            }
            total_revenue_cents += amount_cents

        # Get recent transactions
        cur.execute("""
            SELECT revenue_type, amount_cents, transaction_date, description, payment_method
            FROM league_revenue
            WHERE league_id = %s
            ORDER BY transaction_date DESC, created_at DESC
            LIMIT 10
        """, (league_id_int,))

        recent_transactions = []
        for row in cur.fetchall():
            revenue_type, amount_cents, trans_date, description, payment_method = row
            recent_transactions.append({
                'type': revenue_type,
                'amount_display': _format_cents_to_decimal(amount_cents),
                'date': trans_date.strftime('%Y-%m-%d') if trans_date else '',
                'description': description or '',
                'payment_method': payment_method or ''
            })

        # Get sponsor revenue data (all active deals since sponsors aren't league-specific in current schema)
        cur.execute("""
            SELECT COALESCE(SUM(sd.deal_value_cents), 0) as sponsor_revenue
            FROM sponsorship_deals sd
            WHERE sd.deal_status = 'active'
        """)

        sponsor_revenue_cents = cur.fetchone()[0] or 0

        # Get fee plan and installments (existing logic)
        plan = _get_league_fee_plan(cur, league_id_int)
        installments = []
        fee_revenue_cents = 0

        if plan:
            cur.execute("""
                SELECT installment_id, label, due_date, amount_cents, status, notes
                FROM league_fee_installments
                WHERE plan_id = %s
                ORDER BY due_date IS NULL, due_date
            """, (plan['plan_id'],))

            for row in cur.fetchall():
                installment_id, label, due_date, amount_cents, status, notes = row
                installments.append({
                    'installment_id': installment_id,
                    'label': label,
                    'due_date': due_date.strftime('%Y-%m-%d') if due_date else '',
                    'amount_cents': amount_cents or 0,
                    'amount_display': _format_cents_to_decimal(amount_cents or 0),
                    'status': status or 'pending',
                    'notes': notes
                })
                if status == 'paid':
                    fee_revenue_cents += amount_cents or 0

        # Get in-person payments
        cur.execute("""
            SELECT COALESCE(SUM(amount_cents), 0)
            FROM in_person_payments
            WHERE league_id = %s
        """, (league_id_int,))

        in_person_revenue_cents = cur.fetchone()[0] or 0

        # Get teams for the league
        cur.execute('SELECT team_id, name FROM teams WHERE league_id = %s ORDER BY name', (league_id_int,))
        teams = [{'team_id': row[0], 'name': row[1]} for row in cur.fetchall()]

        return {
            'total_revenue': {
                'amount': total_revenue_cents,
                'display': f'${_format_cents_to_decimal(total_revenue_cents)}'
            },
            'revenue_breakdown': revenue_breakdown,
            'recent_transactions': recent_transactions,
            'sponsor_revenue': {
                'amount': sponsor_revenue_cents,
                'display': f'${_format_cents_to_decimal(sponsor_revenue_cents)}'
            },
            'fee_revenue': {
                'amount': fee_revenue_cents,
                'display': f'${_format_cents_to_decimal(fee_revenue_cents)}'
            },
            'in_person_revenue': {
                'amount': in_person_revenue_cents,
                'display': f'${_format_cents_to_decimal(in_person_revenue_cents)}'
            },
            'fee_plan': plan,
            'installments': installments,
            'teams': teams
        }

    finally:
        cur.close()


def _handle_add_in_person_payment(db):
    """Handle adding in-person payment"""
    try:
        league_id = int(request.form.get('league_id'))
        amount_cents = _parse_amount_to_cents(request.form.get('amount'))
        payment_method = request.form.get('payment_method')
        payer_name = request.form.get('payer_name', '').strip()
        description = request.form.get('description', '').strip()
        payment_date = request.form.get('payment_date') or datetime.now().date()

        if not amount_cents or amount_cents <= 0:
            flash('Valid payment amount is required.', 'error')
            return redirect(url_for('admin.league_finance_hub', league_id=league_id))

        cur = db.cursor()
        try:
            # Insert in-person payment
            cur.execute("""
                INSERT INTO in_person_payments
                (league_id, amount_cents, payment_method, payer_name, description,
                 payment_date, check_number, card_last_four, receipt_number, notes, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING payment_id
            """, (
                league_id, amount_cents, payment_method, payer_name, description,
                payment_date,
                request.form.get('check_number', '').strip() or None,
                request.form.get('card_last_four', '').strip() or None,
                request.form.get('receipt_number', '').strip() or None,
                request.form.get('notes', '').strip() or None,
                current_user.id
            ))

            payment_id = cur.fetchone()[0]

            # Also add to league_revenue for consolidated tracking
            cur.execute("""
                INSERT INTO league_revenue
                (league_id, revenue_type, amount_cents, transaction_date, description,
                 reference_id, reference_type, payment_method, status, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                league_id, 'in_person', amount_cents, payment_date,
                f'{description} - {payer_name}', payment_id, 'in_person_payment',
                payment_method, 'confirmed', current_user.id
            ))

            db.commit()
            flash('In-person payment recorded successfully.', 'success')

        except Exception as e:
            db.rollback()
            flash(f'Failed to record payment: {str(e)}', 'error')
        finally:
            cur.close()

    except (ValueError, TypeError):
        flash('Invalid payment data provided.', 'error')

    return redirect(url_for('admin.league_finance_hub',
                           league_id=request.form.get('league_id')))


def _handle_update_budget(db):
    """Handle budget update"""
    # Implementation for budget management
    flash('Budget feature coming soon.', 'info')
    return redirect(url_for('admin.league_finance_hub',
                           league_id=request.form.get('league_id')))


def _handle_generate_revenue_entry(db):
    """Handle manual revenue entry"""
    # Implementation for manual revenue entries
    flash('Manual revenue entry feature coming soon.', 'info')
    return redirect(url_for('admin.league_finance_hub',
                           league_id=request.form.get('league_id')))


def _handle_legacy_fee_actions(db):
    """Handle existing fee management actions"""
    # This will contain the existing fee management logic
    # For now, redirect to the old system
    flash('Fee management integration in progress.', 'info')
    return redirect(url_for('admin.manage_league_fees',
                           league_id=request.form.get('league_id')))


def _ensure_league_rules_table(db_wrapper):
    cur = db_wrapper.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS league_rules (
                league_id INTEGER PRIMARY KEY REFERENCES leagues(league_id) ON DELETE CASCADE,
                points_win INTEGER NOT NULL DEFAULT 3,
                points_draw INTEGER NOT NULL DEFAULT 1,
                points_loss INTEGER NOT NULL DEFAULT 0,
                tiebreakers TEXT,
                substitution_limit INTEGER,
                foreign_player_limit INTEGER,
                notes TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Create custom fields table for league rules
        cur.execute("""
            CREATE TABLE IF NOT EXISTS league_rules_custom_fields (
                field_id SERIAL PRIMARY KEY,
                league_id INTEGER NOT NULL REFERENCES leagues(league_id) ON DELETE CASCADE,
                field_name VARCHAR(100) NOT NULL,
                field_type VARCHAR(20) NOT NULL DEFAULT 'text',
                field_value TEXT,
                display_order INTEGER DEFAULT 0,
                is_required BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(league_id, field_name)
            )
        """)

        # Create index for better performance
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_league_rules_custom_fields_league_id
            ON league_rules_custom_fields(league_id)
        """)

        db_wrapper.commit()
    except Exception:
        db_wrapper.rollback()
        raise
    finally:
        cur.close()


def _round_robin_schedule(team_ids, double_round=False):
    teams = list(team_ids)
    if len(teams) < 2:
        return []

    if len(teams) % 2:
        teams.append(None)

    rounds = len(teams) - 1
    schedule = []

    for round_index in range(rounds):
        pairings = []
        for idx in range(len(teams) // 2):
            home = teams[idx]
            away = teams[-(idx + 1)]
            if home is None or away is None:
                continue
            if idx == 0 and round_index % 2 == 1:
                home, away = away, home
            pairings.append((home, away))
        schedule.append(pairings)
        teams = [teams[0]] + [teams[-1]] + teams[1:-1]

    if double_round:
        schedule.extend([[(away, home) for home, away in pairings] for pairings in schedule])

    return schedule


def _parse_tiebreakers(raw_value):
    if not raw_value:
        return ['GD', 'GF']
    tokens = []
    for token in str(raw_value).split(','):
        token = token.strip().upper()
        if token:
            tokens.append(token)
    return tokens or ['GD', 'GF']


def _standing_sort_key(record, tiebreakers):
    key = [-record['points']]
    for tb in tiebreakers:
        if tb == 'GD':
            key.append(-record['goal_difference'])
        elif tb == 'GF':
            key.append(-record['goals_for'])
        elif tb == 'GA':
            key.append(record['goals_against'])
        elif tb == 'W':
            key.append(-record['won'])
    key.append(record['team_name'].lower())
    return tuple(key)


def _init_standing_record(team_id, team_name):
    return {
        'team_id': team_id,
        'team_name': team_name,
        'played_games': 0,
        'won': 0,
        'draw': 0,
        'lost': 0,
        'goals_for': 0,
        'goals_against': 0,
        'form': [],
    }


def _get_league_fee_plan(cur, league_id):
    cur.execute(
        "SELECT plan_id, total_fee_cents, deposit_cents, currency, notes, installments_enabled, installment_count FROM league_fee_plans WHERE league_id = %s",
        (league_id,)
    )
    row = cur.fetchone()
    if not row:
        return None
    plan_id, total_fee, deposit, currency, notes, enabled, count = row
    return {
        'plan_id': plan_id,
        'league_id': league_id,
        'total_fee_cents': total_fee or 0,
        'deposit_cents': deposit,
        'currency': (currency or 'USD').upper(),
        'notes': notes,
        'installments_enabled': bool(enabled),
        'installment_count': count or 0,
        'total_fee_display': _format_cents_to_decimal(total_fee or 0),
        'deposit_display': _format_cents_to_decimal(deposit) if deposit is not None else None,
    }


def _ensure_league_fee_plan(cur, league_id, currency='USD'):
    plan = _get_league_fee_plan(cur, league_id)
    if plan:
        return plan
    now = datetime.now(timezone.utc)
    cur.execute(
        "INSERT INTO league_fee_plans (league_id, total_fee_cents, deposit_cents, currency, notes, installments_enabled, installment_count, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
        (league_id, 0, None, currency.upper(), None, False, 0, now, now)
    )
    return _get_league_fee_plan(cur, league_id)




def _load_homepage_config(cur, league_id):
    try:
        cur.execute(
            "SELECT homepage_title, homepage_subtitle, homepage_background_url, homepage_cta_text, homepage_cta_url, homepage_highlights_json FROM leagues WHERE league_id = %s",
            (league_id,)
        )
        row = cur.fetchone()
    except Exception:
        return _default_homepage_config()
    if not row:
        return _default_homepage_config()
    highlights_json = row[5] or '[]'
    try:
        highlights = json.loads(highlights_json)
    except (TypeError, json.JSONDecodeError):
        highlights = []
    return {
        'title': row[0] or '',
        'subtitle': row[1] or '',
        'background_url': row[2] or '',
        'cta_text': row[3] or '',
        'cta_url': row[4] or '',
        'highlights': highlights,
    }



def _save_homepage_config(cur, league_id, config):
    try:
        highlights_json = json.dumps(config.get('highlights', []))
    except (TypeError, ValueError):
        highlights_json = '[]'
    cur.execute(
        "UPDATE leagues SET homepage_title = %s, homepage_subtitle = %s, homepage_background_url = %s, homepage_cta_text = %s, homepage_cta_url = %s, homepage_highlights_json = %s WHERE league_id = %s",
        (
            config.get('title') or None,
            config.get('subtitle') or None,
            config.get('background_url') or None,
            config.get('cta_text') or None,
            config.get('cta_url') or None,
            highlights_json,
            league_id,
        )
    )




NAV_AUDIENCE_CHOICES = ('everyone', 'auth', 'guest', 'admin')
NAV_LAYOUT_CHOICES = ('top', 'sidebar')


def _load_navigation_settings():
    settings = _load_site_settings()
    raw_links = settings.get('navigation_links_raw', []) or []
    layout = settings.get('nav_layout', 'top') or 'top'
    return {
        'layout': layout if layout in NAV_LAYOUT_CHOICES else 'top',
        'links': raw_links,
    }


def _persist_navigation_settings(layout: str, links: list[dict]):
    layout_normalized = layout if layout in NAV_LAYOUT_CHOICES else 'top'
    db = get_db()
    cur = db.cursor()
    try:
        payload = json.dumps(links)
        cur.execute('SELECT id FROM site_settings ORDER BY id ASC LIMIT 1')
        row = cur.fetchone()
        timestamp = datetime.now(timezone.utc)
        if row:
            cur.execute(
                "UPDATE site_settings SET nav_layout = %s, navigation_json = %s, updated_at = %s WHERE id = %s",
                (layout_normalized, payload, timestamp, row[0])
            )
        else:
            cur.execute(
                "INSERT INTO site_settings (nav_layout, navigation_json, updated_at) VALUES (%s, %s, %s)",
                (layout_normalized, payload, timestamp)
            )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        cur.close()
    invalidate_site_settings_cache()
def _build_league_insights():
    db = get_db()
    cur = db.cursor()
    try:
        current_year = datetime.now(timezone.utc).year
        previous_year = current_year - 1

        teams_current = _execute_scalar(
            cur,
            """
            SELECT COUNT(DISTINCT team_id)
            FROM (
                SELECT home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                UNION
                SELECT away_team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
            ) AS teams_year
            """,
            (current_year, current_year),
        )
        teams_previous = _execute_scalar(
            cur,
            """
            SELECT COUNT(DISTINCT team_id)
            FROM (
                SELECT home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                UNION
                SELECT away_team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
            ) AS teams_year
            """,
            (previous_year, previous_year),
        )

        players_current = _execute_scalar(
            cur,
            """
            SELECT COUNT(DISTINCT p.player_id)
            FROM players p
            WHERE p.team_id IN (
                SELECT team_id FROM (
                    SELECT home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                    UNION
                    SELECT away_team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                ) AS teams_year
            )
            """,
            (current_year, current_year),
        )
        players_previous = _execute_scalar(
            cur,
            """
            SELECT COUNT(DISTINCT p.player_id)
            FROM players p
            WHERE p.team_id IN (
                SELECT team_id FROM (
                    SELECT home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                    UNION
                    SELECT away_team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                ) AS teams_year
            )
            """,
            (previous_year, previous_year),
        )

        matches_current = _execute_scalar(
            cur,
            "SELECT COUNT(*) FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s",
            (current_year,),
        )
        matches_previous = _execute_scalar(
            cur,
            "SELECT COUNT(*) FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s",
            (previous_year,),
        )
        matches_completed_current = _execute_scalar(
            cur,
            "SELECT COUNT(*) FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s AND winner IS NOT NULL",
            (current_year,),
        )
        matches_completed_previous = _execute_scalar(
            cur,
            "SELECT COUNT(*) FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s AND winner IS NOT NULL",
            (previous_year,),
        )

        revenue_current_cents = int(matches_current * AVERAGE_MATCH_REVENUE_USD * 100)
        revenue_previous_cents = int(matches_previous * AVERAGE_MATCH_REVENUE_USD * 100)

        average_team_size = (players_current / teams_current) if teams_current else None

        conversion_rate = (
            (matches_completed_current / matches_current) * 100
            if matches_current
            else None
        )

        league_data = {}

        cur.execute(
            """
            SELECT sub.league_id, l.name, COUNT(DISTINCT sub.team_id) AS team_count
            FROM (
                SELECT league_id, home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                UNION
                SELECT league_id, away_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
            ) AS sub
            JOIN leagues l ON l.league_id = sub.league_id
            GROUP BY sub.league_id, l.name
            """,
            (current_year, current_year),
        )
        for league_id, name, team_count in cur.fetchall():
            league_data[league_id] = {
                'name': name,
                'teams_current': int(team_count or 0),
                'teams_previous': 0,
                'registrations_current': 0,
                'revenue_cents': 0,
            }

        cur.execute(
            """
            SELECT sub.league_id, l.name, COUNT(DISTINCT sub.team_id) AS team_count
            FROM (
                SELECT league_id, home_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
                UNION
                SELECT league_id, away_team_id AS team_id FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s
            ) AS sub
            JOIN leagues l ON l.league_id = sub.league_id
            GROUP BY sub.league_id, l.name
            """,
            (previous_year, previous_year),
        )
        for league_id, name, team_count in cur.fetchall():
            entry = league_data.setdefault(
                league_id,
                {
                    'name': name,
                    'teams_current': 0,
                    'teams_previous': 0,
                    'registrations_current': 0,
                    'revenue_cents': 0,
                },
            )
            entry['teams_previous'] = int(team_count or 0)

        cur.execute(
            "SELECT league_id, COUNT(*) FROM matches WHERE EXTRACT(YEAR FROM utc_date) = %s GROUP BY league_id",
            (current_year,),
        )
        for league_id, match_count in cur.fetchall():
            entry = league_data.setdefault(
                league_id,
                {
                    'name': '',
                    'teams_current': 0,
                    'teams_previous': 0,
                    'registrations_current': 0,
                    'revenue_cents': 0,
                },
            )
            entry['registrations_current'] = int(match_count or 0)
            entry['revenue_cents'] = int((match_count or 0) * AVERAGE_MATCH_REVENUE_USD * 100)

        league_breakdown = []
        for entry in league_data.values():
            metric = _calc_metric(entry['teams_current'], entry['teams_previous'])
            league_breakdown.append(
                {
                    'name': entry['name'],
                    'teams_current': entry['teams_current'],
                    'teams_previous': entry['teams_previous'],
                    'team_delta': metric['delta'],
                    'team_delta_pct': metric['delta_pct'],
                    'registrations_current': entry['registrations_current'],
                    'revenue_cents': entry['revenue_cents'],
                }
            )

        league_breakdown.sort(key=lambda item: item['teams_current'], reverse=True)
        league_breakdown = league_breakdown[:6]

        return {
            'current_year': current_year,
            'previous_year': previous_year,
            'totals': {
                'teams': _calc_metric(teams_current, teams_previous),
                'players': _calc_metric(players_current, players_previous),
                'registrations': _calc_metric(matches_current, matches_previous),
                'paid_registrations': _calc_metric(matches_completed_current, matches_completed_previous),
                'revenue': _calc_metric(revenue_current_cents, revenue_previous_cents),
            },
            'conversion_rate': conversion_rate,
            'average_team_size': average_team_size,
            'league_breakdown': league_breakdown,
            'updated_at': datetime.now(timezone.utc),
        }
    finally:
        cur.close()


@admin_bp.route('/')
@admin_required
def admin_dashboard():
    try:
        insights = _build_league_insights()
    except Exception as e:
        print(f"Error building insights: {e}")
        import traceback
        traceback.print_exc()
        insights = None
    return render_template('admin.html', insights=insights)

@admin_bp.route('/manage_stadiums', methods=['GET', 'POST'])
@admin_required
def manage_stadiums():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        action = (request.form.get('action') or 'publish').strip().lower()
        version_label = (request.form.get('version_label') or '').strip() or None
        author_id = getattr(current_user, 'id', None) if current_user.is_authenticated else None

        if action == 'discard_preview':
            discard_site_theme_preview()
            session.pop('theme_preview_active', None)
            flash('Preview discarded.', 'info')
            payload_theme = copy.deepcopy(theme)
        payload = {
            'site_title': site_title,
            'brand_image_url': brand_image_url,
            'primary_color': primary_color,
            'favicon_url': favicon_url,
            'league_tagline': league_tagline,
            'contact_email': contact_email,
            'social_links': social_links,
            'feature_flags': feature_flags,
            'theme': payload_theme,
            'nav_layout': settings.get('nav_layout'),
            'navigation_links_raw': settings.get('navigation_links_raw'),
        }

        if action == 'save_preview':
            try:
                save_site_theme_preview(payload, author_id, version_label)
                session['theme_preview_active'] = True
                flash('Preview saved. Activate preview to review, then publish when ready.', 'success')
            except Exception as exc:
                flash('Failed to save preview: ' + str(exc), 'error')
            return redirect(url_for('admin.site_settings'))

        if action in {'publish', 'save'}:
            try:
                apply_site_payload(payload)
            except Exception as exc:
                flash('Failed to publish theme: ' + str(exc), 'error')
                return redirect(url_for('admin.site_settings'))

            publish_site_theme(payload, author_id, version_label)
            session.pop('theme_preview_active', None)
            flash('Site theme published live.', 'success')
            return redirect(url_for('admin.site_settings'))

        flash('Unknown action requested.', 'error')
        return redirect(url_for('admin.site_settings'))

        if action == 'activate_preview':
            preview = theme_preview_record or get_site_theme_preview()
            if preview:
                session['theme_preview_active'] = True
                flash('Preview mode enabled for this session.', 'success')
            else:
                flash('There is no saved preview to activate.', 'warning')
            return redirect(url_for('admin.site_settings'))

        if action == 'deactivate_preview':
            session.pop('theme_preview_active', None)
            flash('Preview mode disabled. Showing published theme.', 'info')
            return redirect(url_for('admin.site_settings'))

        if action == 'restore_version':
            version_id_raw = request.form.get('version_id')
            try:
                version_id = int(version_id_raw) if version_id_raw else None
            except (TypeError, ValueError):
                version_id = None
            if not version_id:
                flash('Invalid version identifier.', 'error')
                return redirect(url_for('admin.site_settings'))
            version = get_site_theme_version(version_id)
            if not version:
                flash('Version not found.', 'error')
                return redirect(url_for('admin.site_settings'))
            payload = version.get('payload') or {}
            try:
                apply_site_payload(payload)
                publish_site_theme(payload, author_id, version_label or (version.get('label') or f'Restore #{version_id}'))
                session.pop('theme_preview_active', None)
                flash('Theme restored from history and published live.', 'success')
            except Exception as exc:
                flash('Failed to restore theme: ' + str(exc), 'error')
            return redirect(url_for('admin.site_settings'))

        # For preview save/publish actions we continue to collect form data
        try:
            stadium_id = request.form.get('stadium_id')
            name = request.form['name']
            location = request.form['location']
            capacity = request.form['capacity']

            if 'add' in request.form:
                cur.execute('INSERT INTO stadiums (name, location, capacity) VALUES (%s, %s, %s)', 
                            (name, location, capacity))
                flash('Stadium added successfully', 'success')
            elif 'edit' in request.form and stadium_id:
                cur.execute('UPDATE stadiums SET name = %s, location = %s, capacity = %s WHERE stadium_id = %s', 
                            (name, location, capacity, stadium_id))
                flash('Stadium updated successfully', 'success')
            elif 'delete' in request.form and stadium_id:
                cur.execute('DELETE FROM stadiums WHERE stadium_id = %s', (stadium_id,))
                flash('Stadium deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_stadiums'))

    cur.execute('SELECT stadium_id, name, location, capacity FROM stadiums')
    stadiums = cur.fetchall()
    cur.close()
    return render_template('manage_stadiums.html', stadiums=stadiums)

@admin_bp.route('/manage_leagues', methods=['GET', 'POST'])
@admin_required
def manage_leagues():
    db = get_db()
    cur = db.cursor()

    def _clean(value):
        if value is None:
            return None
        value = value.strip()
        return value or None

    if request.method == 'POST':
        try:
            league_id = request.form.get('league_id')
            name = _clean(request.form.get('name'))
            country = _clean(request.form.get('country'))
            sport = _clean(request.form.get('sport')) or 'soccer'
            primary_color = _clean(request.form.get('primary_color'))
            secondary_color = _clean(request.form.get('secondary_color'))
            accent_color = _clean(request.form.get('accent_color'))
            text_color = _clean(request.form.get('text_color'))
            logo_url = _clean(request.form.get('logo_url'))
            hero_image_url = _clean(request.form.get('hero_image_url'))

            if not name or not country:
                flash('Name and country are required', 'error')
                return redirect(url_for('admin.manage_leagues'))

            if 'add' in request.form:
                cur.execute(
                    "INSERT INTO leagues (name, country, sport, primary_color, secondary_color, accent_color, text_color, logo_url, hero_image_url) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (name, country, sport, primary_color, secondary_color, accent_color, text_color, logo_url, hero_image_url)
                )
                flash('League added successfully', 'success')
            elif 'edit' in request.form and league_id:
                cur.execute(
                    "UPDATE leagues SET name = %s, country = %s, sport = %s, primary_color = %s, secondary_color = %s, accent_color = %s, text_color = %s, logo_url = %s, hero_image_url = %s WHERE league_id = %s",
                    (name, country, sport, primary_color, secondary_color, accent_color, text_color, logo_url, hero_image_url, league_id)
                )
                flash('League updated successfully', 'success')
            elif 'delete' in request.form:
                delete_id = request.form.get('deleteItemId') or league_id
                if delete_id:
                    cur.execute('DELETE FROM leagues WHERE league_id = %s', (delete_id,))
                    flash('League deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_leagues'))

    rows = []
    try:
        cur.execute(
            "SELECT league_id, name, country, sport, primary_color, secondary_color, accent_color, text_color, logo_url, hero_image_url FROM leagues ORDER BY name"
        )
        rows = cur.fetchall()
    except Exception as e:
        db.rollback()
        flash('Failed to load leagues: ' + str(e), 'error')
    finally:
        cur.close()

    columns = ['id', 'name', 'country', 'sport', 'primary_color', 'secondary_color', 'accent_color', 'text_color', 'logo_url', 'hero_image_url']
    leagues = [dict(zip(columns, row)) for row in rows]
    return render_template('manage_leagues.html', leagues=leagues)




@admin_bp.route('/homepage_builder', methods=['GET'])
@admin_required
def homepage_builder():
    """New drag-and-drop homepage builder"""
    db = get_db()
    cur = db.cursor()

    # Create homepage_blocks table if it doesn't exist
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS homepage_blocks (
                id SERIAL PRIMARY KEY,
                league_id INTEGER REFERENCES leagues(league_id) ON DELETE CASCADE,
                blocks JSONB NOT NULL DEFAULT '[]'::jsonb,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(league_id)
            )
        """)
        db.commit()
    except Exception as e:
        db.rollback()

    # Get leagues
    cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = [{'id': row[0], 'name': row[1]} for row in cur.fetchall()]

    selected_league_id = request.args.get('league_id')
    if not selected_league_id and leagues:
        selected_league_id = str(leagues[0]['id'])

    # Load blocks for selected league
    blocks = []
    if selected_league_id:
        cur.execute('SELECT blocks FROM homepage_blocks WHERE league_id = %s', (selected_league_id,))
        result = cur.fetchone()
        if result and result[0]:
            blocks = result[0]

    # Get available media for selection
    org = getattr(g, 'org', None)
    media_items = []
    query = MediaAsset.query.order_by(MediaAsset.created_at.desc())
    if org is not None:
        query = query.filter(MediaAsset.org_id == org.id)
    assets = query.all()
    for item in serialize_media_collection(assets):
        entry = dict(item)
        entry['type'] = entry.get('media_type')
        entry['thumbnail'] = entry.get('url')
        media_items.append(entry)

    # Get available seasons
    cur.execute('SELECT season_id, year FROM seasons ORDER BY year DESC')
    seasons = [{'id': row[0], 'year': row[1]} for row in cur.fetchall()]

    # Get upcoming matches for selection
    cur.execute('''
        SELECT m.match_id, t1.name as home_team, t2.name as away_team,
               m.utc_date, s.year, l.name as league_name
        FROM matches m
        JOIN teams t1 ON m.home_team_id = t1.team_id
        JOIN teams t2 ON m.away_team_id = t2.team_id
        JOIN seasons s ON m.season_id = s.season_id
        JOIN leagues l ON m.league_id = l.league_id
        WHERE m.utc_date >= CURRENT_DATE
        ORDER BY m.utc_date ASC
        LIMIT 50
    ''')
    matches = [{'id': row[0], 'home': row[1], 'away': row[2], 'date': row[3].isoformat() if row[3] else '',
                'season': row[4], 'league': row[5]} for row in cur.fetchall()]

    cur.close()

    default_media_categories = {'match_highlights', 'training', 'team_photos', 'events', 'celebrations', 'facilities'}
    site_settings = _load_site_settings()
    theme_payload = copy.deepcopy(site_settings.get('theme', DEFAULT_THEME_CONFIG))
    cta_slots = normalize_cta_slots(site_settings.get('cta_slots') or theme_payload.get('cta_slots'))
    cta_slot_defaults = default_cta_slots()
    cta_slot_meta = [
        {"key": "hero_primary", "label": "Hero Primary", "description": "Main hero action shown in prominent hero sections."},
        {"key": "hero_secondary", "label": "Hero Secondary", "description": "Secondary hero button for alternate action."},
        {"key": "footer_primary", "label": "Footer Primary", "description": "Primary footer CTA displayed near the page footer."},
        {"key": "footer_secondary", "label": "Footer Secondary", "description": "Optional supplemental footer CTA."},
    ]

    media_categories = sorted(default_media_categories.union({item['category'] for item in media_items if item.get('category')}))

    return render_template('manage_homepage_builder.html',
                         leagues=leagues,
                         selected_league_id=selected_league_id,
                         blocks=blocks,
                         media_items=media_items,
                         media_categories=media_categories,
                         seasons=seasons,
                         matches=matches,
                         cta_slots=cta_slots,
                         cta_slot_defaults=cta_slot_defaults,
                         cta_slot_meta=cta_slot_meta)


@admin_bp.route('/save_homepage_blocks', methods=['POST'])
@admin_required
def save_homepage_blocks():
    """Save homepage blocks configuration"""
    from flask import jsonify

    try:
        data = request.get_json()
        league_id = data.get('league_id')
        blocks = data.get('blocks', [])

        db = get_db()
        cur = db.cursor()

        cur.execute("""
            INSERT INTO homepage_blocks (league_id, blocks, updated_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (league_id) DO UPDATE SET
                blocks = EXCLUDED.blocks,
                updated_at = CURRENT_TIMESTAMP
        """, (league_id, json.dumps(blocks)))

        db.commit()
        cur.close()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



@admin_bp.route('/homepage_builder/cta_slots', methods=['POST'])
@admin_required
def save_cta_slots():
    """Persist CTA slot updates to the theme preview or publish live."""
    from flask import jsonify

    data = request.get_json(silent=True) or {}
    slots = data.get('slots') or {}
    mode = (data.get('mode') or 'preview').strip().lower()
    if mode not in {'preview', 'publish'}:
        mode = 'preview'
    label = (data.get('label') or '').strip() or None
    activate_preview = data.get('activate_preview', True)
    if isinstance(activate_preview, str):
        activate_preview = activate_preview.strip().lower() in {'1', 'true', 'yes', 'on'}

    try:
        site_settings = _load_site_settings()
        theme_payload = copy.deepcopy(site_settings.get('theme', DEFAULT_THEME_CONFIG))
        merged_slots = merge_theme_cta_slots(theme_payload, slots)
        payload = {'theme': theme_payload}
        author_id = getattr(current_user, 'id', None) if current_user.is_authenticated else None

        if mode == 'publish':
            apply_site_payload(payload)
            publish_site_theme(payload, author_id, label)
            session.pop('theme_preview_active', None)
            message = 'CTA slots published live.'
            result_mode = 'published'
        else:
            save_site_theme_preview(payload, author_id, label)
            if activate_preview:
                session['theme_preview_active'] = True
            message = 'CTA slots saved to preview.'
            result_mode = 'preview'

        response_slots = normalize_cta_slots(merged_slots, include_defaults=True)
        return jsonify({
            'success': True,
            'slots': response_slots,
            'mode': result_mode,
            'message': message,
        })
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@admin_bp.route('/preview_homepage', methods=['POST'])
@admin_required
def preview_homepage():
    """Store blocks in session for preview"""
    from flask import jsonify, session
    import traceback

    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400

        league_id = data.get('league_id')
        blocks = data.get('blocks', [])

        # Store in session for preview
        session['preview_blocks'] = blocks
        session['preview_league_id'] = league_id
        session.modified = True

        preview_url = url_for('admin.preview_homepage_view')
        return jsonify({'success': True, 'preview_url': preview_url})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/preview_homepage_view')
@admin_required
def preview_homepage_view():
    """Render preview of homepage with blocks from session"""
    from flask import session

    blocks = session.get('preview_blocks', [])
    league_id = session.get('preview_league_id')

    db = get_db()
    cur = db.cursor()

    # Get league info
    league_name = "Preview"
    if league_id:
        cur.execute('SELECT name FROM leagues WHERE league_id = %s', (league_id,))
        result = cur.fetchone()
        if result:
            league_name = result[0]

    # Helper function to process a single block
    def process_block(block):
        block_data = {'type': block['type'], 'settings': block.get('settings', {})}

        # Fetch actual data based on block type
        if block['type'] == 'stats':
            cur.execute('SELECT COUNT(*) FROM teams')
            total_teams = cur.fetchone()[0]
            cur.execute('SELECT COUNT(*) FROM players')
            total_players = cur.fetchone()[0]
            cur.execute('SELECT COUNT(*) FROM matches')
            total_matches = cur.fetchone()[0]
            block_data['data'] = {
                'total_teams': total_teams,
                'total_players': total_players,
                'total_matches': total_matches
            }

        elif block['type'] == 'gallery':
            settings = block.get('settings', {})
            org = getattr(g, 'org', None)
            if settings.get('displayMode') == 'manual' and settings.get('selectedMedia'):
                media_ids = [str(mid) for mid in settings['selectedMedia'] if mid]
                if media_ids:
                    query = MediaAsset.query
                    if org is not None:
                        query = query.filter(MediaAsset.org_id == org.id)
                    assets = query.filter(MediaAsset.id.in_(media_ids)).all()
                    asset_map = {str(asset.id): asset for asset in assets}
                    block_data['data'] = [
                        {
                            'id': mid,
                            'title': asset_map[mid].title if mid in asset_map else '',
                            'url': asset_map[mid].url if mid in asset_map else '',
                            'type': asset_map[mid].media_type if mid in asset_map else 'image',
                            'category': asset_map[mid].category if mid in asset_map else None,
                        }
                        for mid in media_ids
                        if mid in asset_map
                    ]
                else:
                    block_data['data'] = []
            else:
                limit = settings.get('limit', 6)
                category = settings.get('category', '')
                query = MediaAsset.query.filter(MediaAsset.media_type == 'image')
                if org is not None:
                    query = query.filter(MediaAsset.org_id == org.id)
                if category:
                    query = query.filter(MediaAsset.category == category)
                assets = query.order_by(MediaAsset.created_at.desc()).limit(limit).all()
                block_data['data'] = [
                    {
                        'id': asset.id,
                        'title': asset.title,
                        'url': asset.url,
                        'type': asset.media_type,
                        'category': asset.category,
                    }
                    for asset in assets
                ]

        elif block['type'] == 'video':
            settings = block.get('settings', {})
            if settings.get('videoSource') == 'library' and settings.get('selectedVideo'):
                query = MediaAsset.query.filter(MediaAsset.media_type == 'video')
                if org is not None:
                    query = query.filter(MediaAsset.org_id == org.id)
                asset = query.filter(MediaAsset.id == settings['selectedVideo']).first()
                if asset:
                    block_data['data'] = {'url': asset.url}

        elif block['type'] == 'matches':
            settings = block.get('settings', {})
            if settings.get('matchDisplayMode') == 'manual' and settings.get('selectedMatches'):
                match_ids = settings['selectedMatches']
                if match_ids:
                    placeholders = ','.join(['%s'] * len(match_ids))
                    cur.execute(f'''
                        SELECT m.match_id, t1.name as home_team, t2.name as away_team,
                               m.utc_date, m.match_time, l.name as league_name
                        FROM matches m
                        JOIN teams t1 ON m.home_team_id = t1.team_id
                        JOIN teams t2 ON m.away_team_id = t2.team_id
                        JOIN leagues l ON m.league_id = l.league_id
                        WHERE m.match_id IN ({placeholders})
                        ORDER BY m.utc_date ASC
                    ''', match_ids)
                    block_data['data'] = [{'id': r[0], 'home': r[1], 'away': r[2], 'date': r[3],
                                          'time': r[4], 'league': r[5]} for r in cur.fetchall()]
                else:
                    block_data['data'] = []
            else:
                # Auto mode
                limit = settings.get('limit', 5)
                show_past = settings.get('showPast', False)
                filter_season = settings.get('filterBySeason', '')

                query = '''
                    SELECT m.match_id, t1.name as home_team, t2.name as away_team,
                           m.utc_date, m.match_time, l.name as league_name
                    FROM matches m
                    JOIN teams t1 ON m.home_team_id = t1.team_id
                    JOIN teams t2 ON m.away_team_id = t2.team_id
                    JOIN leagues l ON m.league_id = l.league_id
                    WHERE 1=1
                '''
                params = []
                if not show_past:
                    query += ' AND m.utc_date >= CURRENT_DATE'
                if filter_season:
                    query += ' AND m.season_id = %s'
                    params.append(filter_season)
                query += ' ORDER BY m.utc_date ASC LIMIT %s'
                params.append(limit)

                cur.execute(query, params)
                block_data['data'] = [{'id': r[0], 'home': r[1], 'away': r[2], 'date': r[3],
                                      'time': r[4], 'league': r[5]} for r in cur.fetchall()]

        elif block['type'] == 'standings':
            settings = block.get('settings', {})
            limit = settings.get('limit', 10)
            season_id = settings.get('seasonId', '')

            # Get standings data (simplified - you may need to adjust based on your standings calculation)
            query = '''
                SELECT t.name, t.logo_url,
                       COUNT(DISTINCT m.match_id) as played,
                       0 as points
                FROM teams t
                LEFT JOIN matches m ON (t.team_id = m.home_team_id OR t.team_id = m.away_team_id)
            '''
            params = []
            if season_id:
                query += ' AND m.season_id = %s'
                params.append(season_id)
            query += ' GROUP BY t.team_id, t.name, t.logo_url LIMIT %s'
            params.append(limit)

            cur.execute(query, params)
            block_data['data'] = [{'name': r[0], 'logo': r[1], 'played': r[2], 'points': r[3]}
                                 for r in cur.fetchall()]

        return block_data

    # Get data for rendering blocks
    rendered_blocks = []
    for block in blocks:
        if block['type'] == 'container':
            # Process container with nested blocks
            container_data = {
                'type': 'container',
                'layout': block.get('layout', 'split-50-50'),
                'leftBlock': None,
                'rightBlock': None
            }

            # Process left block if it exists
            if block.get('leftBlock'):
                container_data['leftBlock'] = process_block(block['leftBlock'])

            # Process right block if it exists
            if block.get('rightBlock'):
                container_data['rightBlock'] = process_block(block['rightBlock'])

            rendered_blocks.append(container_data)
        else:
            # Regular block
            rendered_blocks.append(process_block(block))

    cur.close()

    return render_template('preview_homepage.html',
                         blocks=rendered_blocks,
                         league_name=league_name)


@admin_bp.route('/league_homepage', methods=['GET', 'POST'])
@admin_required
def manage_league_homepage():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = [{'id': row[0], 'name': row[1]} for row in cur.fetchall()]
    cur.close()

    selected_league_id = request.form.get('league_id') or request.args.get('league_id')
    if not selected_league_id and leagues:
        selected_league_id = str(leagues[0]['id'])
    league_id_int = int(selected_league_id) if selected_league_id else None

    config = _default_homepage_config()
    if league_id_int:
        cur = get_db().cursor()
        config = _load_homepage_config(cur, league_id_int)
        cur.close()

    if request.method == 'POST':
        action = request.form.get('action') or 'save'
        cur = get_db().cursor()
        try:
            if not league_id_int:
                flash('Please choose a league first.', 'error')
                return redirect(url_for('admin.manage_league_homepage'))

            if action == 'reset':
                _save_homepage_config(cur, league_id_int, _default_homepage_config())
                db.commit()
                flash('Homepage settings reset to defaults.', 'success')
            else:
                highlights = []
                for index in range(1, 4):
                    title = (request.form.get(f'highlight_title_{index}') or '').strip()
                    body = (request.form.get(f'highlight_body_{index}') or '').strip()
                    icon = (request.form.get(f'highlight_icon_{index}') or '').strip()
                    if title or body:
                        highlights.append({
                            'title': title,
                            'body': body,
                            'icon': icon or 'ph ph-star',
                        })

                config_payload = {
                    'title': (request.form.get('homepage_title') or '').strip() or None,
                    'subtitle': (request.form.get('homepage_subtitle') or '').strip() or None,
                    'background_url': (request.form.get('homepage_background_url') or '').strip() or None,
                    'cta_text': (request.form.get('homepage_cta_text') or '').strip() or None,
                    'cta_url': (request.form.get('homepage_cta_url') or '').strip() or None,
                    'highlights': highlights,
                }
                _save_homepage_config(cur, league_id_int, config_payload)
                db.commit()
                flash('Homepage content saved.', 'success')
        except Exception as exc:
            db.rollback()
            flash('Failed to update homepage: ' + str(exc), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_league_homepage', league_id=selected_league_id))

    if league_id_int:
        cur = get_db().cursor()
        config = _load_homepage_config(cur, league_id_int)
        cur.close()
    else:
        config = _default_homepage_config()

    highlights = list(config.get('highlights', []))
    while len(highlights) < 3:
        highlights.append({'title': '', 'body': '', 'icon': 'ph ph-star'})

    return render_template(
        'manage_league_homepage.html',
        leagues=leagues,
        selected_league_id=selected_league_id,
        config=config,
        highlights=highlights,
    )

@admin_bp.route('/manage_navigation', methods=['GET', 'POST'])
@admin_required
def manage_navigation():
    nav_config = _load_navigation_settings()
    current_links = nav_config.get('links', [])
    current_layout = nav_config.get('layout', 'top')

    # Load current nav_style from theme settings
    settings = _load_site_settings()
    theme = settings.get('theme', {})
    components = theme.get('components', {})
    current_nav_style = components.get('nav_style', 'glass')

    if request.method == 'POST':
        submitted_layout = (request.form.get('nav_layout') or 'top').strip().lower()
        submitted_nav_style = (request.form.get('nav_style') or 'glass').strip().lower()
        labels = request.form.getlist('nav_label[]')
        types = request.form.getlist('nav_type[]')
        values = request.form.getlist('nav_value[]')
        icons = request.form.getlist('nav_icon[]')
        audiences = request.form.getlist('nav_audience[]')

        nav_items: list[dict] = []
        for idx, label in enumerate(labels):
            label_text = (label or '').strip()
            value_text = (values[idx] if idx < len(values) else '').strip()
            if not label_text or not value_text:
                continue
            item_type = (types[idx] if idx < len(types) else 'url').strip().lower()
            if item_type not in {'url', 'endpoint'}:
                item_type = 'url'
            icon_value = (icons[idx] if idx < len(icons) else '').strip() or None
            audience_value = (audiences[idx] if idx < len(audiences) else 'everyone').strip().lower()
            if audience_value not in NAV_AUDIENCE_CHOICES:
                audience_value = 'everyone'
            open_in_new = bool(request.form.get(f'nav_newtab_{idx}'))
            nav_items.append({
                'label': label_text,
                'type': item_type,
                'value': value_text,
                'icon': icon_value,
                'audience': audience_value,
                'open_in_new': open_in_new,
            })

        if not nav_items:
            nav_items = [dict(item) for item in current_links]

        try:
            _persist_navigation_settings(submitted_layout, nav_items)

            # Also persist nav_style to theme settings
            if submitted_nav_style in ['glass', 'gradient', 'minimal', 'solid']:
                components['nav_style'] = submitted_nav_style
                theme['components'] = components

                # Save theme back to database
                db = get_db()
                cur = db.cursor()
                theme_json = json.dumps(theme)
                timestamp = datetime.now(timezone.utc)
                cur.execute('SELECT id FROM site_settings ORDER BY id ASC LIMIT 1')
                row = cur.fetchone()
                if row:
                    cur.execute(
                        "UPDATE site_settings SET theme_config_json = %s, updated_at = %s WHERE id = %s",
                        (theme_json, timestamp, row[0])
                    )
                db.commit()
                cur.close()
                invalidate_site_settings_cache()

            flash('Navigation settings updated.', 'success')
        except Exception as exc:
            flash('Failed to update navigation: ' + str(exc), 'error')
        return redirect(url_for('admin.manage_navigation'))

    links_for_form = current_links if current_links else [
        {
            'label': '',
            'type': 'url',
            'value': '',
            'icon': '',
            'audience': 'everyone',
            'open_in_new': False,
        }
    ]

    return render_template(
        'manage_navigation.html',
        nav_links=links_for_form,
        nav_layout=current_layout,
        nav_style=current_nav_style,
        nav_styles=['glass', 'gradient', 'minimal', 'solid'],
        layout_choices=NAV_LAYOUT_CHOICES,
        audience_choices=NAV_AUDIENCE_CHOICES,
        icon_choices=ICON_PICKER_CHOICES,
    )


@admin_bp.route('/manage_footer', methods=['GET', 'POST'])
@login_required
@admin_required
@tenant_required
def manage_footer():
    settings = _load_site_settings()
    theme_payload = copy.deepcopy(settings.get('theme', DEFAULT_THEME_CONFIG))
    footer_config = ensure_theme_footer(theme_payload)
    feature_flags = copy.deepcopy(settings.get('feature_flags', {}))
    cta_slots = normalize_cta_slots(settings.get('cta_slots') or theme_payload.get('cta_slots'))

    if request.method == 'POST':
        raw_payload = request.form.get('footer_payload') or '{}'
        try:
            payload_data = json.loads(raw_payload)
        except (TypeError, ValueError, json.JSONDecodeError):
            flash('Footer configuration could not be parsed. Please try again.', 'error')
            return redirect(url_for('admin.manage_footer'))

        normalized_footer = normalize_footer_config(payload_data)
        show_social = bool(request.form.get('show_footer_social'))
        feature_flags['show_footer_social'] = show_social

        cta_updates = {}
        for slot_key in ('footer_primary', 'footer_secondary'):
            prefix = f"{slot_key}_"
            cta_updates[slot_key] = {
                'label': request.form.get(prefix + 'label'),
                'url': request.form.get(prefix + 'url'),
                'style': request.form.get(prefix + 'style'),
                'icon': request.form.get(prefix + 'icon'),
                'enabled': bool(request.form.get(prefix + 'enabled')),
                'new_tab': bool(request.form.get(prefix + 'new_tab')),
            }
        merge_theme_cta_slots(theme_payload, cta_updates)
        theme_payload['footer'] = normalized_footer

        try:
            apply_site_payload({
                'theme': theme_payload,
                'feature_flags': feature_flags,
            })
            flash('Footer updated successfully.', 'success')
        except Exception as exc:
            flash('Failed to update footer: ' + str(exc), 'error')
        return redirect(url_for('admin.manage_footer'))

    footer_styles = sorted(FOOTER_ALLOWED_STYLES)
    show_social_flag = feature_flags.get('show_footer_social', True)

    return render_template(
        'manage_footer.html',
        footer_config=footer_config,
        footer_styles=footer_styles,
        icon_choices=ICON_PICKER_CHOICES,
        cta_slots=cta_slots,
        show_footer_social=show_social_flag,
        max_columns=FOOTER_MAX_COLUMNS,
        max_links=FOOTER_MAX_LINKS_PER_COLUMN,
    )


@admin_bp.route('/league_finance_hub', methods=['GET', 'POST'])
@admin_required
def league_finance_hub():
    """Comprehensive league finance management hub"""
    db = get_db()
    try:
        _ensure_finance_hub_tables(db)
    except Exception as exc:
        flash('Failed to prepare finance hub storage: ' + str(exc), 'error')
        return redirect(url_for('admin.admin_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_in_person_payment':
            return _handle_add_in_person_payment(db)
        elif action == 'update_budget':
            return _handle_update_budget(db)
        elif action == 'generate_revenue_entry':
            return _handle_generate_revenue_entry(db)
        # Continue handling existing fee management actions
        else:
            return _handle_legacy_fee_actions(db)

    # Get league selection
    base_cur = db.cursor()
    base_cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = [{'id': row[0], 'name': row[1]} for row in base_cur.fetchall()]
    base_cur.close()

    selected_league_id = request.form.get('league_id') or request.args.get('league_id')
    if not selected_league_id and leagues:
        selected_league_id = str(leagues[0]['id'])

    league_id_int = int(selected_league_id) if selected_league_id else None

    # Gather comprehensive financial data
    finance_data = _get_comprehensive_finance_data(db, league_id_int)

    return render_template('league_finance_hub.html',
                         leagues=leagues,
                         selected_league_id=selected_league_id,
                         current_datetime=datetime.now().strftime('%Y-%m-%d %H:%M'),
                         current_date=datetime.now().strftime('%Y-%m-%d'),
                         **finance_data)


@admin_bp.route('/manage_league_fees', methods=['GET', 'POST'])
@admin_required
def manage_league_fees():
    db = get_db()
    base_cur = db.cursor()
    base_cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = [{'id': row[0], 'name': row[1]} for row in base_cur.fetchall()]
    base_cur.close()

    selected_league_id = request.form.get('league_id') or request.args.get('league_id')
    if not selected_league_id and leagues:
        selected_league_id = str(leagues[0]['id'])

    league_id_int = int(selected_league_id) if selected_league_id else None

    if request.method == 'POST':
        action = request.form.get('action')
        cur = db.cursor()
        try:
            if not league_id_int:
                flash('Please select a league to manage fees.', 'error')
                return redirect(url_for('admin.manage_league_fees'))

            if action == 'update_plan':
                total_fee_cents = _parse_amount_to_cents(request.form.get('total_fee')) or 0
                deposit_cents = _parse_amount_to_cents(request.form.get('deposit'))
                currency = (request.form.get('currency') or 'USD').upper()[:3]
                installments_enabled = True if request.form.get('installments_enabled') else False
                notes = (request.form.get('notes') or '').strip() or None
                now = datetime.now(timezone.utc).isoformat()

                plan = _get_league_fee_plan(cur, league_id_int)
                installment_count = request.form.get('installment_count')
                try:
                    installment_count_int = int(installment_count) if installment_count else 0
                except ValueError:
                    installment_count_int = 0

                if plan:
                    cur.execute(
                        "UPDATE league_fee_plans SET total_fee_cents = %s, deposit_cents = %s, currency = %s, notes = %s, installments_enabled = %s, installment_count = %s, updated_at = %s WHERE league_id = %s",
                        (total_fee_cents, deposit_cents, currency, notes, installments_enabled, installment_count_int, now, league_id_int)
                    )
                else:
                    cur.execute(
                        "INSERT INTO league_fee_plans (league_id, total_fee_cents, deposit_cents, currency, notes, installments_enabled, installment_count, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        (league_id_int, total_fee_cents, deposit_cents, currency, notes, installments_enabled, installment_count_int, now, now)
                    )
                db.commit()
                flash('League fee plan saved.', 'success')

            elif action == 'add_installment':
                label = (request.form.get('label') or 'Installment').strip()
                amount_cents = _parse_amount_to_cents(request.form.get('amount'))
                if amount_cents is None:
                    raise ValueError('Installment amount is required.')
                due_date = request.form.get('due_date')
                notes = (request.form.get('installment_notes') or '').strip() or None

                plan = _ensure_league_fee_plan(cur, league_id_int)

                cur.execute(
                    "INSERT INTO league_fee_installments (plan_id, label, due_date, amount_cents, status, notes) VALUES (%s, %s, %s, %s, %s, %s)",
                    (plan['plan_id'], label, due_date or None, amount_cents, request.form.get('status') or 'pending', notes)
                )
                db.commit()
                flash('Installment added.', 'success')

            elif action == 'delete_installment':
                installment_id = request.form.get('installment_id')
                if installment_id:
                    cur.execute('DELETE FROM league_installment_status WHERE installment_id = %s', (installment_id,))
                    cur.execute('DELETE FROM league_fee_installments WHERE installment_id = %s', (installment_id,))
                    db.commit()
                    flash('Installment removed.', 'success')

            elif action == 'update_status':
                installment_id = request.form.get('installment_id')
                team_id = request.form.get('team_id')
                status = (request.form.get('status') or 'pending').strip()
                if not installment_id or not team_id:
                    raise ValueError('Installment and team are required for status updates.')
                amount_paid_cents = _parse_amount_to_cents(request.form.get('amount_paid'))
                paid_at_value = request.form.get('paid_at')
                paid_at = None
                if paid_at_value:
                    try:
                        paid_at = datetime.strptime(paid_at_value, '%Y-%m-%d').date().isoformat()
                    except ValueError:
                        raise ValueError('Invalid paid date format.')
                notes = (request.form.get('status_notes') or '').strip() or None

                cur.execute(
                    'SELECT status_id FROM league_installment_status WHERE installment_id = %s AND team_id = %s',
                    (installment_id, team_id)
                )
                existing = cur.fetchone()
                if existing:
                    cur.execute(
                        'UPDATE league_installment_status SET status = %s, amount_paid_cents = %s, paid_at = %s, notes = %s WHERE status_id = %s',
                        (status, amount_paid_cents, paid_at, notes, existing[0])
                    )
                else:
                    cur.execute(
                        'INSERT INTO league_installment_status (installment_id, team_id, status, amount_paid_cents, paid_at, notes) VALUES (%s, %s, %s, %s, %s, %s)',
                        (installment_id, team_id, status, amount_paid_cents, paid_at, notes)
                    )
                db.commit()
                flash('Installment status updated.', 'success')
        except ValueError as exc:
            db.rollback()
            flash(str(exc), 'error')
        except Exception as exc:
            db.rollback()
            flash('Operation failed: ' + str(exc), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_league_fees', league_id=selected_league_id))

    plan = None
    installments = []
    teams = []
    status_by_team = {}
    totals = {
        'installment_total_cents': 0,
        'deposit_display': None,
        'total_fee_display': None,
    }

    if league_id_int:
        cur = db.cursor()
        try:
            plan = _get_league_fee_plan(cur, league_id_int)
            if plan:
                totals['total_fee_display'] = plan['total_fee_display']
                totals['deposit_display'] = plan['deposit_display']

                cur.execute(
                    "SELECT installment_id, label, due_date, amount_cents, status, notes FROM league_fee_installments WHERE plan_id = %s ORDER BY due_date IS NULL, due_date",
                    (plan['plan_id'],)
                )
                rows = cur.fetchall()
                for row in rows:
                    installment_id, label, due_date, amount_cents, status, notes = row
                    if isinstance(due_date, datetime):
                        due_str = due_date.date().isoformat()
                    elif isinstance(due_date, date):
                        due_str = due_date.isoformat()
                    elif due_date:
                        due_str = str(due_date)
                    else:
                        due_str = ''
                    installments.append({
                        'installment_id': installment_id,
                        'label': label,
                        'due_date': due_str,
                        'amount_cents': amount_cents or 0,
                        'amount_display': _format_cents_to_decimal(amount_cents or 0),
                        'status': status or 'pending',
                        'notes': notes,
                    })
                    totals['installment_total_cents'] += amount_cents or 0

                cur.execute('SELECT team_id, name FROM teams WHERE league_id = %s ORDER BY name', (league_id_int,))
                teams = [{'team_id': row[0], 'name': row[1]} for row in cur.fetchall()]

                if installments and teams:
                    inst_ids = [inst['installment_id'] for inst in installments]
                    placeholders = ','.join(['%s'] * len(inst_ids))
                    cur.execute(
                        f'SELECT status_id, installment_id, team_id, status, amount_paid_cents, paid_at, notes FROM league_installment_status WHERE installment_id IN ({placeholders})',
                        tuple(inst_ids)
                    )
                    for status_row in cur.fetchall():
                        status_id, installment_id, team_id, status, amount_paid_cents, paid_at, notes = status_row
                        if isinstance(paid_at, datetime):
                            paid_date = paid_at.date().isoformat()
                        elif isinstance(paid_at, date):
                            paid_date = paid_at.isoformat()
                        elif paid_at:
                            paid_date = str(paid_at)
                        else:
                            paid_date = ''
                        status_by_team.setdefault(team_id, {})[installment_id] = {
                            'status_id': status_id,
                            'status': status or 'pending',
                            'amount_paid_cents': amount_paid_cents,
                            'amount_paid_display': _format_cents_to_decimal(amount_paid_cents) if amount_paid_cents is not None else None,
                            'paid_at': paid_date,
                            'notes': notes,
                        }
        finally:
            cur.close()

    totals['installment_total_display'] = _format_cents_to_decimal(totals['installment_total_cents']) if totals['installment_total_cents'] else None

    return render_template(
        'manage_league_fees.html',
        leagues=leagues,
        selected_league_id=selected_league_id,
        plan=plan,
        installments=installments,
        teams=teams,
        status_by_team=status_by_team,
        totals=totals,
    )




@admin_bp.route('/manage_league_rules', methods=['GET', 'POST'])
@admin_required
def manage_league_rules():
    db = get_db()
    try:
        _ensure_league_rules_table(db)
    except Exception as exc:
        flash('Failed to prepare league rules storage: ' + str(exc), 'error')
        return redirect(url_for('admin.admin_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action', 'save_rules')

        if action == 'add_custom_field':
            return _handle_add_custom_field(db)
        elif action == 'delete_custom_field':
            return _handle_delete_custom_field(db)
        elif action == 'save_rules':
            return _handle_save_rules(db)
        elif action == 'load_rules':
            return _handle_load_rules(db)

    # GET request - load all data
    cur = db.cursor()
    leagues = []
    rules = []
    custom_fields_by_league = {}

    try:
        cur.execute('SELECT league_id, name, sport FROM leagues ORDER BY name')
        leagues_raw = cur.fetchall()
        leagues = [{'id': row[0], 'name': row[1], 'sport': row[2] or 'soccer'} for row in leagues_raw]

        # Get all league rules with custom fields
        cur.execute("""
            SELECT lr.league_id, l.name, lr.points_win, lr.points_draw, lr.points_loss,
                   lr.tiebreakers, lr.substitution_limit, lr.foreign_player_limit, lr.notes
            FROM league_rules lr
            JOIN leagues l ON lr.league_id = l.league_id
            ORDER BY l.name
        """)
        rules = cur.fetchall()

        # Get custom fields for each league
        cur.execute("""
            SELECT league_id, field_id, field_name, field_type, field_value, display_order, is_required
            FROM league_rules_custom_fields
            ORDER BY league_id, display_order, field_name
        """)
        for row in cur.fetchall():
            league_id = row[0]
            if league_id not in custom_fields_by_league:
                custom_fields_by_league[league_id] = []
            custom_fields_by_league[league_id].append({
                'field_id': row[1],
                'field_name': row[2],
                'field_type': row[3],
                'field_value': row[4],
                'display_order': row[5],
                'is_required': row[6]
            })

    except Exception as e:
        flash(f'Error loading league rules data: {str(e)}', 'error')
    finally:
        cur.close()

    # Get sport configurations
    sport_configs = {league['id']: get_sport_config(league['sport']) for league in leagues}

    return render_template('manage_league_rules.html',
                         leagues=leagues,
                         rules=rules,
                         custom_fields_by_league=custom_fields_by_league,
                         sport_configs=sport_configs)


def _handle_add_custom_field(db):
    """Handle adding a new custom field"""
    try:
        league_id = int(request.form.get('league_id'))
        field_name = request.form.get('field_name', '').strip()
        field_type = request.form.get('field_type', 'text')
        field_value = request.form.get('field_value', '').strip()
        is_required = bool(request.form.get('is_required'))

        if not field_name:
            flash('Field name is required.', 'error')
            return redirect(url_for('admin.manage_league_rules'))

        cur = db.cursor()
        try:
            # Get next display order
            cur.execute(
                "SELECT COALESCE(MAX(display_order), 0) + 1 FROM league_rules_custom_fields WHERE league_id = %s",
                (league_id,)
            )
            display_order = cur.fetchone()[0]

            cur.execute("""
                INSERT INTO league_rules_custom_fields
                (league_id, field_name, field_type, field_value, display_order, is_required)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (league_id, field_name, field_type, field_value or None, display_order, is_required))

            db.commit()
            flash(f'Custom field "{field_name}" added successfully.', 'success')
        except Exception as exc:
            db.rollback()
            if 'unique' in str(exc).lower():
                flash(f'Field "{field_name}" already exists for this league.', 'error')
            else:
                flash(f'Failed to add custom field: {str(exc)}', 'error')
        finally:
            cur.close()

    except (ValueError, TypeError):
        flash('Invalid league selection.', 'error')

    return redirect(url_for('admin.manage_league_rules'))


def _handle_delete_custom_field(db):
    """Handle deleting a custom field"""
    try:
        field_id = int(request.form.get('field_id'))

        cur = db.cursor()
        try:
            cur.execute("DELETE FROM league_rules_custom_fields WHERE field_id = %s", (field_id,))

            if cur.rowcount > 0:
                db.commit()
                flash('Custom field deleted successfully.', 'success')
            else:
                flash('Custom field not found.', 'error')
        except Exception as exc:
            db.rollback()
            flash(f'Failed to delete custom field: {str(exc)}', 'error')
        finally:
            cur.close()

    except (ValueError, TypeError):
        flash('Invalid field selection.', 'error')

    return redirect(url_for('admin.manage_league_rules'))


def _handle_save_rules(db):
    """Handle saving league rules and custom fields"""
    cur = db.cursor()
    try:
        league_id_value = request.form.get('league_id')
        if not league_id_value:
            flash('Please select a league before saving rules.', 'error')
            return redirect(url_for('admin.manage_league_rules'))

        try:
            league_id = int(league_id_value)
        except (TypeError, ValueError):
            flash('Invalid league selection.', 'error')
            return redirect(url_for('admin.manage_league_rules'))

        def _optional_int(value):
            if value is None or value == '':
                return None
            return int(value)

        # Save basic league rules
        points_win = int(request.form.get('points_win', 3) or 3)
        points_draw = int(request.form.get('points_draw', 1) or 1)
        points_loss = int(request.form.get('points_loss', 0) or 0)
        tiebreakers = (request.form.get('tiebreakers') or '').strip() or None
        substitution_limit = _optional_int(request.form.get('substitution_limit'))
        foreign_player_limit = _optional_int(request.form.get('foreign_player_limit'))
        notes = (request.form.get('notes') or '').strip() or None

        cur.execute("""
            INSERT INTO league_rules (league_id, points_win, points_draw, points_loss,
                                    tiebreakers, substitution_limit, foreign_player_limit, notes, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (league_id) DO UPDATE SET
                points_win = EXCLUDED.points_win,
                points_draw = EXCLUDED.points_draw,
                points_loss = EXCLUDED.points_loss,
                tiebreakers = EXCLUDED.tiebreakers,
                substitution_limit = EXCLUDED.substitution_limit,
                foreign_player_limit = EXCLUDED.foreign_player_limit,
                notes = EXCLUDED.notes,
                updated_at = CURRENT_TIMESTAMP
        """, (league_id, points_win, points_draw, points_loss, tiebreakers, substitution_limit, foreign_player_limit, notes))

        # Save custom field values
        for key, value in request.form.items():
            if key.startswith('custom_field_'):
                field_id = key.replace('custom_field_', '')
                try:
                    field_id_int = int(field_id)
                    cur.execute(
                        "UPDATE league_rules_custom_fields SET field_value = %s WHERE field_id = %s",
                        (value.strip() or None, field_id_int)
                    )
                except (ValueError, TypeError):
                    continue

        db.commit()
        flash('League rules saved successfully.', 'success')

    except Exception as exc:
        db.rollback()
        flash('Failed to save league rules: ' + str(exc), 'error')
    finally:
        cur.close()

    return redirect(url_for('admin.manage_league_rules'))


def _handle_load_rules(db):
    """Handle loading rules for a specific league (AJAX endpoint)"""
    try:
        league_id = int(request.form.get('league_id'))

        cur = db.cursor()
        try:
            # Get basic rules
            cur.execute("""
                SELECT points_win, points_draw, points_loss, tiebreakers,
                       substitution_limit, foreign_player_limit, notes
                FROM league_rules WHERE league_id = %s
            """, (league_id,))

            rules_data = cur.fetchone()

            # Get custom fields
            cur.execute("""
                SELECT field_id, field_name, field_type, field_value, is_required
                FROM league_rules_custom_fields
                WHERE league_id = %s
                ORDER BY display_order, field_name
            """, (league_id,))

            custom_fields = cur.fetchall()

            return {
                'success': True,
                'rules': rules_data,
                'custom_fields': custom_fields
            }

        finally:
            cur.close()

    except (ValueError, TypeError):
        return {'success': False, 'error': 'Invalid league selection'}
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@admin_bp.route('/manage_seasons', methods=['GET', 'POST'])
@admin_required
def manage_seasons():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            season_id = request.form.get('season_id')
            league_id = request.form['league_id']
            year = request.form['year']

            if 'add' in request.form:
                cur.execute('INSERT INTO seasons (league_id, year) VALUES (%s, %s)', (league_id, year))
                flash('Season added successfully', 'success')
            elif 'edit' in request.form and season_id:
                cur.execute('UPDATE seasons SET league_id = %s, year = %s WHERE season_id = %s', (league_id, year, season_id))
                flash('Season updated successfully', 'success')
            elif 'delete' in request.form:
                season_id = request.form['deleteItemId']
                cur.execute('DELETE FROM seasons WHERE season_id = %s', (season_id,))
                flash('Season deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_seasons'))

    cur.execute('''
        SELECT s.season_id, s.league_id, s.year, l.name
        FROM seasons s
        JOIN leagues l ON s.league_id = l.league_id
    ''')
    seasons = cur.fetchall()
    cur.execute('SELECT league_id, name FROM leagues')
    leagues = cur.fetchall()
    cur.close()
    return render_template('manage_seasons.html', seasons=seasons, leagues=leagues)

@admin_bp.route('/manage_teams', methods=['GET', 'POST'])
@admin_required
def manage_teams():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save_team':
            return _handle_save_team(db)
        elif action == 'delete_team':
            return _handle_delete_team(db)
        elif action == 'upload_logo':
            return _handle_upload_logo(db)

        # Fallback for old-style form submissions
        try:
            team_id = request.form.get('team_id')
            name = request.form.get('name', '')
            founded_year = request.form.get('founded_year')
            stadium_id = request.form.get('stadium_id')
            league_id = request.form.get('league_id')
            coach_id = request.form.get('coach_id')

            cur = db.cursor()

            if 'add' in request.form:
                cur.execute('INSERT INTO teams (name, founded_year, stadium_id, league_id, coach_id) VALUES (%s, %s, %s, %s, %s)',
                            (name, founded_year, stadium_id, league_id, coach_id))
                flash('Team added successfully', 'success')
            elif 'edit' in request.form and team_id:
                cur.execute('UPDATE teams SET name = %s, founded_year = %s, stadium_id = %s, league_id = %s, coach_id = %s WHERE team_id = %s',
                            (name, founded_year, stadium_id, league_id, coach_id, team_id))
                flash('Team updated successfully', 'success')
            elif 'delete' in request.form and team_id:
                cur.execute('DELETE FROM teams WHERE team_id = %s', (team_id,))
                flash('Team deleted successfully', 'success')
            db.commit()
            cur.close()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')

        return redirect(url_for('admin.manage_teams'))

    # Get teams with extended information
    cur.execute("""
        SELECT
            t.team_id,
            t.name,
            t.founded_year,
            t.cresturl,
            l.name as league_name,
            s.name as stadium_name,
            c.name as coach_name,
            t.league_id,
            t.stadium_id,
            t.coach_id,
            COUNT(p.player_id) as player_count
        FROM teams t
        LEFT JOIN leagues l ON t.league_id = l.league_id
        LEFT JOIN stadiums s ON t.stadium_id = s.stadium_id
        LEFT JOIN coaches c ON t.coach_id = c.coach_id
        LEFT JOIN players p ON t.team_id = p.team_id
        GROUP BY t.team_id, t.name, t.founded_year, t.cresturl,
                 l.name, s.name, c.name, t.league_id, t.stadium_id, t.coach_id
        ORDER BY t.name
    """)

    teams_raw = cur.fetchall()

    # Convert to list of dictionaries for easier template access
    teams = []
    for team in teams_raw:
        team_dict = {
            'team_id': team[0],
            'name': team[1],
            'founded_year': team[2],
            'cresturl': team[3],
            'league_name': team[4],
            'stadium_name': team[5],
            'coach_name': team[6],
            'league_id': team[7],
            'stadium_id': team[8],
            'coach_id': team[9],
            'player_count': team[10],
            'recent_players': []
        }

        # Get recent players for this team
        cur.execute("""
            SELECT name, position
            FROM players
            WHERE team_id = %s
            ORDER BY name
            LIMIT 5
        """, (team[0],))

        players = cur.fetchall()
        team_dict['recent_players'] = [{'name': p[0], 'position': p[1]} for p in players]

        teams.append(team_dict)

    # Get reference data for dropdowns
    cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = cur.fetchall()

    cur.execute('SELECT stadium_id, name FROM stadiums ORDER BY name')
    stadiums = cur.fetchall()

    cur.execute('SELECT coach_id, name FROM coaches ORDER BY name')
    coaches = cur.fetchall()

    # Calculate summary statistics
    total_players = sum(team['player_count'] for team in teams)
    avg_squad_size = total_players / len(teams) if teams else 0

    cur.close()

    return render_template('manage_teams.html',
                         teams=teams,
                         leagues=leagues,
                         stadiums=stadiums,
                         coaches=coaches,
                         total_players=total_players,
                         avg_squad_size=avg_squad_size)

def _handle_save_team(db):
    """Handle team creation/editing"""
    try:
        team_id = request.form.get('team_id')
        name = request.form.get('name', '').strip()
        founded_year = request.form.get('founded_year')
        league_id = request.form.get('league_id')
        stadium_id = request.form.get('stadium_id')
        coach_id = request.form.get('coach_id')
        cresturl = request.form.get('cresturl', '').strip()

        if not name or not league_id:
            flash('Team name and league are required.', 'error')
            return redirect(url_for('admin.manage_teams'))

        # Convert empty strings to None for optional fields
        founded_year = int(founded_year) if founded_year else None
        stadium_id = int(stadium_id) if stadium_id else None
        coach_id = int(coach_id) if coach_id else None
        cresturl = cresturl if cresturl else None

        cur = db.cursor()

        if team_id:
            # Update existing team
            cur.execute("""
                UPDATE teams
                SET name = %s, founded_year = %s, league_id = %s, stadium_id = %s, coach_id = %s, cresturl = %s
                WHERE team_id = %s
            """, (name, founded_year, league_id, stadium_id, coach_id, cresturl, team_id))
            flash(f'Team "{name}" updated successfully!', 'success')
        else:
            # Create new team
            cur.execute("""
                INSERT INTO teams (name, founded_year, league_id, stadium_id, coach_id, cresturl)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (name, founded_year, league_id, stadium_id, coach_id, cresturl))
            flash(f'Team "{name}" created successfully!', 'success')

        db.commit()
        cur.close()

    except Exception as e:
        db.rollback()
        flash(f'Error saving team: {str(e)}', 'error')

    return redirect(url_for('admin.manage_teams'))

def _handle_delete_team(db):
    """Handle team deletion"""
    try:
        team_id = request.form.get('team_id')

        if not team_id:
            flash('Invalid team ID.', 'error')
            return redirect(url_for('admin.manage_teams'))

        cur = db.cursor()

        # Get team name for confirmation message
        cur.execute('SELECT name FROM teams WHERE team_id = %s', (team_id,))
        team_row = cur.fetchone()

        if not team_row:
            flash('Team not found.', 'error')
            return redirect(url_for('admin.manage_teams'))

        team_name = team_row[0]

        # Delete team (cascade should handle related records)
        cur.execute('DELETE FROM teams WHERE team_id = %s', (team_id,))

        db.commit()
        cur.close()

        flash(f'Team "{team_name}" deleted successfully.', 'success')

    except Exception as e:
        db.rollback()
        flash(f'Error deleting team: {str(e)}', 'error')

    return redirect(url_for('admin.manage_teams'))

def _handle_upload_logo(db):
    """Handle team logo upload"""
    try:
        team_id = request.form.get('team_id')
        logo_url = request.form.get('logo_url', '').strip()

        if not team_id:
            flash('Invalid team ID.', 'error')
            return redirect(url_for('admin.manage_teams'))

        # Handle file upload
        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename:
            # Here you would typically upload to a cloud service or local storage
            # For now, we'll just use the URL field
            flash('File upload functionality would be implemented here. Please use URL instead.', 'info')
            return redirect(url_for('admin.manage_teams'))

        # Handle URL
        if logo_url:
            cur = db.cursor()
            cur.execute('UPDATE teams SET cresturl = %s WHERE team_id = %s', (logo_url, team_id))
            db.commit()
            cur.close()
            flash('Team logo updated successfully!', 'success')
        else:
            flash('Please provide a logo URL.', 'error')

    except Exception as e:
        db.rollback()
        flash(f'Error updating logo: {str(e)}', 'error')

    return redirect(url_for('admin.manage_teams'))

# Team Management API Routes
@admin_bp.route('/get_team/<int:team_id>')
@admin_required
def get_team(team_id):
    """API endpoint to get team data for editing"""
    try:
        db = get_db()
        cur = db.cursor()

        cur.execute("""
            SELECT team_id, name, founded_year, league_id, stadium_id, coach_id, cresturl
            FROM teams WHERE team_id = %s
        """, (team_id,))

        team_row = cur.fetchone()
        cur.close()

        if not team_row:
            return jsonify({'success': False, 'message': 'Team not found'})

        team_data = {
            'team_id': team_row[0],
            'name': team_row[1],
            'founded_year': team_row[2],
            'league_id': team_row[3],
            'stadium_id': team_row[4],
            'coach_id': team_row[5],
            'cresturl': team_row[6]
        }

        return jsonify({'success': True, 'team': team_data})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@admin_bp.route('/get_team_roster/<int:team_id>')
@admin_required
def get_team_roster(team_id):
    """API endpoint to get team roster and available players"""
    try:
        db = get_db()
        cur = db.cursor()

        # Get team name
        cur.execute('SELECT name FROM teams WHERE team_id = %s', (team_id,))
        team_row = cur.fetchone()
        if not team_row:
            return jsonify({'success': False, 'message': 'Team not found'})

        team_name = team_row[0]

        # Get current roster
        cur.execute("""
            SELECT player_id, name, position, nationality
            FROM players
            WHERE team_id = %s
            ORDER BY position, name
        """, (team_id,))

        roster = []
        for row in cur.fetchall():
            roster.append({
                'player_id': row[0],
                'name': row[1],
                'position': row[2],
                'nationality': row[3]
            })

        # Get available players (no team assigned)
        cur.execute("""
            SELECT player_id, name, position, nationality
            FROM players
            WHERE team_id IS NULL
            ORDER BY position, name
        """)

        available_players = []
        for row in cur.fetchall():
            available_players.append({
                'player_id': row[0],
                'name': row[1],
                'position': row[2],
                'nationality': row[3]
            })

        cur.close()

        return jsonify({
            'success': True,
            'team_name': team_name,
            'roster': roster,
            'available_players': available_players
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@admin_bp.route('/update_team_roster', methods=['POST'])
@admin_required
def update_team_roster():
    """API endpoint to update team roster"""
    try:
        data = request.get_json()
        team_id = data.get('team_id')
        changes = data.get('changes', {})

        if not team_id:
            return jsonify({'success': False, 'message': 'Team ID required'})

        db = get_db()
        cur = db.cursor()

        # Add players to team
        players_to_add = changes.get('add', [])
        for player_id in players_to_add:
            cur.execute('UPDATE players SET team_id = %s WHERE player_id = %s', (team_id, player_id))

        # Remove players from team
        players_to_remove = changes.get('remove', [])
        for player_id in players_to_remove:
            cur.execute('UPDATE players SET team_id = NULL WHERE player_id = %s', (player_id,))

        db.commit()
        cur.close()

        return jsonify({'success': True, 'message': 'Roster updated successfully'})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': str(e)})


@admin_bp.route('/manage_coaches', methods=['GET', 'POST'])
@admin_required
def manage_coaches():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            coach_id = request.form.get('coach_id')
            name = request.form['name']
            nationality = request.form['nationality']
            team_id = request.form['team_id']

            if 'add' in request.form:
                cur.execute('INSERT INTO coaches (name, nationality, team_id) VALUES (%s, %s, %s)', 
                            (name, nationality, team_id))
                flash('Coach added successfully', 'success')
            elif 'submit' in request.form and coach_id:
                cur.execute('UPDATE coaches SET name = %s, nationality = %s, team_id = %s WHERE coach_id = %s', 
                            (name, nationality, team_id, coach_id))
                flash('Coach updated successfully', 'success')
            elif 'delete' in request.form:
                coach_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM coaches WHERE coach_id = %s', (coach_id,))
                flash('Coach deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_coaches'))

    cur.execute('''
        SELECT c.coach_id, c.name, c.team_id, c.nationality, t.name AS team_name
        FROM coaches c
        JOIN teams t ON c.team_id = t.team_id
    ''')
    coaches = cur.fetchall()
    cur.execute('SELECT team_id, name FROM teams')
    teams = cur.fetchall()
    cur.close()
    return render_template('manage_coaches.html', coaches=coaches, teams=teams)



@admin_bp.route('/manage_players', methods=['GET', 'POST'])
@admin_required
def manage_players():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            player_id = request.form.get('player_id')
            team_id = request.form['team_id']
            name = request.form['name']
            position = request.form['position']
            date_of_birth = request.form['date_of_birth']
            nationality = request.form['nationality']
            email = request.form.get('email', '')
            weight_kg = request.form.get('weight_kg') or None
            height_cm = request.form.get('height_cm') or None
            sport = request.form.get('sport', 'Football')

            if 'submit' in request.form:
                if player_id:
                    cur.execute('''UPDATE players SET team_id = %s, name = %s, position = %s, date_of_birth = %s,
                                   nationality = %s, email = %s, weight_kg = %s, height_cm = %s, sport = %s
                                   WHERE player_id = %s''',
                                (team_id, name, position, date_of_birth, nationality, email, weight_kg, height_cm, sport, player_id))
                    flash('Player updated successfully', 'success')
                else:
                    cur.execute('''INSERT INTO players (team_id, name, position, date_of_birth, nationality, email, weight_kg, height_cm, sport)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (team_id, name, position, date_of_birth, nationality, email, weight_kg, height_cm, sport))
                    flash('Player added successfully', 'success')
            elif 'delete' in request.form:
                player_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM players WHERE player_id = %s', (player_id,))
                flash('Player deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_players'))

    cur.execute('''SELECT p.player_id, t.name AS team, p.name, p.position, p.date_of_birth, p.nationality, p.team_id,
                          p.email, p.weight_kg, p.height_cm, p.sport
                   FROM players p JOIN teams t ON p.team_id = t.team_id''')
    players = cur.fetchall()
    cur.execute('SELECT team_id, name FROM teams')
    teams = cur.fetchall()
    cur.execute('SELECT name, positions FROM sports ORDER BY name')
    sports = cur.fetchall()
    cur.close()
    return render_template('manage_players.html', players=players, teams=teams, sports=sports)


@admin_bp.route('/manage_news', methods=['GET', 'POST'])
@admin_required
def manage_news():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            news_id = request.form.get('news_id')
            title = request.form['title']
            content = request.form['content']
            summary = request.form.get('summary', '')
            category = request.form.get('category', 'General')
            status = request.form.get('status', 'draft')
            is_featured = request.form.get('is_featured') == 'on'
            show_on_homepage = request.form.get('show_on_homepage') == 'on'
            image_url = request.form.get('image_url', '')

            if 'submit' in request.form:
                if news_id:
                    # Update existing news
                    cur.execute('''UPDATE news SET title = %s, content = %s, summary = %s, category = %s,
                                   status = %s, is_featured = %s, show_on_homepage = %s, image_url = %s,
                                   updated_at = CURRENT_TIMESTAMP,
                                   published_at = CASE WHEN status = 'published' AND published_at IS NULL
                                                      THEN CURRENT_TIMESTAMP ELSE published_at END
                                   WHERE news_id = %s''',
                                (title, content, summary, category, status, is_featured, show_on_homepage, image_url, news_id))
                    flash('News article updated successfully', 'success')
                else:
                    # Create new news
                    published_at = 'CURRENT_TIMESTAMP' if status == 'published' else 'NULL'
                    cur.execute('''INSERT INTO news (title, content, summary, category, status, is_featured,
                                   show_on_homepage, image_url, published_at)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (title, content, summary, category, status, is_featured, show_on_homepage, image_url,
                                 'now()' if status == 'published' else None))
                    flash('News article created successfully', 'success')
            elif 'delete' in request.form:
                news_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM news WHERE news_id = %s', (news_id,))
                flash('News article deleted successfully', 'success')

            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_news'))

    # Get all news articles
    cur.execute('''SELECT news_id, title, summary, content, category, status, is_featured, show_on_homepage,
                   created_at, published_at, author_id FROM news ORDER BY created_at DESC''')
    news_articles = cur.fetchall()

    cur.close()
    return render_template('manage_news.html', news_articles=news_articles)


@admin_bp.route('/manage_sponsors', methods=['GET', 'POST'])
@admin_required
def manage_sponsors():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            if 'submit_sponsor' in request.form:
                sponsor_id = request.form.get('sponsor_id')
                company_name = request.form['company_name']
                contact_person = request.form.get('contact_person', '')
                email = request.form.get('email', '')
                phone = request.form.get('phone', '')
                website = request.form.get('website', '')
                address = request.form.get('address', '')
                industry = request.form.get('industry', '')
                company_size = request.form.get('company_size', '')
                annual_revenue_range = request.form.get('annual_revenue_range', '')
                status = request.form.get('status', 'prospect')
                priority = request.form.get('priority', 'medium')
                notes = request.form.get('notes', '')
                first_contact_date = request.form.get('first_contact_date') or None
                last_contact_date = request.form.get('last_contact_date') or None

                if sponsor_id:
                    # Update existing sponsor
                    cur.execute('''UPDATE sponsors SET company_name = %s, contact_person = %s, email = %s,
                                   phone = %s, website = %s, address = %s, industry = %s, company_size = %s,
                                   annual_revenue_range = %s, status = %s, priority = %s, notes = %s,
                                   first_contact_date = %s, last_contact_date = %s, updated_at = CURRENT_TIMESTAMP
                                   WHERE sponsor_id = %s''',
                                (company_name, contact_person, email, phone, website, address, industry,
                                 company_size, annual_revenue_range, status, priority, notes,
                                 first_contact_date, last_contact_date, sponsor_id))
                    flash('Sponsor updated successfully', 'success')
                else:
                    # Create new sponsor
                    cur.execute('''INSERT INTO sponsors (company_name, contact_person, email, phone, website,
                                   address, industry, company_size, annual_revenue_range, status, priority,
                                   notes, first_contact_date, last_contact_date, created_by)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (company_name, contact_person, email, phone, website, address, industry,
                                 company_size, annual_revenue_range, status, priority, notes,
                                 first_contact_date, last_contact_date, current_user.user_id if current_user.is_authenticated else None))
                    flash('Sponsor added successfully', 'success')

            elif 'submit_deal' in request.form:
                deal_id = request.form.get('deal_id')
                sponsor_id = request.form['deal_sponsor_id']
                deal_name = request.form['deal_name']
                deal_type = request.form.get('deal_type', 'other')
                sponsorship_level = request.form.get('sponsorship_level', 'bronze')
                deal_value = request.form.get('deal_value', '0')
                deal_value_cents = int(float(deal_value) * 100) if deal_value else 0
                currency = request.form.get('currency', 'USD')
                deal_status = request.form.get('deal_status', 'prospect')
                start_date = request.form.get('start_date') or None
                end_date = request.form.get('end_date') or None
                benefits_provided = request.form.get('benefits_provided', '')
                deliverables = request.form.get('deliverables', '')
                payment_terms = request.form.get('payment_terms', '')
                notes = request.form.get('deal_notes', '')

                if deal_id:
                    # Update existing deal
                    cur.execute('''UPDATE sponsorship_deals SET sponsor_id = %s, deal_name = %s, deal_type = %s,
                                   sponsorship_level = %s, deal_value_cents = %s, currency = %s, deal_status = %s,
                                   start_date = %s, end_date = %s, benefits_provided = %s, deliverables = %s,
                                   payment_terms = %s, notes = %s, updated_at = CURRENT_TIMESTAMP
                                   WHERE deal_id = %s''',
                                (sponsor_id, deal_name, deal_type, sponsorship_level, deal_value_cents, currency,
                                 deal_status, start_date, end_date, benefits_provided, deliverables,
                                 payment_terms, notes, deal_id))
                    flash('Deal updated successfully', 'success')
                else:
                    # Create new deal
                    cur.execute('''INSERT INTO sponsorship_deals (sponsor_id, deal_name, deal_type, sponsorship_level,
                                   deal_value_cents, currency, deal_status, start_date, end_date, benefits_provided,
                                   deliverables, payment_terms, notes, created_by)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (sponsor_id, deal_name, deal_type, sponsorship_level, deal_value_cents, currency,
                                 deal_status, start_date, end_date, benefits_provided, deliverables,
                                 payment_terms, notes, current_user.user_id if current_user.is_authenticated else None))
                    flash('Deal added successfully', 'success')

            elif 'delete_sponsor' in request.form:
                sponsor_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM sponsors WHERE sponsor_id = %s', (sponsor_id,))
                flash('Sponsor deleted successfully', 'success')

            elif 'delete_deal' in request.form:
                deal_id = request.form['deleteDealId']
                cur.execute('DELETE FROM sponsorship_deals WHERE deal_id = %s', (deal_id,))
                flash('Deal deleted successfully', 'success')

            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_sponsors'))

    # Get all sponsors with deal counts
    cur.execute('''SELECT s.sponsor_id, s.company_name, s.contact_person, s.email, s.phone, s.website,
                   s.address, s.industry, s.company_size, s.annual_revenue_range, s.status, s.priority,
                   s.notes, s.first_contact_date, s.last_contact_date, s.created_at,
                   COUNT(d.deal_id) as deal_count,
                   COALESCE(SUM(CASE WHEN d.deal_status = 'active' THEN d.deal_value_cents ELSE 0 END), 0) as active_value
                   FROM sponsors s
                   LEFT JOIN sponsorship_deals d ON s.sponsor_id = d.sponsor_id
                   GROUP BY s.sponsor_id
                   ORDER BY s.priority DESC, s.last_contact_date DESC''')
    sponsors = cur.fetchall()

    # Get all deals with sponsor info
    cur.execute('''SELECT d.deal_id, d.sponsor_id, s.company_name, d.deal_name, d.deal_type,
                   d.sponsorship_level, d.deal_value_cents, d.currency, d.deal_status,
                   d.start_date, d.end_date, d.benefits_provided, d.deliverables,
                   d.payment_terms, d.notes, d.created_at
                   FROM sponsorship_deals d
                   JOIN sponsors s ON d.sponsor_id = s.sponsor_id
                   ORDER BY d.deal_status, d.start_date DESC''')
    deals = cur.fetchall()

    cur.close()
    return render_template('manage_sponsors.html', sponsors=sponsors, deals=deals)


@admin_bp.route('/export_sponsors')
@admin_required
def export_sponsors():
    """Export sponsors data to Excel"""
    try:
        import pandas as pd
        from flask import send_file
        import tempfile
        import os
        from datetime import datetime

        db = get_db()
        cur = db.cursor()

        # Export sponsors
        cur.execute('''SELECT s.sponsor_id, s.company_name, s.contact_person, s.email, s.phone, s.website,
                       s.address, s.industry, s.company_size, s.annual_revenue_range, s.status, s.priority,
                       s.notes, s.first_contact_date, s.last_contact_date, s.created_at,
                       COUNT(d.deal_id) as deal_count,
                       COALESCE(SUM(CASE WHEN d.deal_status = 'active' THEN d.deal_value_cents ELSE 0 END), 0) as active_value
                       FROM sponsors s
                       LEFT JOIN sponsorship_deals d ON s.sponsor_id = d.sponsor_id
                       GROUP BY s.sponsor_id
                       ORDER BY s.company_name''')
        sponsors_data = cur.fetchall()

        # Export deals
        cur.execute('''SELECT d.deal_id, s.company_name as sponsor, d.deal_name, d.deal_type,
                       d.sponsorship_level, d.deal_value_cents, d.currency, d.deal_status,
                       d.start_date, d.end_date, d.benefits_provided, d.deliverables,
                       d.payment_terms, d.notes, d.created_at
                       FROM sponsorship_deals d
                       JOIN sponsors s ON d.sponsor_id = s.sponsor_id
                       ORDER BY s.company_name, d.deal_name''')
        deals_data = cur.fetchall()

        cur.close()

        # Create DataFrames
        sponsors_columns = ['ID', 'Company Name', 'Contact Person', 'Email', 'Phone', 'Website',
                           'Address', 'Industry', 'Company Size', 'Revenue Range', 'Status', 'Priority',
                           'Notes', 'First Contact Date', 'Last Contact Date', 'Created At',
                           'Deal Count', 'Active Value (cents)']

        deals_columns = ['Deal ID', 'Sponsor', 'Deal Name', 'Deal Type', 'Level', 'Value (cents)',
                        'Currency', 'Status', 'Start Date', 'End Date', 'Benefits', 'Deliverables',
                        'Payment Terms', 'Notes', 'Created At']

        sponsors_df = pd.DataFrame(sponsors_data, columns=sponsors_columns)
        deals_df = pd.DataFrame(deals_data, columns=deals_columns)

        # Format dates
        for date_col in ['First Contact Date', 'Last Contact Date', 'Created At']:
            if date_col in sponsors_df.columns:
                sponsors_df[date_col] = pd.to_datetime(sponsors_df[date_col]).dt.strftime('%Y-%m-%d')

        for date_col in ['Start Date', 'End Date', 'Created At']:
            if date_col in deals_df.columns:
                deals_df[date_col] = pd.to_datetime(deals_df[date_col]).dt.strftime('%Y-%m-%d')

        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                sponsors_df.to_excel(writer, sheet_name='Sponsors', index=False)
                deals_df.to_excel(writer, sheet_name='Deals', index=False)

            filename = f"sponsorship_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return send_file(tmp.name, as_attachment=True, download_name=filename,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('admin.manage_sponsors'))


@admin_bp.route('/import_sponsors', methods=['POST'])
@admin_required
def import_sponsors():
    """Import sponsors data from Excel"""
    try:
        import pandas as pd
        from werkzeug.utils import secure_filename

        if 'excel_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('admin.manage_sponsors'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('admin.manage_sponsors'))

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('admin.manage_sponsors'))

        # Read Excel file
        try:
            # Try to read sponsors sheet
            sponsors_df = pd.read_excel(file, sheet_name='Sponsors')
        except:
            # Fallback to first sheet
            sponsors_df = pd.read_excel(file)

        db = get_db()
        cur = db.cursor()

        imported_count = 0
        skipped_count = 0

        for _, row in sponsors_df.iterrows():
            try:
                company_name = str(row.get('Company Name', '')).strip()
                if not company_name:
                    skipped_count += 1
                    continue

                # Check if sponsor already exists
                cur.execute('SELECT sponsor_id FROM sponsors WHERE company_name = %s', (company_name,))
                existing = cur.fetchone()

                contact_person = str(row.get('Contact Person', '')) if pd.notna(row.get('Contact Person')) else None
                email = str(row.get('Email', '')) if pd.notna(row.get('Email')) else None
                phone = str(row.get('Phone', '')) if pd.notna(row.get('Phone')) else None
                website = str(row.get('Website', '')) if pd.notna(row.get('Website')) else None
                address = str(row.get('Address', '')) if pd.notna(row.get('Address')) else None
                industry = str(row.get('Industry', '')) if pd.notna(row.get('Industry')) else None
                company_size = str(row.get('Company Size', '')) if pd.notna(row.get('Company Size')) else None
                revenue_range = str(row.get('Revenue Range', '')) if pd.notna(row.get('Revenue Range')) else None
                status = str(row.get('Status', 'prospect')).lower() if pd.notna(row.get('Status')) else 'prospect'
                priority = str(row.get('Priority', 'medium')).lower() if pd.notna(row.get('Priority')) else 'medium'
                notes = str(row.get('Notes', '')) if pd.notna(row.get('Notes')) else None

                # Handle dates
                first_contact_date = None
                if pd.notna(row.get('First Contact Date')):
                    try:
                        first_contact_date = pd.to_datetime(row.get('First Contact Date')).strftime('%Y-%m-%d')
                    except:
                        pass

                last_contact_date = None
                if pd.notna(row.get('Last Contact Date')):
                    try:
                        last_contact_date = pd.to_datetime(row.get('Last Contact Date')).strftime('%Y-%m-%d')
                    except:
                        pass

                if existing:
                    # Update existing sponsor
                    cur.execute('''UPDATE sponsors SET contact_person = %s, email = %s, phone = %s,
                                   website = %s, address = %s, industry = %s, company_size = %s,
                                   annual_revenue_range = %s, status = %s, priority = %s, notes = %s,
                                   first_contact_date = %s, last_contact_date = %s, updated_at = CURRENT_TIMESTAMP
                                   WHERE company_name = %s''',
                                (contact_person, email, phone, website, address, industry,
                                 company_size, revenue_range, status, priority, notes,
                                 first_contact_date, last_contact_date, company_name))
                else:
                    # Create new sponsor
                    cur.execute('''INSERT INTO sponsors (company_name, contact_person, email, phone, website,
                                   address, industry, company_size, annual_revenue_range, status, priority,
                                   notes, first_contact_date, last_contact_date, created_by)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (company_name, contact_person, email, phone, website, address, industry,
                                 company_size, revenue_range, status, priority, notes,
                                 first_contact_date, last_contact_date, current_user.user_id if current_user.is_authenticated else None))

                imported_count += 1

            except Exception as e:
                print(f"Error processing row: {str(e)}")
                skipped_count += 1
                continue

        db.commit()
        cur.close()

        flash(f'Import completed: {imported_count} sponsors imported/updated, {skipped_count} skipped', 'success')

    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')

    return redirect(url_for('admin.manage_sponsors'))


@admin_bp.route('/export_players')
@admin_required
def export_players():
    """Export players data to Excel"""
    try:
        import pandas as pd
        from flask import send_file
        import tempfile
        from datetime import datetime

        db = get_db()
        cur = db.cursor()

        # Export players with team info
        cur.execute('''SELECT p.player_id, p.name, p.email, p.date_of_birth, p.position,
                       p.nationality, p.weight_kg, p.height_cm, p.sport,
                       t.name as team_name, s.year as season_year, l.name as league_name
                       FROM players p
                       LEFT JOIN teams t ON p.team_id = t.team_id
                       LEFT JOIN seasons s ON t.league_id = s.league_id
                       LEFT JOIN leagues l ON t.league_id = l.league_id
                       ORDER BY p.name''')
        players_data = cur.fetchall()

        cur.close()

        # Create DataFrame
        columns = ['Player ID', 'Name', 'Email', 'Date of Birth', 'Position',
                  'Nationality', 'Weight (kg)', 'Height (cm)', 'Sport',
                  'Team', 'Season Year', 'League']

        players_df = pd.DataFrame(players_data, columns=columns)

        # Format dates
        if 'Date of Birth' in players_df.columns:
            players_df['Date of Birth'] = pd.to_datetime(players_df['Date of Birth']).dt.strftime('%Y-%m-%d')

        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                players_df.to_excel(writer, sheet_name='Players', index=False)

            filename = f"players_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return send_file(tmp.name, as_attachment=True, download_name=filename,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('admin.manage_players'))


@admin_bp.route('/import_players', methods=['POST'])
@admin_required
def import_players():
    """Import players data from Excel"""
    try:
        import pandas as pd

        if 'excel_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('admin.manage_players'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('admin.manage_players'))

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('admin.manage_players'))

        # Read Excel file
        try:
            # Try to read players sheet
            players_df = pd.read_excel(file, sheet_name='Players')
        except:
            # Fallback to first sheet
            players_df = pd.read_excel(file)

        db = get_db()
        cur = db.cursor()

        imported_count = 0
        skipped_count = 0

        for _, row in players_df.iterrows():
            try:
                name = str(row.get('Name', '')).strip()
                if not name:
                    skipped_count += 1
                    continue

                email = str(row.get('Email', '')) if pd.notna(row.get('Email')) else None
                phone = str(row.get('Phone', '')) if pd.notna(row.get('Phone')) else None
                position = str(row.get('Position', '')) if pd.notna(row.get('Position')) else None
                jersey_number = None
                if pd.notna(row.get('Jersey Number')):
                    try:
                        jersey_number = int(row.get('Jersey Number'))
                    except:
                        pass

                emergency_contact_name = str(row.get('Emergency Contact Name', '')) if pd.notna(row.get('Emergency Contact Name')) else None
                emergency_contact_phone = str(row.get('Emergency Contact Phone', '')) if pd.notna(row.get('Emergency Contact Phone')) else None

                # Handle date of birth
                date_of_birth = None
                if pd.notna(row.get('Date of Birth')):
                    try:
                        date_of_birth = pd.to_datetime(row.get('Date of Birth')).strftime('%Y-%m-%d')
                    except:
                        pass

                # Find team by name if provided
                team_id = None
                team_name = str(row.get('Team', '')) if pd.notna(row.get('Team')) else None
                if team_name:
                    cur.execute('SELECT team_id FROM teams WHERE team_name = %s LIMIT 1', (team_name,))
                    team_result = cur.fetchone()
                    if team_result:
                        team_id = team_result[0]

                # Check if player already exists
                cur.execute('SELECT player_id FROM players WHERE name = %s AND email = %s', (name, email))
                existing = cur.fetchone()

                if existing:
                    # Update existing player
                    cur.execute('''UPDATE players SET phone = %s, date_of_birth = %s, position = %s,
                                   jersey_number = %s, emergency_contact_name = %s, emergency_contact_phone = %s,
                                   team_id = %s WHERE player_id = %s''',
                                (phone, date_of_birth, position, jersey_number, emergency_contact_name,
                                 emergency_contact_phone, team_id, existing[0]))
                else:
                    # Create new player
                    cur.execute('''INSERT INTO players (name, email, phone, date_of_birth, position,
                                   jersey_number, emergency_contact_name, emergency_contact_phone, team_id)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (name, email, phone, date_of_birth, position, jersey_number,
                                 emergency_contact_name, emergency_contact_phone, team_id))

                imported_count += 1

            except Exception as e:
                print(f"Error processing row: {str(e)}")
                skipped_count += 1
                continue

        db.commit()
        cur.close()

        flash(f'Import completed: {imported_count} players imported/updated, {skipped_count} skipped', 'success')

    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')

    return redirect(url_for('admin.manage_players'))


@admin_bp.route('/generate_fixtures', methods=['GET', 'POST'])
@admin_required
def generate_fixtures():
    db = get_db()

    if request.method == 'POST':
        cur = db.cursor()
        try:
            league_id_value = request.form.get('league_id')
            season_id_value = request.form.get('season_id')
            start_date_value = request.form.get('start_date')
            interval_value = request.form.get('interval_days') or '7'
            double_round = 'double_round' in request.form
            shuffle_teams = 'shuffle' in request.form

            try:
                league_id = int(league_id_value)
                season_id = int(season_id_value)
            except (TypeError, ValueError):
                flash('Please choose a league and season before generating fixtures.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            try:
                interval_days = int(interval_value)
                if interval_days < 1:
                    raise ValueError
            except (TypeError, ValueError):
                flash('Interval between rounds must be at least one day.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            try:
                start_date = datetime.strptime(start_date_value, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                flash('Please provide a valid start date.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            cur.execute('SELECT league_id FROM seasons WHERE season_id = %s', (season_id,))
            season_row = cur.fetchone()
            if not season_row:
                flash('Selected season could not be found.', 'error')
                return redirect(url_for('admin.generate_fixtures'))
            if season_row[0] != league_id:
                flash('Selected season does not belong to that league.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            cur.execute('SELECT team_id FROM teams WHERE league_id = %s ORDER BY name', (league_id,))
            team_rows = cur.fetchall()
            team_ids = [row[0] for row in team_rows]
            if len(team_ids) < 2:
                flash('At least two teams are required to generate fixtures.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            if shuffle_teams:
                random.shuffle(team_ids)

            schedule = _round_robin_schedule(team_ids, double_round=double_round)
            if not schedule:
                flash('No fixtures generated because there were not enough teams.', 'error')
                return redirect(url_for('admin.generate_fixtures'))

            cur.execute('SELECT home_team_id, away_team_id FROM matches WHERE league_id = %s AND season_id = %s', (league_id, season_id))
            existing_pairs = {(row[0], row[1]) for row in cur.fetchall()}

            created = 0
            skipped = 0
            for round_index, pairings in enumerate(schedule):
                round_datetime = datetime.combine(start_date + timedelta(days=interval_days * round_index), datetime.min.time())
                for home_team_id, away_team_id in pairings:
                    if (home_team_id, away_team_id) in existing_pairs:
                        skipped += 1
                        continue
                    cur.execute(
                        'INSERT INTO matches (utc_date, home_team_id, away_team_id, season_id, league_id) VALUES (%s, %s, %s, %s, %s)',
                        (round_datetime, home_team_id, away_team_id, season_id, league_id)
                    )
                    existing_pairs.add((home_team_id, away_team_id))
                    created += 1

            if created:
                db.commit()
            else:
                db.rollback()

            if created:
                message = f"Generated {created} fixture{'s' if created != 1 else ''}."
                if skipped:
                    message += f" Skipped {skipped} existing pairing{'s' if skipped != 1 else ''}."
                flash(message, 'success')
            else:
                info_message = 'No new fixtures were generated; an identical schedule already exists.' if skipped else 'No fixtures were generated.'
                flash(info_message, 'info')
        except Exception as exc:
            db.rollback()
            flash('Failed to generate fixtures: ' + str(exc), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.generate_fixtures'))

    cur = db.cursor()
    try:
        cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
        leagues = cur.fetchall()
        cur.execute('SELECT season_id, year, league_id FROM seasons ORDER BY year DESC, season_id DESC')
        seasons = cur.fetchall()
    finally:
        cur.close()
    return render_template('generate_fixtures.html', leagues=leagues, seasons=seasons)


@admin_bp.route('/download_bulk_template')
@admin_required
def download_bulk_template():
    """Generate and download Excel template for bulk match upload"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from flask import send_file
        import io

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Match Upload Template"

        # Define headers
        headers = [
            "League Name",
            "Season Name",
            "Home Team",
            "Away Team",
            "Match Date",
            "Match Time",
            "Venue",
            "Matchday"
        ]

        # Style the header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Add example row
        example_data = [
            "Premier League",
            "2024-25 Season",
            "Manchester United",
            "Liverpool",
            "2025-10-15",
            "14:30",
            "Old Trafford",
            "1"
        ]

        for col_num, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bulk_match_template.xlsx'
        )

    except ImportError:
        flash('Excel processing library not available. Please install openpyxl.', 'error')
        return redirect(url_for('admin.manage_matches'))
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('admin.manage_matches'))

@admin_bp.route('/manage_matches', methods=['GET', 'POST'])
@admin_required
def manage_matches():
    db = get_db()
    cur = db.cursor()

    # Add match display settings table if it doesn't exist
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS match_display_settings (
                setting_id SERIAL PRIMARY KEY,
                league_id INTEGER REFERENCES leagues(league_id) ON DELETE CASCADE,
                default_view VARCHAR(20) DEFAULT 'cards',
                show_venue BOOLEAN DEFAULT TRUE,
                show_time BOOLEAN DEFAULT TRUE,
                show_scores BOOLEAN DEFAULT TRUE,
                show_status BOOLEAN DEFAULT TRUE,
                public_calendar_enabled BOOLEAN DEFAULT TRUE,
                public_list_enabled BOOLEAN DEFAULT TRUE,
                public_cards_enabled BOOLEAN DEFAULT TRUE,
                public_timeline_enabled BOOLEAN DEFAULT FALSE,
                calendar_theme VARCHAR(50) DEFAULT 'default',
                card_style VARCHAR(50) DEFAULT 'modern',
                timezone VARCHAR(50) DEFAULT 'UTC',
                date_format VARCHAR(20) DEFAULT 'MM/DD/YYYY',
                time_format VARCHAR(10) DEFAULT '12h',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(league_id)
            )
        """)

        # Add additional columns to matches table if they don't exist
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS venue VARCHAR(255)")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS match_time TIME")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS tournament_id INTEGER")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS matchday INTEGER")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS public_notes TEXT")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS ticket_url VARCHAR(500)")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS weather_info JSONB")

        db.commit()
    except Exception as e:
        db.rollback()

    if request.method == 'POST':
        try:
            # Handle display settings update
            if 'update_display_settings' in request.form:
                # Debug logging
                print(f"DEBUG: Form data received: {dict(request.form)}")

                # Check if this is an intentional save (not accidental form submission)
                intentional_save = request.form.get('intentional_save', '0')
                print(f"DEBUG: intentional_save = '{intentional_save}'")

                if intentional_save != '1':
                    # Ignore accidental form submissions
                    print("DEBUG: Ignoring accidental form submission")
                    return redirect(url_for('admin.manage_matches'))

                league_id = request.form.get('settings_league_id')

                # Validate that a league is selected
                if not league_id or league_id == '':
                    flash('Please select a league to update settings', 'error')
                    return redirect(url_for('admin.manage_matches'))

                # Convert to integer to ensure it's valid
                try:
                    league_id = int(league_id)
                except (ValueError, TypeError):
                    print(f"DEBUG: Failed to convert league_id to int: {league_id}")
                    flash('Invalid league selection', 'error')
                    return redirect(url_for('admin.manage_matches'))

                default_view = request.form.get('default_view', 'cards')
                show_venue = 'show_venue' in request.form
                show_time = 'show_time' in request.form
                show_scores = 'show_scores' in request.form
                show_status = 'show_status' in request.form
                public_calendar_enabled = 'public_calendar_enabled' in request.form
                public_list_enabled = 'public_list_enabled' in request.form
                public_cards_enabled = 'public_cards_enabled' in request.form
                public_timeline_enabled = 'public_timeline_enabled' in request.form
                calendar_theme = request.form.get('calendar_theme', 'default')
                card_style = request.form.get('card_style', 'modern')
                timezone = request.form.get('timezone', 'UTC')
                date_format = request.form.get('date_format', 'MM/DD/YYYY')
                time_format = request.form.get('time_format', '12h')

                # Final safeguard: ensure league_id is a valid integer
                if not isinstance(league_id, int) or league_id <= 0:
                    print(f"DEBUG: Invalid league_id at INSERT time: {league_id}")
                    flash('Invalid league selection', 'error')
                    return redirect(url_for('admin.manage_matches'))

                print(f"DEBUG: About to INSERT with league_id={league_id}")

                # Upsert display settings
                cur.execute("""
                    INSERT INTO match_display_settings
                    (league_id, default_view, show_venue, show_time, show_scores, show_status,
                     public_calendar_enabled, public_list_enabled, public_cards_enabled, public_timeline_enabled,
                     calendar_theme, card_style, timezone, date_format, time_format, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (league_id) DO UPDATE SET
                        default_view = EXCLUDED.default_view,
                        show_venue = EXCLUDED.show_venue,
                        show_time = EXCLUDED.show_time,
                        show_scores = EXCLUDED.show_scores,
                        show_status = EXCLUDED.show_status,
                        public_calendar_enabled = EXCLUDED.public_calendar_enabled,
                        public_list_enabled = EXCLUDED.public_list_enabled,
                        public_cards_enabled = EXCLUDED.public_cards_enabled,
                        public_timeline_enabled = EXCLUDED.public_timeline_enabled,
                        calendar_theme = EXCLUDED.calendar_theme,
                        card_style = EXCLUDED.card_style,
                        timezone = EXCLUDED.timezone,
                        date_format = EXCLUDED.date_format,
                        time_format = EXCLUDED.time_format,
                        updated_at = CURRENT_TIMESTAMP
                """, (league_id, default_view, show_venue, show_time, show_scores, show_status,
                      public_calendar_enabled, public_list_enabled, public_cards_enabled, public_timeline_enabled,
                      calendar_theme, card_style, timezone, date_format, time_format))

                flash('Display settings updated successfully', 'success')
                db.commit()

            # Handle bulk operations
            elif 'bulk_action' in request.form:
                action = request.form['bulk_action']
                selected_matches = request.form.getlist('selected_matches')

                if action == 'bulk_reschedule' and selected_matches:
                    new_date = request.form.get('bulk_new_date')
                    new_time = request.form.get('bulk_new_time')

                    for match_id in selected_matches:
                        update_query = "UPDATE matches SET utc_date = %s"
                        params = [new_date]

                        if new_time:
                            update_query += ", match_time = %s"
                            params.append(new_time)

                        update_query += " WHERE match_id = %s"
                        params.append(match_id)

                        cur.execute(update_query, params)

                    flash(f'Rescheduled {len(selected_matches)} matches', 'success')

                elif action == 'bulk_status_update' and selected_matches:
                    new_status = request.form.get('bulk_new_status')

                    for match_id in selected_matches:
                        cur.execute("UPDATE matches SET status = %s WHERE match_id = %s", (new_status, match_id))

                    flash(f'Updated status for {len(selected_matches)} matches', 'success')

                elif action == 'bulk_delete' and selected_matches:
                    for match_id in selected_matches:
                        cur.execute("DELETE FROM matches WHERE match_id = %s", (match_id,))

                    flash(f'Deleted {len(selected_matches)} matches', 'success')

                db.commit()

            # Handle bulk upload from Excel
            elif request.form.get('action') == 'bulk_upload':
                if 'bulk_file' not in request.files:
                    flash('No file uploaded', 'error')
                    return redirect(url_for('admin.manage_matches'))

                file = request.files['bulk_file']
                if file.filename == '':
                    flash('No file selected', 'error')
                    return redirect(url_for('admin.manage_matches'))

                if not file.filename.endswith(('.xlsx', '.xls')):
                    flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'error')
                    return redirect(url_for('admin.manage_matches'))

                try:
                    import openpyxl
                    from datetime import datetime, time as datetime_time

                    # Load workbook
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active

                    skip_errors = 'skip_errors' in request.form
                    created_count = 0
                    error_count = 0
                    errors = []

                    # Skip header row, process data rows
                    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or not any(row):  # Skip empty rows
                            continue

                        if row_num > 502:  # Max 500 matches + header + 1
                            errors.append(f"Row {row_num}: Exceeded maximum of 500 matches")
                            break

                        try:
                            # Extract data from columns
                            league_name = row[0]
                            season_name = row[1]
                            home_team_name = row[2]
                            away_team_name = row[3]
                            match_date = row[4]
                            match_time = row[5] if len(row) > 5 else None
                            venue_name = row[6] if len(row) > 6 else None
                            matchday = row[7] if len(row) > 7 else None

                            # Validate required fields
                            if not all([league_name, season_name, home_team_name, away_team_name, match_date]):
                                error_msg = f"Row {row_num}: Missing required fields"
                                errors.append(error_msg)
                                error_count += 1
                                if not skip_errors:
                                    raise ValueError(error_msg)
                                continue

                            # Look up league
                            cur.execute("SELECT league_id FROM leagues WHERE name = %s", (league_name,))
                            league_result = cur.fetchone()
                            if not league_result:
                                error_msg = f"Row {row_num}: League '{league_name}' not found"
                                errors.append(error_msg)
                                error_count += 1
                                if not skip_errors:
                                    raise ValueError(error_msg)
                                continue
                            league_id = league_result[0]

                            # Look up season
                            cur.execute("SELECT season_id FROM seasons WHERE name = %s AND league_id = %s",
                                       (season_name, league_id))
                            season_result = cur.fetchone()
                            if not season_result:
                                error_msg = f"Row {row_num}: Season '{season_name}' not found for league '{league_name}'"
                                errors.append(error_msg)
                                error_count += 1
                                if not skip_errors:
                                    raise ValueError(error_msg)
                                continue
                            season_id = season_result[0]

                            # Look up home team
                            cur.execute("SELECT team_id FROM teams WHERE name = %s", (home_team_name,))
                            home_team_result = cur.fetchone()
                            if not home_team_result:
                                error_msg = f"Row {row_num}: Home team '{home_team_name}' not found"
                                errors.append(error_msg)
                                error_count += 1
                                if not skip_errors:
                                    raise ValueError(error_msg)
                                continue
                            home_team_id = home_team_result[0]

                            # Look up away team
                            cur.execute("SELECT team_id FROM teams WHERE name = %s", (away_team_name,))
                            away_team_result = cur.fetchone()
                            if not away_team_result:
                                error_msg = f"Row {row_num}: Away team '{away_team_name}' not found"
                                errors.append(error_msg)
                                error_count += 1
                                if not skip_errors:
                                    raise ValueError(error_msg)
                                continue
                            away_team_id = away_team_result[0]

                            # Parse date
                            if isinstance(match_date, datetime):
                                match_date_str = match_date.strftime('%Y-%m-%d')
                            else:
                                match_date_str = str(match_date)

                            # Parse time if provided
                            match_time_str = None
                            if match_time:
                                if isinstance(match_time, datetime_time):
                                    match_time_str = match_time.strftime('%H:%M')
                                elif isinstance(match_time, datetime):
                                    match_time_str = match_time.strftime('%H:%M')
                                else:
                                    match_time_str = str(match_time)

                            # Venue is stored as VARCHAR, not an ID
                            venue = venue_name if venue_name else None

                            # Insert match
                            cur.execute('''
                                INSERT INTO matches
                                (home_team_id, away_team_id, season_id, league_id, utc_date, match_time, venue, matchday, status)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'scheduled')
                            ''', (home_team_id, away_team_id, season_id, league_id, match_date_str,
                                  match_time_str, venue, matchday))

                            created_count += 1

                        except Exception as e:
                            error_msg = f"Row {row_num}: {str(e)}"
                            errors.append(error_msg)
                            error_count += 1
                            if not skip_errors:
                                db.rollback()
                                flash(f'Upload failed: {error_msg}', 'error')
                                return redirect(url_for('admin.manage_matches'))

                    db.commit()

                    # Show results
                    if created_count > 0:
                        flash(f'Successfully created {created_count} matches', 'success')
                    if error_count > 0:
                        error_summary = f'{error_count} rows had errors'
                        if len(errors) <= 10:
                            error_summary += ': ' + '; '.join(errors)
                        else:
                            error_summary += ': ' + '; '.join(errors[:10]) + f' (and {len(errors)-10} more)'
                        flash(error_summary, 'warning')

                    return redirect(url_for('admin.manage_matches'))

                except ImportError:
                    flash('Excel processing library not available. Please install openpyxl.', 'error')
                    return redirect(url_for('admin.manage_matches'))
                except Exception as e:
                    db.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
                    return redirect(url_for('admin.manage_matches'))

            # Handle regular match CRUD
            else:
                match_id = request.form.get('match_id')
                date = request.form['date']
                time = request.form.get('time')
                team1_id = request.form['team1_id']
                team2_id = request.form['team2_id']
                season_id = request.form['season_id']
                league_id = request.form['league_id']
                venue = request.form.get('venue', '')
                tournament_id = request.form.get('tournament_id') if request.form.get('tournament_id') else None
                status = request.form.get('status', 'scheduled')
                home_score = request.form.get('home_score')
                away_score = request.form.get('away_score')
                matchday = request.form.get('matchday')
                public_notes = request.form.get('public_notes', '')
                ticket_url = request.form.get('ticket_url', '')

                if 'submit' in request.form:
                    if match_id:
                        # Update existing match
                        cur.execute('''
                            UPDATE matches SET
                                utc_date = %s, match_time = %s, home_team_id = %s, away_team_id = %s,
                                season_id = %s, league_id = %s, venue = %s, tournament_id = %s,
                                status = %s, matchday = %s, public_notes = %s, ticket_url = %s
                            WHERE match_id = %s
                        ''', (date, time, team1_id, team2_id, season_id, league_id, venue,
                              tournament_id, status, matchday, public_notes, ticket_url, match_id))

                        # Update scores if provided
                        if home_score is not None and away_score is not None:
                            cur.execute('''
                                INSERT INTO scores (match_id, full_time_home, full_time_away, half_time_home, half_time_away)
                                VALUES (%s, %s, %s, 0, 0)
                                ON CONFLICT (match_id) DO UPDATE SET
                                    full_time_home = EXCLUDED.full_time_home,
                                    full_time_away = EXCLUDED.full_time_away,
                                    updated_at = CURRENT_TIMESTAMP
                            ''', (match_id, home_score, away_score))

                        flash('Match updated successfully', 'success')
                    else:
                        # Create new match
                        cur.execute('''
                            INSERT INTO matches
                            (utc_date, match_time, home_team_id, away_team_id, season_id, league_id,
                             venue, tournament_id, status, matchday, public_notes, ticket_url)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ''', (date, time, team1_id, team2_id, season_id, league_id, venue,
                              tournament_id, status, matchday, public_notes, ticket_url))

                        flash('Match added successfully', 'success')

                elif 'delete' in request.form:
                    match_id = request.form['deleteEntityId']
                    cur.execute('DELETE FROM matches WHERE match_id = %s', (match_id,))
                    flash('Match deleted successfully', 'success')

                elif 'quick_status_update' in request.form:
                    match_id = request.form['match_id']
                    status = request.form['status']
                    home_score = request.form.get('home_score')
                    away_score = request.form.get('away_score')

                    cur.execute('UPDATE matches SET status = %s WHERE match_id = %s', (status, match_id))

                    if home_score is not None and away_score is not None and status == 'completed':
                        cur.execute('''
                            INSERT INTO scores (match_id, full_time_home, full_time_away, half_time_home, half_time_away)
                            VALUES (%s, %s, %s, 0, 0)
                            ON CONFLICT (match_id) DO UPDATE SET
                                full_time_home = EXCLUDED.full_time_home,
                                full_time_away = EXCLUDED.full_time_away,
                                updated_at = CURRENT_TIMESTAMP
                        ''', (match_id, home_score, away_score))

                    flash('Match status updated successfully', 'success')

                db.commit()

        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_matches'))

    # Get enhanced match data with all related information
    cur.execute('''
        SELECT
            m.match_id, m.utc_date, t1.name AS team1, t2.name AS team2,
            s.year AS season, l.name AS league, m.home_team_id, m.away_team_id,
            m.tournament_id, m.venue,
            CASE
                WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                ELSE 'scheduled'
            END AS status,
            m.match_time, m.matchday,
            sc.full_time_home, sc.full_time_away, sc.half_time_home, sc.half_time_away,
            m.public_notes, m.ticket_url, l.league_id,
            '' AS tournament_name
        FROM matches m
        JOIN teams t1 ON m.home_team_id = t1.team_id
        JOIN teams t2 ON m.away_team_id = t2.team_id
        JOIN seasons s ON m.season_id = s.season_id
        JOIN leagues l ON m.league_id = l.league_id
        LEFT JOIN scores sc ON m.match_id = sc.match_id
        ORDER BY m.utc_date DESC, m.match_time DESC NULLS LAST
    ''')
    matches_raw = cur.fetchall()

    # Convert matches to serializable format
    matches = []
    for match in matches_raw:
        match_dict = []
        for i, value in enumerate(match):
            if hasattr(value, 'isoformat'):  # datetime/date objects
                match_dict.append(value.isoformat())
            else:
                match_dict.append(value)
        matches.append(match_dict)

    # Get all supporting data
    cur.execute('SELECT team_id, name FROM teams ORDER BY name')
    teams = cur.fetchall()

    cur.execute('SELECT season_id, year FROM seasons ORDER BY year DESC')
    seasons = cur.fetchall()

    cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = cur.fetchall()

    # Check if tournaments table exists
    try:
        cur.execute('SELECT tournament_id, name FROM tournaments ORDER BY name')
        tournaments = cur.fetchall()
    except Exception:
        db.rollback()  # Reset transaction state after failed query
        tournaments = []

    # Get display settings for all leagues
    cur.execute('''
        SELECT
            mds.*, l.name as league_name
        FROM match_display_settings mds
        JOIN leagues l ON mds.league_id = l.league_id
        ORDER BY l.name
    ''')
    display_settings = cur.fetchall()

    # Get match statistics
    try:
        cur.execute('''
            SELECT
                COUNT(*) as total_matches,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'scheduled' THEN 1 END) as scheduled,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'live' THEN 1 END) as live,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'completed' THEN 1 END) as completed,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'postponed' THEN 1 END) as postponed,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'cancelled' THEN 1 END) as cancelled
            FROM matches m
            LEFT JOIN scores sc ON m.match_id = sc.match_id
        ''')
        match_stats = cur.fetchone()
    except Exception:
        db.rollback()
        # Fallback statistics if query fails
        match_stats = (0, 0, 0, 0, 0, 0)

    # Get venues
    try:
        cur.execute('SELECT stadium_id, name FROM stadiums ORDER BY name')
        venues = cur.fetchall()
    except Exception:
        db.rollback()
        venues = []

    # Get recent matches for sidebar
    cur.execute('''
        SELECT
            m.match_id,
            t1.name AS home_team,
            t2.name AS away_team,
            m.utc_date,
            m.matchday,
            sc.full_time_home AS home_score,
            sc.full_time_away AS away_score,
            CASE
                WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                ELSE 'scheduled'
            END AS status
        FROM matches m
        JOIN teams t1 ON m.home_team_id = t1.team_id
        JOIN teams t2 ON m.away_team_id = t2.team_id
        LEFT JOIN scores sc ON m.match_id = sc.match_id
        ORDER BY m.utc_date DESC, m.match_time DESC NULLS LAST
        LIMIT 10
    ''')
    recent_matches_raw = cur.fetchall()

    # Convert to list of dicts for template
    recent_matches = []
    for match in recent_matches_raw:
        recent_matches.append({
            'match_id': match[0],
            'home_team': match[1],
            'away_team': match[2],
            'utc_date': match[3],
            'matchday': match[4],
            'home_score': match[5],
            'away_score': match[6],
            'status': match[7]
        })

    # Get total matches count
    total_matches = match_stats[0] if match_stats else 0

    cur.close()

    return render_template('manage_matches.html',
                         matches=matches_raw, matches_json=matches, teams=teams, seasons=seasons, leagues=leagues,
                         tournaments=tournaments, display_settings=display_settings,
                         match_stats=match_stats, venues=venues, recent_matches=recent_matches,
                         total_matches=total_matches)



@admin_bp.route('/manage_media', methods=['GET', 'POST'])
@admin_required
def manage_media():
    """Media gallery management - images and videos."""
    org = getattr(g, 'org', None)
    if org is None:
        flash('No organization selected for media library', 'error')
        return redirect(url_for('admin.admin_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'add':
                file = request.files.get('file')
                media_type = request.form.get('media_type') or None
                create_media_asset(
                    title=request.form.get('title', ''),
                    description=request.form.get('description'),
                    category=request.form.get('category'),
                    media_type=media_type,
                    file=file if file and file.filename else None,
                    source_url=request.form.get('source_url'),
                    alt_text=request.form.get('alt_text'),
                    uploaded_by_user_id=getattr(current_user, 'id', None),
                )
                flash('Media added successfully', 'success')
            elif action == 'edit':
                asset_id = request.form.get('media_id')
                asset = MediaAsset.query.filter_by(org_id=org.id, id=asset_id).first_or_404()
                file = request.files.get('file')
                media_type = request.form.get('media_type') or None
                update_media_asset(
                    asset,
                    title=request.form.get('title'),
                    description=request.form.get('description'),
                    category=request.form.get('category'),
                    media_type=media_type,
                    file=file if file and file.filename else None,
                    source_url=request.form.get('source_url') if 'source_url' in request.form else None,
                    alt_text=request.form.get('alt_text'),
                )
                flash('Media updated successfully', 'success')
            elif action == 'delete':
                asset_id = request.form.get('media_id')
                asset = MediaAsset.query.filter_by(org_id=org.id, id=asset_id).first_or_404()
                delete_media_asset(asset)
                flash('Media deleted successfully', 'success')
        except ValueError as exc:
            flash(str(exc), 'error')
        except Exception as exc:
            flash(f'An error occurred: {exc}', 'error')
        return redirect(url_for('admin.manage_media'))

    assets = (MediaAsset.query
              .filter_by(org_id=org.id)
              .order_by(MediaAsset.created_at.desc())
              .all())
    media_items = serialize_media_collection(assets)
    default_categories = {'match_highlights', 'training', 'team_photos', 'events', 'celebrations', 'facilities'}
    categories = sorted(default_categories.union({item['category'] for item in media_items if item.get('category')}))

    return render_template(
        'manage_media.html',
        media_items=media_items,
        categories=categories,
    )





@admin_bp.route('/manage_countries', methods=['GET', 'POST'])
@admin_required
def manage_countries():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            country_id = request.form.get('country_id')
            name = request.form['name']
            flag_url = request.form['flag_url']

            if 'submit' in request.form:
                if country_id:
                    cur.execute('UPDATE countries SET name = %s, flag_url = %s WHERE country_id = %s', 
                                (name, flag_url, country_id))
                    flash('Country updated successfully', 'success')
                else:
                    cur.execute('INSERT INTO countries (name, flag_url) VALUES (%s, %s)', 
                                (name, flag_url))
                    flash('Country added successfully', 'success')
            elif 'delete' in request.form:
                country_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM countries WHERE country_id = %s', (country_id,))
                flash('Country deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_countries'))

    cur.execute('SELECT country_id, name, flag_url FROM countries')
    countries = cur.fetchall()
    cur.close()
    return render_template('manage_countries.html', countries=countries)



@admin_bp.route('/manage_referees', methods=['GET', 'POST'])
@admin_required
def manage_referees():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            referee_id = request.form.get('referee_id')
            name = request.form['name']
            nationality = request.form['nationality']

            if 'submit' in request.form:
                if referee_id:
                    cur.execute('UPDATE referees SET name = %s, nationality = %s WHERE referee_id = %s', 
                                (name, nationality, referee_id))
                    flash('Referee updated successfully', 'success')
                else:
                    cur.execute('INSERT INTO referees (name, nationality) VALUES (%s, %s)', 
                                (name, nationality))
                    flash('Referee added successfully', 'success')
            elif 'delete' in request.form:
                referee_id = request.form['deleteEntityId']
                cur.execute('DELETE FROM referees WHERE referee_id = %s', (referee_id,))
                flash('Referee deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_referees'))

    cur.execute('SELECT referee_id, name, nationality FROM referees')
    referees = cur.fetchall()
    cur.close()
    return render_template('manage_referees.html', referees=referees)



@admin_bp.route('/player_stats', methods=['GET', 'POST'])
@admin_required
def player_stats():
    db = get_db()
    _ensure_scorer_metrics_table(db)

    if request.method == 'POST':
        action = request.form.get('action', 'basic').lower()
        post_cur = db.cursor()

        def _to_int(value, default=0):
            try:
                return int(value)
            except (TypeError, ValueError):
                return default

        def _to_float(value, default=None):
            try:
                if value is None or value == '':
                    return default
                return float(value)
            except (TypeError, ValueError):
                return default

        try:
            if 'delete' in request.form:
                scorer_id = request.form.get('deleteEntityId')
                if scorer_id:
                    post_cur.execute('DELETE FROM scorer_metrics WHERE scorer_id = %s', (scorer_id,))
                    post_cur.execute('DELETE FROM scorers WHERE scorer_id = %s', (scorer_id,))
                    flash('Scorer entry removed.', 'success')
                else:
                    flash('Unable to determine scorer record to delete.', 'error')
            else:
                # Combined form handling - both basic and advanced stats
                scorer_id = request.form.get('scorer_id')
                player_id = request.form.get('player_id')
                season_id = request.form.get('season_id')
                league_id = request.form.get('league_id')
                goals = _to_int(request.form.get('goals'))
                assists = _to_int(request.form.get('assists'))
                penalties = _to_int(request.form.get('penalties'))

                if not player_id or not season_id or not league_id:
                    raise ValueError('Player, season, and league are required.')

                # Save basic stats to scorers table
                if scorer_id:
                    post_cur.execute(
                        'UPDATE scorers SET player_id = %s, season_id = %s, league_id = %s, goals = %s, assists = %s, penalties = %s WHERE scorer_id = %s',
                        (player_id, season_id, league_id, goals, assists, penalties, scorer_id)
                    )
                else:
                    post_cur.execute(
                        'INSERT INTO scorers (player_id, season_id, league_id, goals, assists, penalties) VALUES (%s, %s, %s, %s, %s, %s) RETURNING scorer_id',
                        (player_id, season_id, league_id, goals, assists, penalties)
                    )
                    scorer_id = post_cur.fetchone()[0]

                # Save advanced stats to scorer_metrics table
                games_played = _to_int(request.form.get('games_played'))
                minutes_played = _to_int(request.form.get('minutes_played'))
                shots_on_target = _to_int(request.form.get('shots_on_target'))
                shot_attempts = _to_int(request.form.get('shot_attempts'))
                passes_completed = _to_int(request.form.get('passes_completed'))
                passes_attempted = _to_int(request.form.get('passes_attempted'))
                saves = _to_int(request.form.get('saves'))
                rating = _to_float(request.form.get('rating'))

                custom_metrics_raw = request.form.get('custom_metrics_json') or '{}'
                try:
                    custom_metrics = json.loads(custom_metrics_raw)
                    if not isinstance(custom_metrics, dict):
                        custom_metrics = {}
                except json.JSONDecodeError:
                    custom_metrics = {}

                post_cur.execute("""
                    INSERT INTO scorer_metrics (scorer_id, games_played, minutes_played, shots_on_target, shot_attempts, passes_completed, passes_attempted, saves, rating, custom_metrics)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (scorer_id) DO UPDATE SET
                        games_played = EXCLUDED.games_played,
                        minutes_played = EXCLUDED.minutes_played,
                        shots_on_target = EXCLUDED.shots_on_target,
                        shot_attempts = EXCLUDED.shot_attempts,
                        passes_completed = EXCLUDED.passes_completed,
                        passes_attempted = EXCLUDED.passes_attempted,
                        saves = EXCLUDED.saves,
                        rating = EXCLUDED.rating,
                        custom_metrics = EXCLUDED.custom_metrics
                """,
                    (scorer_id, games_played, minutes_played, shots_on_target, shot_attempts, passes_completed, passes_attempted, saves, rating, json.dumps(custom_metrics) if custom_metrics else None)
                )

                flash('Player statistics saved successfully.', 'success')
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        else:
            db.commit()
        finally:
            post_cur.close()
        return redirect(url_for('admin.player_stats'))

    cur = db.cursor()
    cur.execute("""
        SELECT
            s.scorer_id,
            p.player_id,
            p.name,
            se.season_id,
            se.year,
            l.league_id,
            l.name,
            s.goals,
            s.assists,
            s.penalties,
            COALESCE(sm.games_played, 0) AS games_played,
            COALESCE(sm.minutes_played, 0) AS minutes_played,
            COALESCE(sm.shots_on_target, 0) AS shots_on_target,
            COALESCE(sm.shot_attempts, 0) AS shot_attempts,
            COALESCE(sm.passes_completed, 0) AS passes_completed,
            COALESCE(sm.passes_attempted, 0) AS passes_attempted,
            COALESCE(sm.saves, 0) AS saves,
            sm.rating,
            sm.custom_metrics
        FROM scorers s
        JOIN players p ON s.player_id = p.player_id
        JOIN seasons se ON s.season_id = se.season_id
        JOIN leagues l ON s.league_id = l.league_id
        LEFT JOIN scorer_metrics sm ON sm.scorer_id = s.scorer_id
    """)
    scorer_rows = cur.fetchall()

    cur.execute('SELECT player_id, name FROM players ORDER BY name ASC')
    players = cur.fetchall()
    cur.execute('SELECT season_id, year FROM seasons ORDER BY year DESC')
    seasons = cur.fetchall()
    cur.execute('SELECT league_id, name, sport FROM leagues ORDER BY name ASC')
    leagues_raw = cur.fetchall()
    leagues = [{'id': row[0], 'name': row[1], 'sport': row[2] or 'soccer'} for row in leagues_raw]
    cur.close()

    leaderboard = []
    total_goals = 0
    total_assists = 0
    total_penalties = 0
    ratings = []
    custom_metric_keys = set()

    for row in scorer_rows:
        (scorer_id, player_id, player_name, season_id, season_year, league_id, league_name, goals, assists, penalties,
         games_played, minutes_played, shots_on_target, shot_attempts, passes_completed, passes_attempted, saves, rating, custom_metrics_raw) = row

        goals = goals or 0
        assists = assists or 0
        penalties = penalties or 0
        games_played = games_played or 0
        minutes_played = minutes_played or 0
        shots_on_target = shots_on_target or 0
        shot_attempts = shot_attempts or 0
        passes_completed = passes_completed or 0
        passes_attempted = passes_attempted or 0
        saves = saves or 0

        total_goals += goals
        total_assists += assists
        total_penalties += penalties
        if rating is not None:
            ratings.append(float(rating))

        goal_contributions = goals + assists
        per_game_contribution = goal_contributions / games_played if games_played else None
        per_90_contribution = (goal_contributions * 90 / minutes_played) if minutes_played else None
        shot_accuracy = (shots_on_target / shot_attempts) if shot_attempts else None
        pass_accuracy = (passes_completed / passes_attempted) if passes_attempted else None
        defensive_value = (saves / games_played) if games_played and saves else None

        try:
            custom_metrics = json.loads(custom_metrics_raw) if custom_metrics_raw else {}
        except (TypeError, json.JSONDecodeError):
            custom_metrics = {}
        if isinstance(custom_metrics, dict):
            custom_metric_keys.update(custom_metrics.keys())
        else:
            custom_metrics = {}

        analytics_score = (goal_contributions * 3)
        if per_game_contribution:
            analytics_score += per_game_contribution * 5
        if rating is not None:
            analytics_score += float(rating) * 2
        if shot_accuracy:
            analytics_score += shot_accuracy * 10
        if pass_accuracy:
            analytics_score += pass_accuracy * 5

        leaderboard.append({
            'scorer_id': scorer_id,
            'player_id': player_id,
            'player_name': player_name,
            'season_id': season_id,
            'season_year': season_year,
            'league_id': league_id,
            'league_name': league_name,
            'goals': goals,
            'assists': assists,
            'penalties': penalties,
            'goal_contributions': goal_contributions,
            'games_played': games_played,
            'minutes_played': minutes_played,
            'per_game_contribution': per_game_contribution,
            'per_90_contribution': per_90_contribution,
            'shot_accuracy': shot_accuracy,
            'pass_accuracy': pass_accuracy,
            'defensive_value': defensive_value,
            'rating': float(rating) if rating is not None else None,
            'shots_on_target': shots_on_target,
            'shot_attempts': shot_attempts,
            'passes_completed': passes_completed,
            'passes_attempted': passes_attempted,
            'saves': saves,
            'custom_metrics': custom_metrics,
            'analytics_score': analytics_score
        })

    leaderboard.sort(key=lambda item: (item['analytics_score'], item['goal_contributions'], item['goals'], item['assists']), reverse=True)
    for idx, entry in enumerate(leaderboard, start=1):
        entry['rank'] = idx
        if leaderboard and leaderboard[0]['analytics_score']:
            top_score = leaderboard[0]['analytics_score'] or 1
            entry['momentum'] = min(100, round((entry['analytics_score'] / top_score) * 100, 2)) if entry['analytics_score'] else 0
        else:
            entry['momentum'] = 0

    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else None
    top_goal_scorer = max(leaderboard, key=lambda item: item['goals'], default=None)
    top_assist_player = max(leaderboard, key=lambda item: item['assists'], default=None)
    most_efficient = max((item for item in leaderboard if item['per_game_contribution']), key=lambda item: item['per_game_contribution'], default=None)
    best_rating = max((item for item in leaderboard if item['rating'] is not None), key=lambda item: item['rating'], default=None)

    analytics_summary = {
        'avg_rating': avg_rating,
        'top_goal_scorer': top_goal_scorer,
        'top_assist_player': top_assist_player,
        'most_efficient': most_efficient,
        'best_rating': best_rating,
        'total_players': len(leaderboard),
        'total_goals': total_goals,
        'total_assists': total_assists,
        'total_penalties': total_penalties
    }

    advanced_metrics_map = {
        entry['scorer_id']: {
            'games_played': entry['games_played'],
            'minutes_played': entry['minutes_played'],
            'shots_on_target': entry['shots_on_target'],
            'shot_attempts': entry['shot_attempts'],
            'passes_completed': entry['passes_completed'],
            'passes_attempted': entry['passes_attempted'],
            'saves': entry['saves'],
            'rating': entry['rating'],
            'custom_metrics': entry['custom_metrics'],
        }
        for entry in leaderboard
    }

    custom_metric_summary = {}
    for key in sorted(custom_metric_keys):
        leader_entry = None
        leader_value = None
        numeric_values = []
        for entry in leaderboard:
            value = entry['custom_metrics'].get(key)
            if value is None:
                continue
            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                continue
            numeric_values.append(numeric_value)
            if leader_value is None or numeric_value > leader_value:
                leader_value = numeric_value
                leader_entry = entry
        if numeric_values:
            custom_metric_summary[key] = {
                'average': round(sum(numeric_values) / len(numeric_values), 2),
                'max': leader_value,
                'leader': leader_entry['player_name'] if leader_entry else None
            }

    # Get sport configurations for all leagues
    sport_configs = {league['id']: get_sport_config(league['sport']) for league in leagues}

    return render_template(
        'player_stats.html',
        leaderboard=leaderboard,
        players=players,
        seasons=seasons,
        leagues=leagues,
        analytics_summary=analytics_summary,
        custom_metric_keys=sorted(custom_metric_keys),
        advanced_metrics_map=advanced_metrics_map,
        custom_metric_summary=custom_metric_summary,
        sport_configs=sport_configs,
        get_sport_config=get_sport_config
    )
@admin_bp.route('/recalculate_standings', methods=['GET', 'POST'])
@admin_required
def recalculate_standings():
    db = get_db()
    results = []
    selected_league_id = None
    selected_season_id = None

    try:
        _ensure_league_rules_table(db)
    except Exception as exc:
        flash('Failed to prepare league rules storage: ' + str(exc), 'error')
        return redirect(url_for('admin.admin_dashboard'))

    if request.method == 'POST':
        cur = db.cursor()
        try:
            league_id_value = request.form.get('league_id')
            season_id_value = request.form.get('season_id')
            save_flag = bool(request.form.get('save'))

            try:
                league_id = int(league_id_value)
                season_id = int(season_id_value)
            except (TypeError, ValueError):
                flash('Please choose both a league and season.', 'error')
                return redirect(url_for('admin.recalculate_standings'))

            selected_league_id = league_id
            selected_season_id = season_id

            cur.execute('SELECT league_id FROM seasons WHERE season_id = %s', (season_id,))
            season_row = cur.fetchone()
            if not season_row:
                flash('Selected season could not be found.', 'error')
                return redirect(url_for('admin.recalculate_standings'))
            if season_row[0] != league_id:
                flash('Season does not belong to the selected league.', 'error')
                return redirect(url_for('admin.recalculate_standings'))

            # Get league's sport to determine point system
            cur.execute('SELECT sport FROM leagues WHERE league_id = %s', (league_id,))
            league_row = cur.fetchone()
            league_sport = league_row[0] if league_row else 'soccer'

            # Get sport-specific point configuration
            from slms.services.sport_config import get_standings_points_config
            sport_points = get_standings_points_config(league_sport)

            # Check for custom league rules, otherwise use sport defaults
            cur.execute('SELECT points_win, points_draw, points_loss, tiebreakers FROM league_rules WHERE league_id = %s', (league_id,))
            rules_row = cur.fetchone()
            points_win = int(rules_row[0]) if rules_row and rules_row[0] is not None else sport_points.get('win', 3)
            points_draw = int(rules_row[1]) if rules_row and rules_row[1] is not None else sport_points.get('draw', 1)
            points_loss = int(rules_row[2]) if rules_row and rules_row[2] is not None else sport_points.get('loss', 0)
            tiebreakers = _parse_tiebreakers(rules_row[3] if rules_row else None)

            cur.execute('SELECT team_id, name FROM teams WHERE league_id = %s', (league_id,))
            team_rows = cur.fetchall()
            team_names = {row[0]: row[1] for row in team_rows}
            stats = {team_id: _init_standing_record(team_id, name) for team_id, name in team_names.items()}

            cur.execute(
                """
                SELECT m.match_id, m.utc_date, m.home_team_id, m.away_team_id,
                       sc.full_time_home, sc.full_time_away
                FROM matches m
                LEFT JOIN scores sc ON sc.match_id = m.match_id
                WHERE m.league_id = %s AND m.season_id = %s
                ORDER BY m.utc_date ASC, m.match_id ASC
                """,
                (league_id, season_id)
            )
            match_rows = cur.fetchall()

            for row in match_rows:
                home_team_id = row[2]
                away_team_id = row[3]
                home_score = row[4]
                away_score = row[5]

                if home_team_id is None or away_team_id is None:
                    continue

                if home_team_id not in stats:
                    stats[home_team_id] = _init_standing_record(home_team_id, team_names.get(home_team_id, f'Team {home_team_id}'))
                if away_team_id not in stats:
                    stats[away_team_id] = _init_standing_record(away_team_id, team_names.get(away_team_id, f'Team {away_team_id}'))

                if home_score is None or away_score is None:
                    continue

                home_stat = stats[home_team_id]
                away_stat = stats[away_team_id]

                home_stat['played_games'] += 1
                away_stat['played_games'] += 1
                home_stat['goals_for'] += home_score
                home_stat['goals_against'] += away_score
                away_stat['goals_for'] += away_score
                away_stat['goals_against'] += home_score

                if home_score > away_score:
                    home_stat['won'] += 1
                    away_stat['lost'] += 1
                    home_stat['form'].append('W')
                    away_stat['form'].append('L')
                elif home_score < away_score:
                    home_stat['lost'] += 1
                    away_stat['won'] += 1
                    home_stat['form'].append('L')
                    away_stat['form'].append('W')
                else:
                    home_stat['draw'] += 1
                    away_stat['draw'] += 1
                    home_stat['form'].append('D')
                    away_stat['form'].append('D')

            if not stats:
                flash('No teams found for the selected league.', 'info')
            else:
                for team_id, record in stats.items():
                    goal_difference = record['goals_for'] - record['goals_against']
                    points = record['won'] * points_win + record['draw'] * points_draw + record['lost'] * points_loss
                    recent_form = record['form'][-5:]
                    results.append({
                        'team_id': team_id,
                        'team_name': record['team_name'],
                        'played_games': record['played_games'],
                        'won': record['won'],
                        'draw': record['draw'],
                        'lost': record['lost'],
                        'goals_for': record['goals_for'],
                        'goals_against': record['goals_against'],
                        'goal_difference': goal_difference,
                        'points': points,
                        'form': ' '.join(recent_form),
                    })

                results.sort(key=lambda r, tb=tiebreakers: _standing_sort_key(r, tb))
                for position, entry in enumerate(results, start=1):
                    entry['position'] = position

                if save_flag and results:
                    try:
                        for entry in results:
                            cur.execute('DELETE FROM standings WHERE team_id = %s', (entry['team_id'],))
                        for entry in results:
                            cur.execute(
                                'INSERT INTO standings (position, team_id, played_games, won, draw, lost, points, goals_for, goals_against, goal_difference, form) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                                (
                                    entry['position'], entry['team_id'], entry['played_games'], entry['won'],
                                    entry['draw'], entry['lost'], entry['points'], entry['goals_for'],
                                    entry['goals_against'], entry['goal_difference'], entry['form']
                                )
                            )
                        db.commit()
                        flash('Standings recalculated and saved.', 'success')
                    except Exception as exc:
                        db.rollback()
                        results = []
                        flash('Failed to save standings: ' + str(exc), 'error')
                elif results:
                    flash('Standings recalculated.', 'success')
                else:
                    flash('No completed matches found to calculate standings.', 'info')
        except Exception as exc:
            db.rollback()
            results = []
            flash('Failed to recalculate standings: ' + str(exc), 'error')
        finally:
            cur.close()

    cur = db.cursor()
    try:
        cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
        leagues = cur.fetchall()
        cur.execute('SELECT season_id, year, league_id FROM seasons ORDER BY year DESC, season_id DESC')
        seasons = cur.fetchall()
    finally:
        cur.close()
    return render_template('recalculate_standings.html', leagues=leagues, seasons=seasons, results=results, league_id=selected_league_id, season_id=selected_season_id)


@admin_bp.route('/manage_scores', methods=['GET', 'POST'])
@admin_required
def manage_scores():
    db = get_db()
    cur = db.cursor()

    # Add match status column if it doesn't exist
    try:
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'scheduled'")
        cur.execute("ALTER TABLE matches ADD COLUMN IF NOT EXISTS match_time TIME")
        cur.execute("ALTER TABLE scores ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        cur.execute("ALTER TABLE scores ADD COLUMN IF NOT EXISTS match_status VARCHAR(20) DEFAULT 'completed'")
        db.commit()
    except Exception as e:
        db.rollback()

    if request.method == 'POST':
        try:
            if 'ajax_update' in request.form:
                # Handle AJAX score updates
                match_id = request.form['match_id']
                field = request.form['field']
                value = int(request.form['value'])

                # Get or create score record
                cur.execute('SELECT score_id FROM scores WHERE match_id = %s', (match_id,))
                score_record = cur.fetchone()

                if score_record:
                    score_id = score_record[0]
                    cur.execute(f'UPDATE scores SET {field} = %s, updated_at = CURRENT_TIMESTAMP WHERE score_id = %s',
                                (value, score_id))
                else:
                    # Create new score record with default values
                    defaults = {'full_time_home': 0, 'full_time_away': 0, 'half_time_home': 0, 'half_time_away': 0}
                    defaults[field] = value
                    cur.execute('INSERT INTO scores (match_id, full_time_home, full_time_away, half_time_home, half_time_away) VALUES (%s, %s, %s, %s, %s)',
                                (match_id, defaults['full_time_home'], defaults['full_time_away'],
                                 defaults['half_time_home'], defaults['half_time_away']))

                # Update match status
                cur.execute('UPDATE matches SET status = %s WHERE match_id = %s', ('live', match_id))
                db.commit()

                return jsonify({'success': True})

            elif 'update_match_status' in request.form:
                # Handle match status updates
                match_id = request.form['match_id']
                status = request.form['status']
                cur.execute('UPDATE matches SET status = %s WHERE match_id = %s', (status, match_id))
                db.commit()
                flash(f'Match status updated to {status}', 'success')

            elif 'bulk_finalize' in request.form:
                # Bulk finalize matches
                match_ids = request.form.getlist('selected_matches')
                for match_id in match_ids:
                    cur.execute('UPDATE matches SET status = %s WHERE match_id = %s', ('completed', match_id))
                db.commit()
                flash(f'Finalized {len(match_ids)} matches', 'success')

            else:
                # Handle traditional form submission
                score_id = request.form.get('score_id')
                match_id = request.form['match_id']
                full_time_home = request.form['full_time_home']
                full_time_away = request.form['full_time_away']
                half_time_home = request.form['half_time_home']
                half_time_away = request.form['half_time_away']

                if 'submit' in request.form:
                    if score_id:
                        cur.execute('UPDATE scores SET match_id = %s, full_time_home = %s, full_time_away = %s, half_time_home = %s, half_time_away = %s, updated_at = CURRENT_TIMESTAMP WHERE score_id = %s',
                                    (match_id, full_time_home, full_time_away, half_time_home, half_time_away, score_id))
                        flash('Score updated successfully', 'success')
                    else:
                        cur.execute('INSERT INTO scores (match_id, full_time_home, full_time_away, half_time_home, half_time_away) VALUES (%s, %s, %s, %s, %s)',
                                    (match_id, full_time_home, full_time_away, half_time_home, half_time_away))
                        flash('Score added successfully', 'success')

                    # Update match status to completed
                    cur.execute('UPDATE matches SET status = %s WHERE match_id = %s', ('completed', match_id))

                elif 'delete' in request.form:
                    score_id = request.form['deleteEntityId']
                    cur.execute('DELETE FROM scores WHERE score_id = %s', (score_id,))
                    flash('Score deleted successfully', 'success')
                db.commit()

        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_scores'))

    # Get filter parameters
    league_filter = request.args.get('league_filter', '')

    # Enhanced data retrieval with match details
    base_query = '''
        SELECT
            m.match_id,
            m.utc_date,
            m.match_time,
            m.status,
            ht.name as home_team,
            at.name as away_team,
            ht.cresturl as home_crest,
            at.cresturl as away_crest,
            s.full_time_home as home_score,
            s.full_time_away as away_score,
            s.half_time_home,
            s.half_time_away,
            s.score_id,
            s.updated_at,
            l.name as league_name,
            se.year as season_year
        FROM matches m
        LEFT JOIN teams ht ON m.home_team_id = ht.team_id
        LEFT JOIN teams at ON m.away_team_id = at.team_id
        LEFT JOIN scores s ON m.match_id = s.match_id
        LEFT JOIN leagues l ON m.league_id = l.league_id
        LEFT JOIN seasons se ON m.season_id = se.season_id
    '''

    # Add WHERE clause for filtering
    where_conditions = []
    params = []

    if league_filter:
        where_conditions.append('m.league_id = %s')
        params.append(league_filter)

    if where_conditions:
        base_query += ' WHERE ' + ' AND '.join(where_conditions)

    base_query += '''
        ORDER BY
            CASE m.status
                WHEN 'live' THEN 1
                WHEN 'scheduled' THEN 2
                WHEN 'completed' THEN 3
                ELSE 4
            END,
            m.utc_date DESC, m.match_id DESC
    '''

    cur.execute(base_query, params)

    matches = cur.fetchall()

    # Get summary statistics
    try:
        cur.execute('''
            SELECT
                COUNT(*) as total_matches,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'scheduled' THEN 1 END) as scheduled,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'live' THEN 1 END) as live,
                COUNT(CASE WHEN
                    CASE
                        WHEN sc.full_time_home IS NOT NULL AND sc.full_time_away IS NOT NULL THEN 'completed'
                        WHEN m.utc_date < CURRENT_DATE THEN 'completed'
                        ELSE 'scheduled'
                    END = 'completed' THEN 1 END) as completed
            FROM matches m
            LEFT JOIN scores sc ON m.match_id = sc.match_id
        ''')
        stats = cur.fetchone()
    except Exception:
        db.rollback()
        # Fallback statistics if query fails
        stats = (0, 0, 0, 0)

    # Get leagues for filter dropdown
    cur.execute('SELECT league_id, name FROM leagues ORDER BY name')
    leagues = cur.fetchall()

    cur.close()
    return render_template('manage_scores.html', matches=matches, stats=stats, leagues=leagues)



@admin_bp.route('/manage_standings', methods=['GET', 'POST'])
@admin_required
def manage_standings():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        try:
            standing_id = request.form.get('standing_id')
            position = request.form['position']
            team_id = request.form['team_id']
            played_games = request.form['played_games']
            won = request.form['won']
            draw = request.form['draw']
            lost = request.form['lost']
            points = request.form['points']
            goals_for = request.form['goals_for']
            goals_against = request.form['goals_against']
            goal_difference = request.form['goal_difference']
            form = request.form['form']

            if 'add' in request.form:
                cur.execute('''
                    INSERT INTO standings (position, team_id, played_games, won, draw, lost, points, goals_for, goals_against, goal_difference, form)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (position, team_id, played_games, won, draw, lost, points, goals_for, goals_against, goal_difference, form))
                flash('Standing added successfully', 'success')
            elif 'edit' in request.form and standing_id:
                cur.execute('''
                    UPDATE standings
                    SET position = %s, team_id = %s, played_games = %s, won = %s, draw = %s, lost = %s, points = %s, goals_for = %s, goals_against = %s, goal_difference = %s, form = %s
                    WHERE standing_id = %s
                ''', (position, team_id, played_games, won, draw, lost, points, goals_for, goals_against, goal_difference, form, standing_id))
                flash('Standing updated successfully', 'success')
            elif 'delete' in request.form:
                standing_id = request.form['deleteItemId']
                cur.execute('DELETE FROM standings WHERE standing_id = %s', (standing_id,))
                flash('Standing deleted successfully', 'success')
            db.commit()
        except Exception as e:
            db.rollback()
            flash('An error occurred: ' + str(e), 'error')
        finally:
            cur.close()
        return redirect(url_for('admin.manage_standings'))

    cur.execute('''
        SELECT s.standing_id, s.position, t.name, s.played_games, s.won, s.draw, s.lost, s.points, s.goals_for, s.goals_against, s.goal_difference, s.form, s.team_id
        FROM standings s
        JOIN teams t ON s.team_id = t.team_id
    ''')
    standings = cur.fetchall()
    cur.execute('SELECT team_id, name FROM teams')
    teams = cur.fetchall()
    cur.close()
    return render_template('manage_standings.html', standings=standings, teams=teams)

@admin_bp.route('/manage_users', methods=['GET', 'POST'])
@admin_required
def manage_users():
    org = getattr(g, 'org', None)

    user_id_raw = (request.form.get('user_id') or '').strip() if request.method == 'POST' else ''
    is_admin = request.form.get('is_admin') == 'true' if request.method == 'POST' else False

    def _update_orm_user(orm_id: str) -> bool:
        query = db.session.query(User)
        if org is not None:
            query = query.filter_by(org_id=org.id)
        target = query.filter(User.id == orm_id).first()
        if not target:
            return False
        if target.role == UserRole.OWNER:
            flash('Owners already have full privileges.', 'info')
            return True
        target.role = UserRole.ADMIN if is_admin else UserRole.VIEWER
        db.session.commit()
        flash('User privilege updated successfully.', 'success')
        return True

    def _update_legacy_user(legacy_id: int) -> bool:
        sql_db = get_db()
        cur = sql_db.cursor()
        try:
            cur.execute('UPDATE users SET is_admin = %s WHERE user_id = %s', (is_admin, legacy_id))
            if cur.rowcount:
                sql_db.commit()
                flash('User privilege updated successfully.', 'success')
                return True
            sql_db.rollback()
            return False
        except Exception as exc:
            sql_db.rollback()
            flash('Failed to update legacy user: ' + str(exc), 'error')
            return True
        finally:
            cur.close()

    if request.method == 'POST':
        if not user_id_raw:
            flash('Please select a user before updating privileges.', 'error')
            return redirect(url_for('admin.manage_users'))

        updated = False
        if user_id_raw.startswith('legacy:'):
            try:
                legacy_id = int(user_id_raw.split(':', 1)[1])
            except ValueError:
                flash('Invalid user selection.', 'error')
                return redirect(url_for('admin.manage_users'))
            updated = _update_legacy_user(legacy_id)
        else:
            updated = _update_orm_user(user_id_raw)

        if not updated:
            flash('User not found for this organization.', 'error')
            db.session.rollback()
        db.session.expire_all()
        return redirect(url_for('admin.manage_users'))

    query = db.session.query(User)
    if org is not None:
        query = query.filter_by(org_id=org.id)
    orm_users = query.order_by(User.email.asc()).all()

    legacy_rows = []
    try:
        sql_db = get_db()
        cur = sql_db.cursor()
        cur.execute('SELECT user_id, username, is_admin FROM users ORDER BY username ASC')
        legacy_rows = cur.fetchall()
    except Exception:
        legacy_rows = []
    finally:
        try:
            cur.close()
        except Exception:
            pass

    user_rows = [
        {
            'id': orm_user.id,
            'email': orm_user.email,
            'is_admin': orm_user.role in (UserRole.ADMIN, UserRole.OWNER),
        }
        for orm_user in orm_users
    ]

    for legacy_id, username, is_admin_flag in legacy_rows:
        user_rows.append(
            {
                'id': f'legacy:{legacy_id}',
                'email': username,
                'is_admin': bool(is_admin_flag),
            }
        )

    if current_user.is_authenticated:
        current_id = getattr(current_user, 'id', None)
        current_email = getattr(current_user, 'email', None)
        if current_id and current_email:
            if current_id not in {row['id'] for row in user_rows}:
                user_rows.append(
                    {
                        'id': current_id,
                        'email': current_email,
                        'is_admin': current_user.has_role(UserRole.OWNER, UserRole.ADMIN) if hasattr(current_user, 'has_role') else False,
                    }
                )

    user_rows.sort(key=lambda row: row['email'].lower())

    flash(f'Loaded {len(user_rows)} users', 'info')
    return render_template('manage_users.html', users=user_rows)










@admin_bp.route('/settings', methods=['GET', 'POST'])
@admin_required
def site_settings():
    db = get_db()
    settings = _load_site_settings()
    theme_preview_record = get_site_theme_preview()
    theme_versions = list_site_theme_versions(15)
    preview_active = False
    if session.get('theme_preview_active') and theme_preview_record:
        preview_active = True
        settings = _apply_payload_to_settings(settings, theme_preview_record['payload'])
    elif session.get('theme_preview_active'):
        session.pop('theme_preview_active', None)
    theme_template = copy.deepcopy(DEFAULT_THEME_CONFIG)
    theme = copy.deepcopy(settings.get('theme', DEFAULT_THEME_CONFIG))
    palette_defaults = theme_template.get('palette', {})
    typography_defaults = theme_template.get('typography', {})
    icon_defaults = theme_template.get('iconography', {})
    component_defaults = theme_template.get('components', {})

    hex_pattern = re.compile(r'^#[0-9A-Fa-f]{6}$')
    url_schemes = {'http', 'https'}

    def _clean_text(value: str | None, fallback: str = '', max_length: int = 255) -> str:
        if value is None:
            return fallback
        cleaned = value.strip()
        if not cleaned:
            return fallback
        cleaned = cleaned.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        cleaned = ' '.join(cleaned.split())
        if len(cleaned) > max_length:
            cleaned = cleaned[:max_length]
        return cleaned

    def _clean_multiline(value: str | None, fallback: str = '', limit: int = 12000) -> str:
        if value is None:
            return fallback
        normalised = value.replace('\r\n', '\n').replace('\r', '\n').strip()
        if not normalised:
            return ''
        if len(normalised) > limit:
            normalised = normalised[:limit]
        return normalised

    def _clean_email(value: str | None) -> str | None:
        cleaned = _clean_text(value, '', 255)
        if not cleaned or '@' not in cleaned or ' ' in cleaned:
            return None
        return cleaned.lower()

    def _clean_url(value: str | None) -> str | None:
        if not value:
            return None
        candidate = value.strip()
        if not candidate:
            return None
        parsed = urlparse(candidate)
        if parsed.scheme.lower() in url_schemes and parsed.netloc:
            return candidate
        return None

    def _clean_color(value: str | None, fallback: str | None, allow_empty: bool = False) -> str | None:
        if value is None:
            return None if allow_empty else fallback
        candidate = value.strip()
        if not candidate:
            return None if allow_empty else fallback
        candidate = candidate.replace('"', '').replace("'", '')
        if hex_pattern.match(candidate):
            return candidate.lower()
        lowered = candidate.lower()
        if lowered.startswith(('rgb(', 'rgba(')):
            return candidate
        if allow_empty and lowered in {'none', 'transparent'}:
            return None
        return fallback

    def _as_float(value, default: float) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def _clean_float(value, fallback: float, minimum: float, maximum: float) -> float:
        number = _as_float(value, fallback)
        if number < minimum:
            return minimum
        if number > maximum:
            return maximum
        return number

    def _clean_font(value: str | None, fallback: str, max_length: int = 200) -> str:
        cleaned = _clean_text(value, fallback, max_length)
        cleaned = cleaned.replace('{', '').replace('}', '').replace(';', '')
        return cleaned or fallback

    def _normalise_font_stack(value: str | None) -> str:
        if not value:
            return ''
        collapsed = value.replace('"', '').replace("'", '').replace(' ', '')
        return collapsed.lower()

    def _match_font_preset(value: str | None) -> str | None:
        normalised = _normalise_font_stack(value)
        if not normalised:
            return None
        for preset_key, preset_stack in font_preset_map.items():
            if _normalise_font_stack(preset_stack) == normalised:
                return preset_key
        return None

    def _select_font_preset(current_value: str | None, fallback_key: str | None = None) -> tuple[str, str | None]:
        matched = _match_font_preset(current_value)
        if matched:
            return f"preset:{matched}", None
        if fallback_key and fallback_key in font_preset_map:
            return f"preset:{fallback_key}", None
        return 'custom', current_value or ''

    def _resolve_font_choice(choice: str | None, custom_value: str | None, fallback: str) -> str:
        if not choice:
            return fallback
        choice = choice.strip().lower()
        if choice.startswith('preset:'):
            preset_key = choice.split(':', 1)[1]
            return font_preset_map.get(preset_key, fallback)
        if choice == 'custom':
            return _clean_font(custom_value, fallback, 200)
        return fallback

    def _clean_choice(value: str | None, fallback: str, allowed: set[str]) -> str:
        if not value:
            return fallback
        choice = value.strip().lower()
        return choice if choice in allowed else fallback

    feature_flag_fields = [
        'show_hero',
        'show_stats',
        'show_leagues',
        'show_recent_games',
        'show_cta_panel',
        'show_team_logos',
        'show_breadcrumbs',
        'show_footer_social',
        'enable_dark_mode',
        'show_season_filter',
        'show_live_scores',
        'show_venue_details',
        'show_player_stats',
        'show_standings_preview',
        'show_featured_players',
        'show_latest_results',
        'show_highlight_reel',
        'show_partner_logos',
        'show_match_filters',
        'enable_animated_icons',
        'enable_card_glow',
    ]
    social_keys = list(DEFAULT_SOCIAL_LINKS.keys())

    theme_options = {
        'button_shapes': ['rounded', 'soft', 'pill', 'square'],
        'button_styles': ['solid', 'outline', 'gradient', 'ghost'],
        'button_text_transforms': ['none', 'uppercase', 'capitalize'],
        'card_styles': ['flat', 'bordered', 'elevated', 'layered'],
        'card_shadows': ['none', 'light', 'medium', 'heavy'],
        'input_styles': ['soft', 'solid', 'underline', 'outline'],
        'border_radius_scales': ['xs', 'sm', 'md', 'lg', 'xl'],
        'layout_densities': ['compact', 'comfortable', 'spacious'],
        'nav_styles': ['glass', 'solid', 'gradient', 'minimal'],
        'surface_tints': ['subtle', 'solid', 'none', 'frosted'],
        'chip_styles': ['soft', 'outline', 'solid'],
        'icon_weights': ['thin', 'light', 'regular', 'bold', 'fill', 'duotone'],
    }

    font_presets = [
        {
            'key': 'inter',
            'label': 'Inter (Regular)',
            'stack': "Inter, system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'inter-thin',
            'label': 'Inter Thin (200)',
            'stack': "Inter, system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
            'body_weight': '200',
            'heading_weight': '400',
            'group': 'Sans Serif',
        },
        {
            'key': 'inter-bold',
            'label': 'Inter Bold (700)',
            'stack': "Inter, system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
            'body_weight': '600',
            'heading_weight': '700',
            'group': 'Sans Serif',
        },
        {
            'key': 'poppins',
            'label': 'Poppins',
            'stack': "'Poppins', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'poppins-semibold',
            'label': 'Poppins SemiBold',
            'stack': "'Poppins', 'Segoe UI', sans-serif",
            'body_weight': '500',
            'heading_weight': '700',
            'group': 'Sans Serif',
        },
        {
            'key': 'montserrat',
            'label': 'Montserrat',
            'stack': "'Montserrat', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '700',
            'group': 'Sans Serif',
        },
        {
            'key': 'roboto',
            'label': 'Roboto',
            'stack': "'Roboto', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '500',
            'group': 'Sans Serif',
        },
        {
            'key': 'nunito',
            'label': 'Nunito',
            'stack': "'Nunito', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '700',
            'group': 'Sans Serif',
        },
        {
            'key': 'nunito-extra-bold',
            'label': 'Nunito ExtraBold',
            'stack': "'Nunito', 'Segoe UI', sans-serif",
            'body_weight': '500',
            'heading_weight': '800',
            'group': 'Sans Serif',
        },
        {
            'key': 'raleway',
            'label': 'Raleway',
            'stack': "'Raleway', 'Segoe UI', sans-serif",
            'body_weight': '300',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'work-sans',
            'label': 'Work Sans',
            'stack': "'Work Sans', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'work-sans-light',
            'label': 'Work Sans Light',
            'stack': "'Work Sans', 'Segoe UI', sans-serif",
            'body_weight': '300',
            'heading_weight': '500',
            'group': 'Sans Serif',
        },
        {
            'key': 'source-sans',
            'label': 'Source Sans Pro',
            'stack': "'Source Sans Pro', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'ibm-plex-sans',
            'label': 'IBM Plex Sans',
            'stack': "'IBM Plex Sans', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'fira-sans',
            'label': 'Fira Sans',
            'stack': "'Fira Sans', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'space-grotesk',
            'label': 'Space Grotesk',
            'stack': "'Space Grotesk', 'Segoe UI', sans-serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Sans Serif',
        },
        {
            'key': 'roboto-slab',
            'label': 'Roboto Slab',
            'stack': "'Roboto Slab', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Serif & Display',
        },
        {
            'key': 'playfair',
            'label': 'Playfair Display',
            'stack': "'Playfair Display', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '700',
            'group': 'Serif & Display',
        },
        {
            'key': 'merriweather',
            'label': 'Merriweather',
            'stack': "'Merriweather', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '700',
            'group': 'Serif & Display',
        },
        {
            'key': 'lora',
            'label': 'Lora',
            'stack': "'Lora', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Serif & Display',
        },
        {
            'key': 'abril-fatface',
            'label': 'Abril Fatface',
            'stack': "'Abril Fatface', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '700',
            'group': 'Serif & Display',
        },
        {
            'key': 'ibm-plex-serif',
            'label': 'IBM Plex Serif',
            'stack': "'IBM Plex Serif', 'Georgia', serif",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Serif & Display',
        },
        {
            'key': 'space-mono',
            'label': 'Space Mono',
            'stack': "'Space Mono', 'Fira Code', monospace",
            'body_weight': '400',
            'heading_weight': '600',
            'group': 'Monospace',
        },
    ]

    weight_option_map = {
        'inter': [
            {'label': 'Thin', 'value': '200'},
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'poppins': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'montserrat': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'roboto': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'Bold', 'value': '700'},
        ],
        'nunito': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
            {'label': 'ExtraBold', 'value': '800'},
        ],
        'raleway': [
            {'label': 'Thin', 'value': '200'},
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'work-sans': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'source-sans': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'ibm-plex-sans': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'fira-sans': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'space-grotesk': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'roboto-slab': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'playfair': [
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'merriweather': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Bold', 'value': '700'},
        ],
        'lora': [
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'ibm-plex-serif': [
            {'label': 'Light', 'value': '300'},
            {'label': 'Regular', 'value': '400'},
            {'label': 'Medium', 'value': '500'},
            {'label': 'SemiBold', 'value': '600'},
            {'label': 'Bold', 'value': '700'},
        ],
        'space-mono': [
            {'label': 'Regular', 'value': '400'},
            {'label': 'Bold', 'value': '700'},
        ],
    }
    variant_weight_sources = {
        'inter-thin': 'inter',
        'inter-bold': 'inter',
        'poppins-semibold': 'poppins',
        'nunito-extra-bold': 'nunito',
        'work-sans-light': 'work-sans',
    }
    for preset in font_presets:
        source_key = variant_weight_sources.get(preset['key'], preset['key'])
        options = weight_option_map.get(source_key)
        if options:
            preset['weights'] = [dict(option) for option in options]

    font_preset_map = {item['key']: item['stack'] for item in font_presets}
    font_preset_weights = {
        item['key']: {
            'body': item.get('body_weight'),
            'heading': item.get('heading_weight'),
            'options': [dict(option) for option in item.get('weights', [])],
        }
        for item in font_presets
    }
    font_preset_lookup = {item['key']: item for item in font_presets}

    if request.method == 'POST':
        form = request.form

        site_title = _clean_text(form.get('site_title'), 'Sports League Management System', 150)
        brand_image_url = _clean_url(form.get('brand_image_url'))
        favicon_url = _clean_url(form.get('favicon_url'))
        league_tagline = _clean_text(form.get('league_tagline'), '', 180) or None
        contact_email = _clean_email(form.get('contact_email'))

        palette = theme.setdefault('palette', {})
        palette_map = [
            ('primary', 'primary_color', palette_defaults.get('primary', DEFAULT_THEME_CONFIG['palette']['primary']), False),
            ('secondary', 'secondary_color', palette_defaults.get('secondary'), False),
            ('accent', 'accent_color', palette_defaults.get('accent'), False),
            ('neutral', 'neutral_color', palette_defaults.get('neutral'), False),
            ('surface', 'surface_color', palette_defaults.get('surface'), False),
            ('background', 'background_color', palette_defaults.get('background'), False),
            ('text', 'text_color', palette_defaults.get('text'), False),
            ('muted', 'muted_text_color', palette_defaults.get('muted'), False),
            ('heading', 'heading_color', palette_defaults.get('heading'), False),
            ('nav_background', 'nav_background_color', palette_defaults.get('nav_background'), True),
            ('nav_text', 'nav_text_color', palette_defaults.get('nav_text'), False),
            ('nav_hover', 'nav_hover_color', palette_defaults.get('nav_hover'), False),
            ('card_border', 'card_border_color', palette_defaults.get('card_border'), False),
            ('success', 'success_color', palette_defaults.get('success'), False),
            ('warning', 'warning_color', palette_defaults.get('warning'), False),
            ('danger', 'danger_color', palette_defaults.get('danger'), False),
            ('info', 'info_color', palette_defaults.get('info'), False),
            ('gradient_start', 'gradient_start_color', palette_defaults.get('gradient_start'), False),
            ('gradient_end', 'gradient_end_color', palette_defaults.get('gradient_end'), False),
            ('highlight', 'highlight_color', palette_defaults.get('highlight'), False),
            ('shadow', 'shadow_color', palette_defaults.get('shadow'), False),
        ]
        for key, field, default, allow_none in palette_map:
            fallback_value = palette.get(key, default)
            palette[key] = _clean_color(form.get(field), fallback_value, allow_none)
            if palette[key] is None and not allow_none:
                palette[key] = fallback_value

        primary_color = palette.get('primary') or DEFAULT_THEME_CONFIG['palette']['primary']
        palette['primary'] = primary_color

        typography = theme.setdefault('typography', {})
        base_font_choice = form.get('font_family_base_select')
        base_font_custom = form.get('font_family_base_custom')
        default_base = typography.get('base_family', typography_defaults.get('base_family'))
        typography['base_family'] = _resolve_font_choice(base_font_choice, base_font_custom, default_base)

        heading_font_choice = form.get('font_family_heading_select')
        heading_font_custom = form.get('font_family_heading_custom')
        default_heading = typography.get('heading_family', typography_defaults.get('heading_family'))
        typography['heading_family'] = _resolve_font_choice(heading_font_choice, heading_font_custom, default_heading or typography['base_family'])

        typography['base_weight'] = _clean_text(form.get('font_weight_body'), typography.get('base_weight', typography_defaults.get('base_weight')), 4)
        typography['heading_weight'] = _clean_text(form.get('font_weight_heading'), typography.get('heading_weight', typography_defaults.get('heading_weight')), 4)
        typography['letter_spacing'] = _clean_text(form.get('body_letter_spacing'), typography.get('letter_spacing', typography_defaults.get('letter_spacing')), 12)
        typography['heading_letter_spacing'] = _clean_text(form.get('heading_letter_spacing'), typography.get('heading_letter_spacing', typography_defaults.get('heading_letter_spacing')), 12)
        typography['heading_transform'] = _clean_choice(form.get('heading_transform'), typography.get('heading_transform', typography_defaults.get('heading_transform')), {'none', 'uppercase', 'capitalize'})
        base_size_raw = typography.get('base_size', typography_defaults.get('base_size', '16px'))
        base_size_default = _as_float(str(base_size_raw).replace('px', ''), 16.0)
        base_size_value = round(_clean_float(form.get('font_size_base'), base_size_default, 12.0, 24.0))
        typography['base_size'] = f"{base_size_value}px"
        typography['scale'] = round(_clean_float(form.get('font_scale'), _as_float(typography.get('scale', typography_defaults.get('scale', 1.0)), 1.0), 0.6, 1.8), 2)
        typography['line_height'] = round(_clean_float(form.get('line_height'), _as_float(typography.get('line_height', typography_defaults.get('line_height', 1.6)), 1.6), 1.2, 2.0), 2)
        typography['heading_line_height'] = round(_clean_float(form.get('heading_line_height'), _as_float(typography.get('heading_line_height', typography_defaults.get('heading_line_height', 1.3)), 1.3), 1.0, 1.8), 2)

        iconography = theme.setdefault('iconography', {})
        iconography['weight'] = _clean_choice(form.get('icon_weight'), iconography.get('weight', icon_defaults.get('weight', 'regular')), set(theme_options['icon_weights']))
        iconography['size_scale'] = round(_clean_float(form.get('icon_size_scale'), _as_float(iconography.get('size_scale', icon_defaults.get('size_scale', 1.0)), 1.0), 0.5, 2.0), 2)
        iconography['primary_color'] = _clean_color(form.get('icon_primary_color'), iconography.get('primary_color', icon_defaults.get('primary_color')), True)
        iconography['accent_color'] = _clean_color(form.get('icon_accent_color'), iconography.get('accent_color', icon_defaults.get('accent_color')), True)
        iconography['hover_color'] = _clean_color(form.get('icon_hover_color'), iconography.get('hover_color', icon_defaults.get('hover_color')), True)
        iconography['active_color'] = _clean_color(form.get('icon_active_color'), iconography.get('active_color', icon_defaults.get('active_color')), True)

        components = theme.setdefault('components', {})
        components['button_shape'] = _clean_choice(form.get('button_shape'), components.get('button_shape', component_defaults.get('button_shape', 'rounded')), set(theme_options['button_shapes']))
        components['button_style'] = _clean_choice(form.get('button_style'), components.get('button_style', component_defaults.get('button_style', 'solid')), set(theme_options['button_styles']))
        components['button_text_transform'] = _clean_choice(form.get('button_text_transform'), components.get('button_text_transform', component_defaults.get('button_text_transform', 'none')), set(theme_options['button_text_transforms']))
        components['card_style'] = _clean_choice(form.get('card_style'), components.get('card_style', component_defaults.get('card_style', 'elevated')), set(theme_options['card_styles']))
        components['card_shadow'] = _clean_choice(form.get('card_shadow'), components.get('card_shadow', component_defaults.get('card_shadow', 'medium')), set(theme_options['card_shadows']))
        components['input_style'] = _clean_choice(form.get('input_style'), components.get('input_style', component_defaults.get('input_style', 'soft')), set(theme_options['input_styles']))
        components['border_radius_scale'] = _clean_choice(form.get('border_radius_scale'), components.get('border_radius_scale', component_defaults.get('border_radius_scale', 'md')), set(theme_options['border_radius_scales']))
        components['layout_density'] = _clean_choice(form.get('layout_density'), components.get('layout_density', component_defaults.get('layout_density', 'comfortable')), set(theme_options['layout_densities']))
        components['nav_style'] = _clean_choice(form.get('nav_style'), components.get('nav_style', component_defaults.get('nav_style', 'glass')), set(theme_options['nav_styles']))
        components['surface_tint'] = _clean_choice(form.get('surface_tint'), components.get('surface_tint', component_defaults.get('surface_tint', 'subtle')), set(theme_options['surface_tints']))
        components['chip_style'] = _clean_choice(form.get('chip_style'), components.get('chip_style', component_defaults.get('chip_style', 'soft')), set(theme_options['chip_styles']))
        components['card_border'] = _clean_text(form.get('card_border'), components.get('card_border', component_defaults.get('card_border', '1px solid rgba(15, 23, 42, 0.08)')), 120)
        components['navbar_transparency'] = round(_clean_float(form.get('navbar_transparency'), _as_float(components.get('navbar_transparency', component_defaults.get('navbar_transparency', 0.9)), 0.9), 0.0, 1.0), 2)
        components['navbar_blur'] = _clean_text(form.get('navbar_blur'), components.get('navbar_blur', component_defaults.get('navbar_blur', '18px')), 8)
        components['use_gradients'] = bool(form.get('use_gradients'))
        components['section_dividers'] = bool(form.get('section_dividers'))
        components['button_glow'] = bool(form.get('button_glow'))

        social_labels = theme.setdefault('social_labels', {})
        social_labels['custom_1'] = _clean_text(form.get('custom_1_label'), social_labels.get('custom_1', 'Custom Link 1'), 40)
        social_labels['custom_2'] = _clean_text(form.get('custom_2_label'), social_labels.get('custom_2', 'Custom Link 2'), 40)

        theme['custom_css'] = _clean_multiline(form.get('custom_css'), theme.get('custom_css', ''))

        social_links = dict(DEFAULT_SOCIAL_LINKS)
        social_links.update(settings.get('social_links', {}))
        for key in social_keys:
            social_links[key] = _clean_url(form.get(f'{key}_url'))

        feature_flags = dict(DEFAULT_FEATURE_FLAGS)
        feature_flags.update(settings.get('feature_flags', {}))
        for key in feature_flag_fields:
            feature_flags[key] = bool(form.get(key))


        return redirect(url_for('admin.site_settings'))

    theme_for_display = copy.deepcopy(settings.get('theme', DEFAULT_THEME_CONFIG))
    social_labels = theme_for_display.setdefault('social_labels', {})
    social_labels.setdefault('custom_1', 'Custom Link 1')
    social_labels.setdefault('custom_2', 'Custom Link 2')
    typography = theme_for_display.setdefault('typography', {})
    typography_defaults = theme_template.get('typography', {})

    base_font_value = typography.get('base_family') or typography_defaults.get('base_family') or font_presets[0]['stack']
    heading_font_value = typography.get('heading_family') or typography_defaults.get('heading_family') or base_font_value

    default_base_key = _match_font_preset(typography_defaults.get('base_family')) or font_presets[0]['key']
    default_heading_key = _match_font_preset(typography_defaults.get('heading_family')) or default_base_key

    base_font_choice, base_font_custom = _select_font_preset(base_font_value, default_base_key)
    heading_font_choice, heading_font_custom = _select_font_preset(heading_font_value, default_heading_key)

    if base_font_choice == 'custom':
        base_font_custom = base_font_value or base_font_custom or ''
    if heading_font_choice == 'custom':
        heading_font_custom = heading_font_value or heading_font_custom or ''


    settings_for_template = copy.deepcopy(settings)
    settings_for_template['theme'] = theme_for_display
    preview_label_default = theme_preview_record.get('label') if theme_preview_record and theme_preview_record.get('label') else ''
    settings_for_template['social_labels'] = social_labels
    settings_for_template['font_selection'] = {
        'base_choice': base_font_choice,
        'base_custom': base_font_custom or '',
        'heading_choice': heading_font_choice,
        'heading_custom': heading_font_custom or '',
    }

    return render_template(
        'manage_settings.html',
        settings=settings_for_template,
        theme_options=theme_options,
        social_keys=social_keys,
        feature_flag_fields=feature_flag_fields,
        font_presets=font_presets,
        font_preset_weights=font_preset_weights,
        font_preset_lookup=font_preset_lookup,
        theme_preview=theme_preview_record,
        theme_versions=theme_versions,
        theme_preview_active=preview_active,
        preview_label_default=preview_label_default,
    )


@admin_bp.route('/data_management')
@login_required
@admin_required
@tenant_required
def data_management():
    """Data export/import management page."""
    return render_template('data_management.html')


@admin_bp.route('/automation_integrations')
@login_required
@admin_required
@tenant_required
def automation_integrations():
    """Automation and integrations management page."""
    return render_template('automation_integrations.html')


@admin_bp.route('/manage_branding')
@login_required
@admin_required
@tenant_required
def manage_branding():
    """Organization branding and landing page management."""
    from slms.services.domain_loader import (
        get_org_hero_config,
        get_org_modules,
        get_footer_cta
    )

    org = g.org if hasattr(g, 'org') and g.org else None

    hero = get_org_hero_config(org)
    modules = get_org_modules(org)
    footer_cta = get_footer_cta(org)

    return render_template('manage_branding.html',
                         hero=hero,
                         modules=modules,
                         footer_cta=footer_cta)


@admin_bp.route('/api/branding/save', methods=['POST'])
@login_required
@admin_required
@tenant_required
def save_branding():
    """Save branding configuration."""
    try:
        data = request.json
        org = g.org if hasattr(g, 'org') and g.org else None

        if not org:
            return jsonify({'success': False, 'message': 'Organization not found'}), 404

        # Update organization branding
        if 'branding' in data:
            if data['branding'].get('primary'):
                org.primary_color = data['branding']['primary']
            if data['branding'].get('secondary'):
                org.secondary_color = data['branding']['secondary']
            if data['branding'].get('logo'):
                org.logo_url = data['branding']['logo']
            if data['branding'].get('favicon'):
                org.favicon_url = data['branding']['favicon']
            if data['branding'].get('banner'):
                org.banner_image_url = data['branding']['banner']

        # Update custom domain
        if 'custom_domain' in data and data['custom_domain']:
            org.custom_domain = data['custom_domain'].lower().strip()

        # Save hero configuration
        if 'hero' in data:
            org.hero_config = data['hero']

        # Save modules configuration
        if 'modules' in data:
            org.modules_config = data['modules']

        db.session.commit()

        return jsonify({'success': True, 'message': 'Branding saved successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500






